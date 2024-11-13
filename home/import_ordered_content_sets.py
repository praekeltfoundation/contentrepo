import csv
import io
from dataclasses import dataclass
from io import BytesIO
from logging import getLogger
from queue import Queue

from django.core.files.base import File  # type: ignore
from openpyxl import load_workbook

from home.import_helpers import ImportException
from home.models import ContentPage, OrderedContentSet

logger = getLogger(__name__)


@dataclass
class OrderedContentSetPage:
    time: str
    unit: str
    before_or_after: str
    page_slug: str
    contact_field: str


class OrderedContentSetImporter:
    def __init__(
        self,
        file: File,
        file_type: str,
        progress_queue: Queue[int],
    ):
        """
        Initialize an instance of OrderedContentSetImporter.

        :param file: The file to be imported, as a django.core.files.base.File.
        :param file_type: The type of the file, e.g. 'CSV' or 'XLSX'.
        :param progress_queue: A queue.Queue to put progress information on.
        """
        self.file = file
        self.filetype = file_type
        self.progress_queue = progress_queue

    def _get_or_init_ordered_content_set(
        self, row: dict[str, str], set_name: str
    ) -> OrderedContentSet:
        """
        Get or initialize an instance of OrderedContentSet from a row of a CSV file.

        :param row: The row of the CSV file, as a dict.
        :param set_name: The name of the ordered content set.
        :return: An instance of OrderedContentSet.
        """
        ordered_set = OrderedContentSet.objects.filter(name=set_name).first()
        if not ordered_set:
            ordered_set = OrderedContentSet(name=set_name)

        return ordered_set

    def _add_profile_fields(
        self, ordered_set: OrderedContentSet, row: dict[str, str]
    ) -> None:
        """
        Add profile fields to an instance of OrderedContentSet from a row of a CSV file.

        :param ordered_set: An instance of OrderedContentSet.
        :param row: The row of the CSV file, as a dict.
        """
        ordered_set.profile_fields = []
        for field in [f.strip() for f in (row["Profile Fields"] or "").split(",")]:
            if field and field != "-":
                [field_name, field_value] = field.split(":")
                ordered_set.profile_fields.append((field_name, field_value))

    def _extract_ordered_content_set_pages(
        self, row: dict[str, str], index: int
    ) -> list[OrderedContentSetPage]:
        times = self._csv_to_list(row["Time"])
        units = self._csv_to_list(row["Unit"])
        before_or_afters = self._csv_to_list(row["Before Or After"])
        page_slugs = self._csv_to_list(row["Page Slugs"])
        contact_fields = self._csv_to_list(row["Contact Field"])
        # backwards compatiblilty if there's only one contact field
        if len(contact_fields) == 1:
            contact_fields = [contact_fields[0]] * len(times)

        fields = [times, units, before_or_afters, page_slugs, contact_fields]
        if len({len(item) for item in fields}) != 1:
            raise ImportException(
                f"Row {row['Name']} has {len(times)} times, {len(units)} units, {len(before_or_afters)} before_or_afters, {len(page_slugs)} page_slugs and {len(contact_fields)} contact_fields and they should all be equal.",
                index,
            )

        return [
            OrderedContentSetPage(
                time=time,
                unit=unit,
                before_or_after=before_or_after,
                page_slug=page_slug,
                contact_field=contact_field,
            )
            for time, unit, before_or_after, page_slug, contact_field in zip(
                times, units, before_or_afters, page_slugs, contact_fields, strict=False
            )
        ]

    def _csv_to_list(self, column: str) -> list[str]:
        """
        Return a list of strings parsed from a column of a CSV row.

        :param column: The column to parse.
        :return: A list of strings, split from the value of row[column].
        """
        return [p.strip() for p in column.split(",")]

    def _add_pages(
        self,
        ordered_set: OrderedContentSet,
        pages: list[OrderedContentSetPage],
    ) -> None:
        """
        Given the extracted values from a row of the file, create the corresponding ordered content set pages.

        Note that this adds the pages to the OrderedContentSet object, hence there is no return value.

        :param ordered_set: The ordered content set to add the pages to.
        :param pages: A list of OrderedContentSetPage objects.
        """
        ordered_set.pages = []
        for page in pages:
            if page.page_slug and page.page_slug != "-":
                content_page = ContentPage.objects.filter(slug=page.page_slug).first()
                if content_page:
                    os_page = {
                        "contentpage": content_page,
                        "time": page.time or "",
                        "unit": page.unit or "",
                        "before_or_after": page.before_or_after or "",
                        "contact_field": page.contact_field or "",
                    }
                    ordered_set.pages.append(("pages", os_page))
                else:
                    raise ImportException(
                        f"Content page not found for slug '{page.page_slug}'"
                    )

    def _create_ordered_set_from_row(
        self, index: int, row: dict[str, str]
    ) -> OrderedContentSet:
        """
        Create an ordered content set from a single row of the file.

        :param index: The row number of the row being processed.
        :param row: The row of the file as a dictionary.
        :return: An instance of OrderedContentSet.
        :raises ImportException: If time, units, before_or_afters, page_slugs and contact_fields are not all equal length.
        """
        ordered_set = self._get_or_init_ordered_content_set(row, row["Name"])
        self._add_profile_fields(ordered_set, row)

        pages = self._extract_ordered_content_set_pages(row, index)

        self._add_pages(ordered_set, pages)

        ordered_set.save()
        return ordered_set

    def _set_progress(self, progress: int) -> None:
        self.progress_queue.put_nowait(progress)

    def _get_xlsx_rows(self, file: File) -> list[dict[str, str]]:
        """
        Return a list of dictionaries representing the rows in the XLSX file.

        :return: A list of dictionaries representing the rows in the XLSX file.
        """
        lines = []
        wb = load_workbook(filename=BytesIO(file))
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        for row in ws.iter_rows(values_only=True):
            row_dict = {
                "Name": str(row[0]),
                "Profile Fields": str(row[1]),
                "Page Slugs": str(row[2]),
                "Time": str(row[3]),
                "Unit": str(row[4]),
                "Before Or After": str(row[5]),
                "Contact Field": str(row[6]),
            }
            lines.append(row_dict)
        return lines

    def _get_csv_rows(self, file: File) -> list[dict[str, str]]:
        """
        Return a list of dictionaries representing the rows in the CSV file.

        :return: A list of dictionaries representing the rows in the CSV file.
        """
        lines = []
        if isinstance(file, bytes):
            try:
                file = file.decode("utf-8")
            except UnicodeDecodeError:
                file = file.decode("latin-1")

        reader = csv.DictReader(io.StringIO(file))
        for dictionary in reader:
            lines.append(dictionary)

        return lines

    def perform_import(self) -> None:
        """
        Import ordered content sets from a file.

        This method reads the file associated with the importer, processes its
        content based on the specified file type (XLSX or CSV), and creates
        ordered content sets by extracting relevant information from each row.
        The progress of the import operation is reported via a queue.

        Raises:
            ImportException: If any inconsistencies are found in the data
            row while creating ordered content sets.
        """
        file = self.file.read()

        rows = (
            self._get_xlsx_rows(file)
            if self.filetype == "XLSX"
            else self._get_csv_rows(file)
        )

        # 10% progress for loading file
        self._set_progress(10)

        for index, row in enumerate(rows):  # type: ignore
            os = self._create_ordered_set_from_row(index, row)  # type: ignore
            if not os:
                raise ImportException("Ordered Content Set not created", index)
            # 10-100% for loading ordered content sets
            self._set_progress(10 + index * 90 // len(rows))
