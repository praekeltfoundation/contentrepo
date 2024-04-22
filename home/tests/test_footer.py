import csv
import json
import re
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from queue import Queue
from typing import Any

import pytest
from django.core import serializers  # type: ignore
from openpyxl import load_workbook
from pytest_django.fixtures import SettingsWrapper
from wagtail.models import Locale, Page  # type: ignore

from home.content_import_export import import_content, import_ordered_sets
from home.models import (
    ContentPage,
    ContentPageIndex,
)

from .helpers import set_profile_field_options

IMP_EXP_DATA_BASE = Path("home/tests/import-export-data")

ExpDict = dict[str, Any]
ExpPair = tuple[ExpDict, ExpDict]
ExpDicts = Iterable[ExpDict]
ExpDictsPair = tuple[ExpDicts, ExpDicts]


def filter_both(
    filter_func: Callable[[ExpDict], ExpDict]
) -> Callable[[ExpDict, ExpDict], ExpPair]:
    @wraps(filter_func)
    def ff(src: ExpDict, dst: ExpDict) -> ExpPair:
        return filter_func(src), filter_func(dst)

    return ff


@filter_both
def add_new_fields(entry: ExpDict) -> ExpDict:
    # FIXME: This should probably be in a separate test for importing old exports.
    return {
        "whatsapp_template_name": "",
        **entry,
        "whatsapp_template_category": entry.get("whatsapp_template_category")
        or "UTILITY",
        "example_values": entry.get("example_values") or "[]",
        "sms_body": entry.get("sms_body") or "",
        "ussd_body": entry.get("ussd_body") or "",
    }


@filter_both
def ignore_certain_fields(entry: ExpDict) -> ExpDict:
    # FIXME: Do we need page.id to be imported? At the moment nothing in the
    #        import reads that.
    # FIXME: Implement import/export for doc_link, image_link, media_link.
    ignored_fields = {
        "page_id",
        "doc_link",
        "image_link",
        "media_link",
    }
    return {k: v for k, v in entry.items() if k not in ignored_fields}


@filter_both
def strip_leading_whitespace(entry: ExpDict) -> ExpDict:
    # FIXME: Do we expect imported content to have leading spaces removed?
    bodies = {k: v.lstrip(" ") for k, v in entry.items() if k.endswith("_body")}
    return {**entry, **bodies}


EXPORT_FILTER_FUNCS = [
    add_new_fields,
    ignore_certain_fields,
    strip_leading_whitespace,
]


def filter_exports(srcs: ExpDicts, dsts: ExpDicts) -> ExpDictsPair:
    fsrcs, fdsts = [], []
    for src, dst in zip(srcs, dsts, strict=True):
        for ff in EXPORT_FILTER_FUNCS:
            src, dst = ff(src, dst)
        fsrcs.append(src)
        fdsts.append(dst)
    return fsrcs, fdsts


def csv2dicts(csv_bytes: bytes) -> ExpDicts:
    return list(csv.DictReader(StringIO(csv_bytes.decode())))


DbDict = dict[str, Any]
DbDicts = Iterable[DbDict]


def _models2dicts(model_instances: Any) -> DbDicts:
    return json.loads(serializers.serialize("json", model_instances))


def get_page_json() -> DbDicts:
    page_objs = Page.objects.type(ContentPage, ContentPageIndex).all()
    pages = {p["pk"]: p["fields"] for p in _models2dicts(page_objs)}
    content_pages = [
        *_models2dicts(ContentPage.objects.all()),
        *_models2dicts(ContentPageIndex.objects.all()),
    ]
    return [p | {"fields": {**pages[p["pk"]], **p["fields"]}} for p in content_pages]


def per_page(filter_func: Callable[[DbDict], DbDict]) -> Callable[[DbDicts], DbDicts]:
    @wraps(filter_func)
    def fp(pages: DbDicts) -> DbDicts:
        return [filter_func(page) for page in pages]

    return fp


def _is_json_field(field_name: str) -> bool:
    return field_name.endswith("body") or field_name in {"related_pages"}


@per_page
def decode_json_fields(page: DbDict) -> DbDict:
    fields = {
        k: json.loads(v) if _is_json_field(k) else v for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def _normalise_button_pks(body: DbDict, min_pk: int) -> DbDict:
    value = body["value"]
    if "buttons" in value:
        buttons = []
        for button in value["buttons"]:
            if button["type"] == "go_to_page":
                if button.get("value").get("page") is None:
                    continue
                v = button["value"]
                button = button | {"value": v | {"page": v["page"] - min_pk}}
            buttons.append(button)
        value = value | {"buttons": buttons}
    return body | {"value": value}


def _normalise_pks(page: DbDict, min_pk: int) -> DbDict:
    fields = page["fields"]
    if "related_pages" in fields:
        related_pages = [
            rp | {"value": rp["value"] - min_pk} for rp in fields["related_pages"]
        ]
        fields = fields | {"related_pages": related_pages}
    if "whatsapp_body" in fields:
        body = [_normalise_button_pks(b, min_pk) for b in fields["whatsapp_body"]]
        fields = fields | {"whatsapp_body": body}
    return page | {"fields": fields, "pk": page["pk"] - min_pk}


def normalise_pks(pages: DbDicts) -> DbDicts:
    min_pk = min(p["pk"] for p in pages)
    return [_normalise_pks(p, min_pk) for p in pages]


def _update_field(
    pages: DbDicts, field_name: str, update_fn: Callable[[Any], Any]
) -> DbDicts:
    for p in pages:
        fields = p["fields"]
        yield p | {"fields": {**fields, field_name: update_fn(fields[field_name])}}


def normalise_revisions(pages: DbDicts) -> DbDicts:
    if "latest_revision" not in list(pages)[0]["fields"]:
        return pages
    min_latest = min(p["fields"]["latest_revision"] for p in pages)
    min_live = min(p["fields"]["live_revision"] for p in pages)
    pages = _update_field(pages, "latest_revision", lambda v: v - min_latest)
    pages = _update_field(pages, "live_revision", lambda v: v - min_live)
    return pages


def _remove_fields(pages: DbDicts, field_names: set[str]) -> DbDicts:
    for p in pages:
        fields = {k: v for k, v in p["fields"].items() if k not in field_names}
        yield p | {"fields": fields}


PAGE_TIMESTAMP_FIELDS = {
    "first_published_at",
    "last_published_at",
    "latest_revision_created_at",
}


def remove_timestamps(pages: DbDicts) -> DbDicts:
    return _remove_fields(pages, PAGE_TIMESTAMP_FIELDS)


def _normalise_varmsg_ids(page_id: str, var_list: list[dict[str, Any]]) -> None:
    for i, varmsg in enumerate(var_list):
        assert "id" in varmsg
        varmsg["id"] = f"{page_id}:var:{i}"
        for ir, rest in enumerate(varmsg["value"]["variation_restrictions"]):
            rest["id"] = f"{page_id}:var:{i}:var:{ir}"


def _normalise_list_item_ids(page_id: str, var_list: list[dict[str, Any]]) -> None:
    for i, list_item in enumerate(var_list):
        assert "id" in list_item
        list_item["id"] = f"{page_id}:li:{i}"


def _normalise_body_field_ids(
    page: DbDict, body_name: str, body_list: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    for i, body in enumerate(body_list):
        assert "id" in body
        body["id"] = f"fake:{page['pk']}:{body_name}:{i}"
        if "variation_messages" in body["value"]:
            _normalise_varmsg_ids(body["id"], body["value"]["variation_messages"])
        if "list_items" in body["value"]:
            _normalise_list_item_ids(body["id"], body["value"]["list_items"])
    return body_list


@per_page
def normalise_body_ids(page: DbDict) -> DbDict:
    # FIXME: Does it matter if these change?
    fields = {
        k: _normalise_body_field_ids(page, k, v) if k.endswith("body") else v
        for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def _normalise_related_page_ids(
    page: DbDict, rp_list: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    for i, rp in enumerate(rp_list):
        assert "id" in rp
        rp["id"] = f"fake:{page['pk']}:related_page:{i}"
    return rp_list


@per_page
def normalise_related_page_ids(page: DbDict) -> DbDict:
    # FIXME: Does it matter if these change?
    fields = {
        k: _normalise_related_page_ids(page, v) if k == "related_pages" else v
        for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


@per_page
def null_to_emptystr(page: DbDict) -> DbDict:
    # FIXME: Confirm that there's no meaningful difference here, potentially
    #        make these fields non-nullable.
    fields = {**page["fields"]}
    for k in [
        "subtitle",
        "whatsapp_title",
        "messenger_title",
        "viber_title",
    ]:
        if k in fields and fields[k] is None:
            fields[k] = ""
    if "whatsapp_body" in fields:
        for body in fields["whatsapp_body"]:
            if "next_prompt" in body["value"] and not body["value"]["next_prompt"]:
                body["value"]["next_prompt"] = ""
    return page | {"fields": fields}


WEB_PARA_RE = re.compile(r'^<div class="block-paragraph">(.*)</div>$')


@per_page
def clean_web_paragraphs(page: DbDict) -> DbDict:
    # FIXME: Handle this at export time, I guess.
    if "body" in page["fields"]:
        body = [
            b | {"value": WEB_PARA_RE.sub(r"\1", b["value"])}
            for b in page["fields"]["body"]
        ]
        page = page | {"fields": page["fields"] | {"body": body}}
    return page


@per_page
def remove_button_ids(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            buttons = body["value"].get("buttons", [])
            for button in buttons:
                button.pop("id", None)
    return page


@per_page
def remove_example_value_ids(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            example_values = body["value"].get("example_values", [])
            for example_value in example_values:
                example_value.pop("id", None)
    return page


PAGE_FILTER_FUNCS = [
    normalise_pks,
    normalise_revisions,
    remove_timestamps,
    normalise_body_ids,
    normalise_related_page_ids,
    clean_web_paragraphs,
    null_to_emptystr,
    remove_button_ids,
    remove_example_value_ids,
]


@dataclass
class ImportExport:
    admin_client: Any
    format: str

    @property
    def _filter_export(self) -> Callable[..., bytes]:
        return {
            "csv": self._filter_export_CSV,
            "xlsx": self._filter_export_XLSX,
        }[self.format]

    def _filter_export_row(self, row: ExpDict, locale: str | None) -> bool:
        """
        Determine whether to keep a given export row.
        """
        if locale:
            if row["locale"] not in [None, "", locale]:
                return False
        return True

    def _filter_export_CSV(self, content: bytes, locale: str | None) -> bytes:
        reader = csv.DictReader(StringIO(content.decode()))
        assert reader.fieldnames is not None
        out_content = StringIO()
        writer = csv.DictWriter(out_content, fieldnames=reader.fieldnames)
        writer.writeheader()
        for row in reader:
            if self._filter_export_row(row, locale=locale):
                writer.writerow(row)
        return out_content.getvalue().encode()

    def _filter_export_XLSX(self, content: bytes, locale: str | None) -> bytes:
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.worksheets[0]
        header = next(worksheet.iter_rows(max_row=1, values_only=True))

        rows_to_remove: list[int] = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {k: v for k, v in zip(header, row, strict=True) if isinstance(k, str)}
            if not self._filter_export_row(r, locale=locale):
                rows_to_remove.append(i + 2)
        for row_num in reversed(rows_to_remove):
            worksheet.delete_rows(row_num)

        out_content = BytesIO()
        workbook.save(out_content)
        return out_content.getvalue()

    def export_content(self, locale: str | None = None) -> bytes:
        """
        Export all (or filtered) content in the configured format.

        FIXME:
         * If we filter the export by locale, we only get ContentPage entries
           for the given language, but we still get ContentPageIndex rows for
           all languages.
        """
        url = f"/admin/home/contentpage/?export={self.format}"
        if locale:
            loc = Locale.objects.get(language_code=locale)
            locale = str(loc)
            url = f"{url}&locale__id__exact={loc.id}"
        content = self.admin_client.get(url).content
        # Hopefully we can get rid of this at some point.
        if locale:
            content = self._filter_export(content, locale=locale)
        if self.format == "csv":
            print("-v-CONTENT-v-")
            print(content.decode())
            print("-^-CONTENT-^-")
        return content

    def import_content(self, content_bytes: bytes, **kw: Any) -> None:
        """
        Import given content in the configured format with the configured importer.
        """
        import_content(BytesIO(content_bytes), self.format.upper(), Queue(), **kw)

    def import_ordered_sets(self, content_bytes: bytes, purge: bool = False) -> None:
        import_ordered_sets(BytesIO(content_bytes), self.format.upper(), Queue(), purge)

    def read_bytes(self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE) -> bytes:
        return (path_base / path_str).read_bytes()

    def import_file(
        self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE, **kw: Any
    ) -> bytes:
        """
        Import given content file in the configured format with the configured importer.
        """
        content = self.read_bytes(path_str, path_base)
        self.import_content(content, **kw)
        return content

    def export_reimport(self) -> None:
        """
        Export all content, then immediately reimport it.
        """
        self.import_content(self.export_content())

    def get_page_json(self, locale: str | None = None) -> DbDicts:
        """
        Serialize all ContentPage and ContentPageIndex instances and normalize
        things that vary across import/export.
        """
        pages = decode_json_fields(get_page_json())
        for ff in PAGE_FILTER_FUNCS:
            pages = ff(pages)
        if locale is not None:
            loc = Locale.objects.get(language_code=locale)
            pages = [p for p in pages if p["fields"]["locale"] == loc.id]
        return sorted(pages, key=lambda p: p["pk"])

    def csvs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = csv2dicts(src_bytes)
        dst = csv2dicts(dst_bytes)
        return filter_exports(src, dst)


@pytest.fixture()
def csv_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "csv")


@pytest.fixture()
def tmp_media_path(tmp_path: Path, settings: SettingsWrapper) -> None:
    settings.MEDIA_ROOT = tmp_path


@pytest.fixture()
def xlsx_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "xlsx")


@pytest.mark.usefixtures("tmp_media_path")
@pytest.mark.django_db
class TestExportImportRoundtrip:

    def test_list_items_values_with_comma(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file containing list items that has a comma
        page and then exporting it produces a duplicate of the original file.

        (This uses list_items_with_comma.csv.)
        """
        set_profile_field_options()
        csv_bytes = csv_impexp.import_file("list_items_with_comma.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src


# @pytest.mark.django_db
# class TestImportExport:
#     """
#     Text various import and export scenarios that aren't specifically round
#     trips.

#     NOTE: This is not a Django (or even unittest) TestCase. It's just a
#         container for related tests.
#     """

#     def test_import_pages_xlsx(self, xlsx_impexp: ImportExport) -> None:
#         """
#         Importing an XLSX file with content pages should not break
#         """
#         xlsx_impexp.import_file("content_pages.xlsx", purge=False)
#         content_pages = ContentPage.objects.all()
#         assert len(content_pages) > 0
