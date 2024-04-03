import csv
import json
import re
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from queue import Queue
from typing import Any

import pytest, sqlite3
from django.core import serializers  # type: ignore
from django.core.files.base import File  # type: ignore
from django.core.files.images import ImageFile  # type: ignore
from openpyxl import load_workbook
from pytest_django.fixtures import SettingsWrapper
from wagtail.documents.models import Document  # type: ignore
from wagtail.images.models import Image  # type: ignore
from wagtail.models import Locale, Page  # type: ignore
from wagtailmedia.models import Media  # type: ignore

from home.assessment_import_export import import_assessment
from home.import_assessment_pages import ImportAssessmentException
from home.models import (
    ContentPage,
    ContentPageIndex,
    HomePage,
    Assessment,
)
from home.tests.utils import unwagtail

from .helpers import set_profile_field_options
from .page_builder import (
    MBlk,
    MBody,
    NextBtn,
    PageBtn,
    PageBuilder,
    SBlk,
    SBody,
    UBlk,
    UBody,
    VarMsg,
    VBlk,
    VBody,
    WABlk,
    WABody,
)

IMP_EXP_DATA_BASE = Path("home/tests/import-export-data")

ExpDict = dict[str, Any]
ExpPair = tuple[ExpDict, ExpDict]
ExpDicts = Iterable[ExpDict]
ExpDictsPair = tuple[ExpDicts, ExpDicts]


def filter_both(
    filter_func: Callable[[ExpDict], ExpDict]
) -> Callable[[ExpDict, ExpDict], ExpPair]:
    @wraps(filter_func)
    def ff(src: ExpDict, dst: ExpDict) -> ExpPair:
        return filter_func(src), filter_func(dst)

    return ff


@filter_both
def ignore_certain_fields(entry: ExpDict) -> ExpDict:
    # FIXME: Do we need page.id to be imported? At the moment nothing in the
    #        import reads that.
    # FIXME: Implement import/export for doc_link, image_link, media_link.
    ignored_fields = {
        "page_id",
        "doc_link",
        "image_link",
        "media_link",
    }
    return {k: v for k, v in entry.items() if k not in ignored_fields}


@filter_both
def strip_leading_whitespace(entry: ExpDict) -> ExpDict:
    # FIXME: Do we expect imported content to have leading spaces removed?
    bodies = {k: v.lstrip(" ") for k, v in entry.items() if k.endswith("_body")}
    return {**entry, **bodies}


EXPORT_FILTER_FUNCS = [
    # add_new_fields,
    ignore_certain_fields,
    strip_leading_whitespace,
]


def filter_exports(srcs: ExpDicts, dsts: ExpDicts) -> ExpDictsPair:
    fsrcs, fdsts = [], []
    for src, dst in zip(srcs, dsts, strict=True):
        for ff in EXPORT_FILTER_FUNCS:
            src, dst = ff(src, dst)
        fsrcs.append(src)
        fdsts.append(dst)
    return fsrcs, fdsts


def csv2dicts(csv_bytes: bytes) -> ExpDicts:
    return list(csv.DictReader(StringIO(csv_bytes.decode())))


WEB_PARA_RE = re.compile(r'^<div class="block-paragraph">(.*)</div>$')


@dataclass
class ImportExport:
    admin_client: Any
    format: str

    @property
    def _filter_export(self) -> Callable[..., bytes]:
        return {
            "csv": self._filter_export_CSV,
            "xlsx": self._filter_export_XLSX,
        }[self.format]

    def _filter_export_row(self, row: ExpDict, locale: str | None) -> bool:
        """
        Determine whether to keep a given export row.
        """
        if locale:
            if row["locale"] not in [None, "", locale]:
                return False
        return True

    def _filter_export_CSV(self, content: bytes, locale: str | None) -> bytes:
        reader = csv.DictReader(StringIO(content.decode()))
        assert reader.fieldnames is not None
        out_content = StringIO()
        writer = csv.DictWriter(out_content, fieldnames=reader.fieldnames)
        writer.writeheader()
        for row in reader:
            if self._filter_export_row(row, locale=locale):
                writer.writerow(row)
        return out_content.getvalue().encode()

    def _filter_export_XLSX(self, content: bytes, locale: str | None) -> bytes:
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.worksheets[0]
        header = next(worksheet.iter_rows(max_row=1, values_only=True))

        rows_to_remove: list[int] = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {k: v for k, v in zip(header, row, strict=True) if isinstance(k, str)}
            if not self._filter_export_row(r, locale=locale):
                rows_to_remove.append(i + 2)
        for row_num in reversed(rows_to_remove):
            worksheet.delete_rows(row_num)

        out_content = BytesIO()
        workbook.save(out_content)
        return out_content.getvalue()

    def export_assessment(self, locale: str | None = None) -> bytes:
        """
        Export all (or filtered) content in the configured format.

        FIXME:
         * If we filter the export by locale, we only get ContentPage entries
           for the given language, but we still get ContentPageIndex rows for
           all languages.
        """
        url = f"/admin/snippets/home/assessment/?export={self.format}"
        if locale:
            loc = Locale.objects.get(language_code=locale)
            locale = str(loc)
            url = f"{url}&locale__id__exact={loc.id}"
        content = self.admin_client.get(url).content
        # Hopefully we can get rid of this at some point.
        if locale:
            content = self._filter_export(content, locale=locale)
        if self.format == "csv":
            print("-v-CONTENT-v-")
            print(content.decode())
            print("-^-CONTENT-^-")
        return content

    def import_assessment(self, content_bytes: bytes, **kw: Any) -> None:
        """
        Import given content in the configured format with the configured importer.
        """
        import_assessment(BytesIO(content_bytes), self.format.upper(), Queue(), **kw)

    def read_bytes(self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE) -> bytes:
        return (path_base / path_str).read_bytes()

    def import_file(
        self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE, **kw: Any
    ) -> bytes:
        """
        Import given content file in the configured format with the configured importer.
        """
        content = self.read_bytes(path_str, path_base)
        print("Import method content is")
        print(content)
        self.import_assessment(content, **kw)
        print(content)
        return content

    def export_reimport(self) -> None:
        """
        Export all content, then immediately reimport it.
        """
        self.import_assessment(self.export_assessment())

    def csvs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = csv2dicts(src_bytes)
        dst = csv2dicts(dst_bytes)
        return filter_exports(src, dst)


@pytest.fixture(params=["csv", "xlsx"])
def impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, request.param)


@pytest.fixture()
def tmp_media_path(tmp_path: Path, settings: SettingsWrapper) -> None:
    settings.MEDIA_ROOT = tmp_path


@pytest.fixture()
def csv_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "csv")


@pytest.fixture()
def xlsx_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "xlsx")


@pytest.mark.django_db()
class TestImportExportRoundtrip:
    """
    Test importing and reexporting content produces an export that is
    equilavent to the original imported content.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_simple(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with one assessment and
        one question and export it

        (This uses assessment_simple.csv.)

        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        high_inflection = PageBuilder.build_cp(
            parent=main_menu,
            slug="high-inflection",
            title="High Inflection",
            bodies=[
                WABody("High Inflection", [WABlk("*High Inflection Page")]),
                MBody("High inflection", [MBlk("High Inflection Page")]),
            ],
        )
        medium_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="medium-score",
            title="Medium Score",
            bodies=[
                WABody("Medium Score", [WABlk("*Medium Inflection Page")]),
                MBody("Medium Score", [MBlk("Medium Inflection Page")]),
            ],
        )

        low_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="low-score",
            title="Low Score",
            bodies=[
                WABody("Low Score", [WABlk("*Low Inflection Page")]),
                MBody("Low Score", [MBlk("Low Inflection Page")]),
            ],
        )

        csv_bytes = csv_impexp.import_file("assessment_simple.csv")
        content = csv_impexp.export_assessment()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_less_simple_multiple_questions(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with two assessments and
        multiple questions and export it

        (This uses assessment_less_simple.csv.)

        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        high_inflection = PageBuilder.build_cp(
            parent=main_menu,
            slug="high-inflection",
            title="High Inflection",
            bodies=[
                WABody("High Inflection", [WABlk("*High Inflection Page")]),
                MBody("High inflection", [MBlk("High Inflection Page")]),
            ],
        )
        medium_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="medium-score",
            title="Medium Score",
            bodies=[
                WABody("Medium Score", [WABlk("*Medium Inflection Page")]),
                MBody("Medium Score", [MBlk("Medium Inflection Page")]),
            ],
        )

        low_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="low-score",
            title="Low Score",
            bodies=[
                WABody("Low Score", [WABlk("*Low Inflection Page")]),
                MBody("Low Score", [MBlk("Low Inflection Page")]),
            ],
        )

        csv_bytes = csv_impexp.import_file("assessment_less_simple.csv")
        content = csv_impexp.export_assessment()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_multiple_assessments(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with more than one assessment and
        multiple questions and export it

        (This uses multiple_assessments.csv.)

        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        high_inflection = PageBuilder.build_cp(
            parent=main_menu,
            slug="high-inflection",
            title="High Inflection",
            bodies=[
                WABody("High Inflection", [WABlk("*High Inflection Page")]),
                MBody("High inflection", [MBlk("High Inflection Page")]),
            ],
        )
        medium_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="medium-score",
            title="Medium Score",
            bodies=[
                WABody("Medium Score", [WABlk("*Medium Inflection Page")]),
                MBody("Medium Score", [MBlk("Medium Inflection Page")]),
            ],
        )

        low_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="low-score",
            title="Low Score",
            bodies=[
                WABody("Low Score", [WABlk("*Low Inflection Page")]),
                MBody("Low Score", [MBlk("Low Inflection Page")]),
            ],
        )

        csv_bytes = csv_impexp.import_file("multiple_assessments.csv")
        content = csv_impexp.export_assessment()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_bulk_assessments(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with more than five assessments and
        multiple questions and export it

        (This uses bulk_assessments.csv.)

        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        high_inflection = PageBuilder.build_cp(
            parent=main_menu,
            slug="high-inflection",
            title="High Inflection",
            bodies=[
                WABody("High Inflection", [WABlk("*High Inflection Page")]),
                MBody("High inflection", [MBlk("High Inflection Page")]),
            ],
        )
        medium_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="medium-score",
            title="Medium Score",
            bodies=[
                WABody("Medium Score", [WABlk("*Medium Inflection Page")]),
                MBody("Medium Score", [MBlk("Medium Inflection Page")]),
            ],
        )

        low_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="low-score",
            title="Low Score",
            bodies=[
                WABody("Low Score", [WABlk("*Low Inflection Page")]),
                MBody("Low Score", [MBlk("Low Inflection Page")]),
            ],
        )

        csv_bytes = csv_impexp.import_file("bulk_assessments.csv")
        content = csv_impexp.export_assessment()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src


@pytest.mark.django_db()
class TestImportExport:
    """
    Test import and export scenarios that aren't specifically round
    trips.
    """

    def test_missing_related_pages(self, csv_impexp: ImportExport) -> None:
        """
        Related pages are listed as comma separated slugs in imported files. If there
        is a slug listed that we cannot find the page for, then we should show the
        user an error with information about the missing page.
        """

        with pytest.raises(ImportAssessmentException) as e:
            csv_impexp.import_file("assessment_missing_related_page.csv")
        assert (
            e.value.message
            == "You are trying to add an assessment with slug fake-page that does not exist. Please create the fake-page page first."
        )

    def test_import_error(elf, csv_impexp: ImportExport) -> None:
        """
        Importing an ivalid CSV file leaves the db as-is.
        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        high_inflection = PageBuilder.build_cp(
            parent=main_menu,
            slug="high-inflection",
            title="High Inflection",
            bodies=[
                WABody("High Inflection", [WABlk("*High Inflection Page")]),
                MBody("High inflection", [MBlk("High Inflection Page")]),
            ],
        )
        medium_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="medium-score",
            title="Medium Score",
            bodies=[
                WABody("Medium Score", [WABlk("*Medium Inflection Page")]),
                MBody("Medium Score", [MBlk("Medium Inflection Page")]),
            ],
        )

        low_score = PageBuilder.build_cp(
            parent=main_menu,
            slug="low-score",
            title="Low Score",
            bodies=[
                WABody("Low Score", [WABlk("*Low Inflection Page")]),
                MBody("Low Score", [MBlk("Low Inflection Page")]),
            ],
        )
        csv_bytes = csv_impexp.import_file("assessment_simple.csv")
        with pytest.raises(ImportAssessmentException) as e:
            csv_impexp.import_file("broken_assessment.csv")
        assert e.value.message == "The import file is missing some required fields."
        content = csv_impexp.export_assessment()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src
