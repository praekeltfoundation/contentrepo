import csv
import itertools
import json
import re
from collections.abc import Callable, Iterable, Iterator
from dataclasses import dataclass
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from queue import Queue
from typing import Any

import pytest
from django.core import serializers  # type: ignore
from django.core.files.base import File  # type: ignore
from django.core.files.images import ImageFile  # type: ignore
from openpyxl import load_workbook
from pytest_django.fixtures import SettingsWrapper
from wagtail.documents.models import Document  # type: ignore
from wagtail.images.models import Image  # type: ignore
from wagtail.models import Locale, Page  # type: ignore
from wagtailmedia.models import Media  # type: ignore

from home.content_import_export import import_content, import_ordered_sets
from home.import_helpers import ImportException
from home.models import (
    Assessment,
    ContentPage,
    ContentPageIndex,
    GoToFormButton,
    GoToFormListItem,
    GoToPageButton,
    HomePage,
    OrderedContentSet,
)
from home.xlsx_helpers import get_active_sheet

from .helpers import set_profile_field_options
from .page_builder import (
    FormBtn,
    FormListItem,
    MBlk,
    MBody,
    NextBtn,
    NextListItem,
    PageBtn,
    PageBuilder,
    PageListItem,
    SBlk,
    SBody,
    UBlk,
    UBody,
    VarMsg,
    VBlk,
    VBody,
    WABlk,
    WABody,
)
from .utils import unwagtail

IMP_EXP_DATA_BASE = Path("home/tests/import-export-data")

ExpDict = dict[str, Any]
ExpPair = tuple[ExpDict, ExpDict]
ExpDicts = Iterable[ExpDict]
ExpDictsPair = tuple[ExpDicts, ExpDicts]


def filter_both(
    filter_func: Callable[[ExpDict], ExpDict],
) -> Callable[[ExpDict, ExpDict], ExpPair]:
    @wraps(filter_func)
    def ff(src: ExpDict, dst: ExpDict) -> ExpPair:
        return filter_func(src), filter_func(dst)

    return ff


@filter_both
def add_new_fields(entry: ExpDict) -> ExpDict:
    # FIXME: This should probably be in a separate test for importing old exports.
    return {
        "whatsapp_template_name": "",
        **entry,
        "whatsapp_template_category": entry.get("whatsapp_template_category")
        or "UTILITY",
        "example_values": entry.get("example_values") or "[]",
        "sms_body": entry.get("sms_body") or "",
        "ussd_body": entry.get("ussd_body") or "",
    }


@filter_both
def ignore_certain_fields(entry: ExpDict) -> ExpDict:
    # FIXME: Do we need page.id to be imported? At the moment nothing in the
    #        import reads that.
    # FIXME: Implement import/export for doc_link, image_link, media_link.
    ignored_fields = {
        "page_id",
        "doc_link",
        "image_link",
        "media_link",
    }
    return {k: v for k, v in entry.items() if k not in ignored_fields}


@filter_both
def strip_leading_whitespace(entry: ExpDict) -> ExpDict:
    # FIXME: Do we expect imported content to have leading spaces removed?
    bodies = {k: v.lstrip(" ") for k, v in entry.items() if k.endswith("_body")}
    return {**entry, **bodies}


EXPORT_FILTER_FUNCS = [
    add_new_fields,
    ignore_certain_fields,
    strip_leading_whitespace,
]


def filter_exports(srcs: ExpDicts, dsts: ExpDicts) -> ExpDictsPair:
    fsrcs, fdsts = [], []
    for src, dst in zip(srcs, dsts, strict=True):
        for ff in EXPORT_FILTER_FUNCS:
            src, dst = ff(src, dst)
        fsrcs.append(src)
        fdsts.append(dst)
    return fsrcs, fdsts


def lower_first(iterator: Iterator[Any]) -> Iterable[Any]:
    return itertools.chain([next(iterator).lower()], iterator)


def csv2dicts(csv_bytes: bytes) -> ExpDicts:
    return list(csv.DictReader(lower_first(StringIO(csv_bytes.decode()))))


def xlsx2dicts(xlsx_bytes: bytes) -> ExpDicts:
    workbook = load_workbook(BytesIO(xlsx_bytes))
    worksheet = get_active_sheet(workbook)
    header = next(worksheet.iter_rows(max_row=1, values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in header]

    for row in worksheet.iter_rows(values_only=True):  # type: ignore
        yield {headers[i]: row[i] for i in range(len(headers))}


DbDict = dict[str, Any]
DbDicts = Iterable[DbDict]


def _models2dicts(model_instances: Any) -> DbDicts:
    return json.loads(serializers.serialize("json", model_instances))


def get_page_json() -> DbDicts:
    page_objs = Page.objects.type(
        ContentPage, ContentPageIndex, OrderedContentSet
    ).all()
    pages = {p["pk"]: p["fields"] for p in _models2dicts(page_objs)}
    content_pages = [
        *_models2dicts(ContentPage.objects.all()),
        *_models2dicts(ContentPageIndex.objects.all()),
        *_models2dicts(OrderedContentSet.objects.all()),
    ]
    return [p | {"fields": {**pages[p["pk"]], **p["fields"]}} for p in content_pages]


def per_page(filter_func: Callable[[DbDict], DbDict]) -> Callable[[DbDicts], DbDicts]:
    @wraps(filter_func)
    def fp(pages: DbDicts) -> DbDicts:
        return [filter_func(page) for page in pages]

    return fp


def _is_json_field(field_name: str) -> bool:
    return field_name.endswith("body") or field_name in {"related_pages"}


@per_page
def decode_json_fields(page: DbDict) -> DbDict:
    fields = {
        k: json.loads(v) if _is_json_field(k) else v for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def _normalise_button_pks(body: DbDict, min_pk: int) -> DbDict:
    value = body["value"]
    if "buttons" in value:
        buttons = []
        for button in value["buttons"]:
            if button["type"] == "go_to_page":
                if button.get("value").get("page") is None:
                    continue
                v = button["value"]
                button = button | {"value": v | {"page": v["page"] - min_pk}}
            buttons.append(button)
        value = value | {"buttons": buttons}
    return body | {"value": value}


def _normalise_list_pks(body: DbDict, min_pk: int) -> DbDict:
    value = body["value"]
    if "list_items" in value:
        list_items = []
        for list_item in value["list_items"]:
            if list_item["type"] == "go_to_page":
                if list_item.get("value").get("page") is None:
                    continue
                v = list_item["value"]
                list_item = list_item | {"value": v | {"page": v["page"] - min_pk}}
            list_items.append(list_item)
        value = value | {"list_items": list_items}
    return body | {"value": value}


def _normalise_pks(page: DbDict, min_pk: int) -> DbDict:
    fields = page["fields"]
    if "related_pages" in fields:
        related_pages = [
            rp | {"value": rp["value"] - min_pk} for rp in fields["related_pages"]
        ]
        fields = fields | {"related_pages": related_pages}
    if "whatsapp_body" in fields:
        body = [_normalise_button_pks(b, min_pk) for b in fields["whatsapp_body"]]
        body = [_normalise_list_pks(b, min_pk) for b in body]
        fields = fields | {"whatsapp_body": body}
    return page | {"fields": fields, "pk": page["pk"] - min_pk}


def normalise_pks(pages: DbDicts) -> DbDicts:
    min_pk = min(p["pk"] for p in pages)
    return [_normalise_pks(p, min_pk) for p in pages]


def _update_field(
    pages: DbDicts, field_name: str, update_fn: Callable[[Any], Any]
) -> DbDicts:
    for p in pages:
        fields = p["fields"]
        yield p | {"fields": {**fields, field_name: update_fn(fields[field_name])}}


def normalise_revisions(pages: DbDicts) -> DbDicts:
    if "latest_revision" not in list(pages)[0]["fields"]:
        return pages
    min_latest = min(p["fields"]["latest_revision"] for p in pages)
    min_live = min(p["fields"]["live_revision"] for p in pages)
    pages = _update_field(pages, "latest_revision", lambda v: v - min_latest)
    pages = _update_field(pages, "live_revision", lambda v: v - min_live)
    return pages


def _remove_fields(pages: DbDicts, field_names: set[str]) -> DbDicts:
    for p in pages:
        fields = {k: v for k, v in p["fields"].items() if k not in field_names}
        yield p | {"fields": fields}


PAGE_TIMESTAMP_FIELDS = {
    "first_published_at",
    "last_published_at",
    "latest_revision_created_at",
}


def remove_timestamps(pages: DbDicts) -> DbDicts:
    return _remove_fields(pages, PAGE_TIMESTAMP_FIELDS)


def _normalise_varmsg_ids(page_id: str, var_list: list[dict[str, Any]]) -> None:
    for i, varmsg in enumerate(var_list):
        assert "id" in varmsg
        varmsg["id"] = f"{page_id}:var:{i}"
        for ir, rest in enumerate(varmsg["value"]["variation_restrictions"]):
            rest["id"] = f"{page_id}:var:{i}:var:{ir}"


def _normalise_list_item_ids(page_id: str, var_list: list[dict[str, Any]]) -> None:
    for i, list_item in enumerate(var_list):
        assert "id" in list_item
        list_item["id"] = f"{page_id}:li:{i}"


def _normalise_body_field_ids(
    page: DbDict, body_name: str, body_list: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    for i, body in enumerate(body_list):
        assert "id" in body
        body["id"] = f"fake:{page['pk']}:{body_name}:{i}"
        if "variation_messages" in body["value"]:
            _normalise_varmsg_ids(body["id"], body["value"]["variation_messages"])
        if "list_items" in body["value"]:
            _normalise_list_item_ids(body["id"], body["value"]["list_items"])
    return body_list


@per_page
def normalise_body_ids(page: DbDict) -> DbDict:
    # FIXME: Does it matter if these change?
    fields = {
        k: _normalise_body_field_ids(page, k, v) if k.endswith("body") else v
        for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def _normalise_related_page_ids(
    page: DbDict, rp_list: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    for i, rp in enumerate(rp_list):
        assert "id" in rp
        rp["id"] = f"fake:{page['pk']}:related_page:{i}"
    return rp_list


@per_page
def normalise_related_page_ids(page: DbDict) -> DbDict:
    # FIXME: Does it matter if these change?
    fields = {
        k: _normalise_related_page_ids(page, v) if k == "related_pages" else v
        for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


@per_page
def null_to_emptystr(page: DbDict) -> DbDict:
    # FIXME: Confirm that there's no meaningful difference here, potentially
    #        make these fields non-nullable.
    fields = {**page["fields"]}
    for k in [
        "subtitle",
        "whatsapp_title",
        "messenger_title",
        "viber_title",
    ]:
        if k in fields and fields[k] is None:
            fields[k] = ""
    if "whatsapp_body" in fields:
        for body in fields["whatsapp_body"]:
            if "next_prompt" in body["value"] and not body["value"]["next_prompt"]:
                body["value"]["next_prompt"] = ""
    return page | {"fields": fields}


WEB_PARA_RE = re.compile(r'^<div class="block-paragraph">(.*)</div>$')


@per_page
def clean_web_paragraphs(page: DbDict) -> DbDict:
    # FIXME: Handle this at export time, I guess.
    if "body" in page["fields"]:
        body = [
            b | {"value": WEB_PARA_RE.sub(r"\1", b["value"])}
            for b in page["fields"]["body"]
        ]
        page = page | {"fields": page["fields"] | {"body": body}}
    return page


@per_page
def remove_button_ids(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            buttons = body["value"].get("buttons", [])
            for button in buttons:
                button.pop("id", None)
    return page


@per_page
def remove_example_value_ids(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            example_values = body["value"].get("example_values", [])
            for example_value in example_values:
                example_value.pop("id", None)
    return page


PAGE_FILTER_FUNCS = [
    normalise_pks,
    normalise_revisions,
    remove_timestamps,
    normalise_body_ids,
    normalise_related_page_ids,
    clean_web_paragraphs,
    null_to_emptystr,
    remove_button_ids,
    remove_example_value_ids,
]


@dataclass
class ImportExport:
    admin_client: Any
    format: str

    @property
    def _filter_export(self) -> Callable[..., bytes]:
        return {
            "csv": self._filter_export_CSV,
            "xlsx": self._filter_export_XLSX,
        }[self.format]

    def _filter_export_row(self, row: ExpDict, locale: str | None) -> bool:
        """
        Determine whether to keep a given export row.
        """
        if locale:
            if row.get("locale") not in [None, "", locale]:
                return False
        return True

    def _filter_export_CSV(self, content: bytes, locale: str | None) -> bytes:
        reader = csv.DictReader(StringIO(content.decode()))
        assert reader.fieldnames is not None
        out_content = StringIO()
        writer = csv.DictWriter(out_content, fieldnames=reader.fieldnames)
        writer.writeheader()
        for row in reader:
            if self._filter_export_row(row, locale=locale):
                writer.writerow(row)
        return out_content.getvalue().encode()

    def _filter_export_XLSX(self, content: bytes, locale: str | None) -> bytes:
        workbook = load_workbook(BytesIO(content))
        worksheet = get_active_sheet(workbook)
        header = next(worksheet.iter_rows(max_row=1, values_only=True))

        rows_to_remove: list[int] = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {k: v for k, v in zip(header, row, strict=True) if isinstance(k, str)}
            if not self._filter_export_row(r, locale=locale):
                rows_to_remove.append(i + 2)
        for row_num in reversed(rows_to_remove):
            worksheet.delete_rows(row_num)

        out_content = BytesIO()
        workbook.save(out_content)
        return out_content.getvalue()

    def export_content(self, locale: str | None = None) -> bytes:
        """
        Export all (or filtered) content in the configured format.

        FIXME:
         * If we filter the export by locale, we only get ContentPage entries
           for the given language, but we still get ContentPageIndex rows for
           all languages.
        """
        url = f"/admin/home/contentpage/?export={self.format}"
        if locale:
            loc = Locale.objects.get(language_code=locale)
            locale = str(loc)
            url = f"{url}&locale__id__exact={loc.id}"
        content = self.admin_client.get(url).content
        # Hopefully we can get rid of this at some point.
        if locale:
            content = self._filter_export(content, locale=locale)
        if self.format == "csv":
            print("-v-CONTENT-v-")
            print(content.decode())
            print("-^-CONTENT-^-")

        return content

    def export_ordered_content(self) -> bytes:
        """
        Export ordered content in the configured format.
        """
        url = f"/admin/snippets/home/orderedcontentset/?export={self.format}"

        stream = self.admin_client.get(url)
        content = b"".join(stream.streaming_content)
        return content

    def import_content(self, content_bytes: bytes, **kw: Any) -> Any:
        """
        Import given content in the configured format with the configured importer.
        """
        return import_content(
            BytesIO(content_bytes), self.format.upper(), Queue(), **kw
        )

    def import_ordered_sets(self, content_bytes: bytes) -> None:
        import_ordered_sets(BytesIO(content_bytes), self.format.upper(), Queue())

    def read_bytes(self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE) -> bytes:
        return (path_base / path_str).read_bytes()

    def import_file(
        self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE, **kw: Any
    ) -> bytes:
        """
        Import given content file in the configured format with the configured importer.
        """
        content = self.read_bytes(path_str, path_base)
        self.import_content(content, **kw)
        return content

    def import_ordered_file(
        self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE, **kw: Any
    ) -> bytes:
        """
        Import given ordered content file in the configured format with the configured importer.
        """
        content = self.read_bytes(path_str, path_base)
        self.import_ordered_sets(content, **kw)
        return content

    def export_reimport(self) -> None:
        """
        Export all content, then immediately reimport it.
        """
        self.import_content(self.export_content())

    def get_page_json(self, locale: str | None = None) -> DbDicts:
        """
        Serialize all ContentPage and ContentPageIndex instances and normalize
        things that vary across import/export.
        """
        pages = decode_json_fields(get_page_json())
        for ff in PAGE_FILTER_FUNCS:
            pages = ff(pages)
        if locale is not None:
            loc = Locale.objects.get(language_code=locale)
            pages = [p for p in pages if p["fields"]["locale"] == loc.id]
        return sorted(pages, key=lambda p: p["pk"])

    def csvs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = csv2dicts(src_bytes)
        dst = csv2dicts(dst_bytes)
        return filter_exports(src, dst)

    def xlsxs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = xlsx2dicts(src_bytes)
        dst = xlsx2dicts(dst_bytes)
        return filter_exports(src, dst)


@pytest.fixture()
def csv_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "csv")


@pytest.fixture()
def xlsx_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "xlsx")


@pytest.mark.django_db
class TestImportExportRoundtrip:
    """
    Test importing and reexporting content produces an export that is
    equilavent to the original imported content.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_simple(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file and then exporting it produces the correct new export format.

        (This uses content2.csv from test_api.py.)

        FIXME:
         * This should probably be in a separate test for importing old exports.
         * Do we need page.id to be exported? At the moment nothing in the
           import reads that.
         * Do we expect imported content to have leading spaces removed?
         * Should we set enable_web and friends based on body, title, or an
           enable field that we'll need to add to the export?
        """
        csv_bytes = csv_impexp.import_file("content2.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_less_simple(self, csv_impexp: ImportExport) -> None:
        """
        Importing a less simple CSV file and then exporting it produces a
        duplicate of the original file.

        (This uses exported_content_20230911-variations-linked-page.csv.)

        FIXME:
         * Implement import/export for doc_link, image_link, media_link.
        """
        set_profile_field_options()
        csv_bytes = csv_impexp.import_file(
            "exported_content_20230911-variations-linked-page.csv"
        )
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_multiple_messages(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file containing multiple messages of each type for a
        page and then exporting it produces a duplicate of the original file.

        (This uses multiple_messages.csv.)
        """
        set_profile_field_options()
        csv_bytes = csv_impexp.import_file("multiple_messages.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_default_locale(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with multiple languages and specifying a locale
        and then exporting it produces a duplicate of the original file but
        with only pages from the specifyied specified locale included.

        (This uses translations.csv and the en language-specific subset thereof.)
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        set_profile_field_options()
        csv_impexp.import_file("translations.csv", locale="en")
        csv_bytes = csv_impexp.read_bytes("translations-en.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_translated_locale(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with multiple languages and specifying a locale
        and then exporting it produces a duplicate of the original file but
        with only pages from the specifyied specified locale included.

        (This uses translations.csv and the pt language-specific subset thereof.)
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        set_profile_field_options()
        csv_impexp.import_file("translations.csv", locale="pt")
        csv_bytes = csv_impexp.read_bytes("translations-pt.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_all_locales(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file containing translations and then exporting it
        produces a duplicate of the original file.

        (This uses translations.csv.)
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        set_profile_field_options()
        csv_bytes = csv_impexp.import_file("translations.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_all_locales_split(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file split into separate parts per locale and then
        exporting it produces a duplicate of the original file.

        (This uses translations.csv and the two language-specific subsets thereof.)
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        set_profile_field_options()
        csv_bytes = csv_impexp.read_bytes("translations.csv")
        csv_impexp.import_file("translations.csv", locale="en")
        csv_impexp.import_file("translations.csv", purge=False, locale="pt")

        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_list_items_values(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file containing a string of list items of each type for a
        page and then exporting it produces the correct new export format.

        (This uses list_items.csv and list_items_output.csv.)
        """
        set_profile_field_options()
        csv_impexp.import_file("list_items.csv")
        content = csv_impexp.export_content()

        new_export_content = csv_impexp.read_bytes(
            "list_items_output.csv", IMP_EXP_DATA_BASE
        )

        src, dst = csv_impexp.csvs2dicts(new_export_content, content)
        assert dst == src

    def test_list_items_values_with_comma(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file containing a string of list items that has a comma
        page and then exporting it produces the correct new export format.

        (This uses list_items_with_comma.csv list_items_with_comma_output.csv.)
        """
        set_profile_field_options()
        # We import list_items.csv so that we can use the page as a reference in the
        csv_impexp.import_file("list_items.csv")
        csv_impexp.import_file("list_items_with_comma.csv")
        content = csv_impexp.export_content()

        new_export_content = csv_impexp.read_bytes(
            "list_items_with_comma_output.csv", IMP_EXP_DATA_BASE
        )

        src, dst = csv_impexp.csvs2dicts(new_export_content, content)

        assert dst == src


@pytest.mark.django_db
class TestImportExport:
    """
    Text various import and export scenarios that aren't specifically round
    trips.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_import_error(self, csv_impexp: ImportExport) -> None:
        """
        Importing an invalid CSV file leaves the db as-is.

        (This uses content2.csv from test_api.py and broken.csv.)
        """
        # Start with some existing content.
        csv_bytes = csv_impexp.import_file("content2.csv")

        # This CSV doesn't have any of the fields we expect.
        with pytest.raises((KeyError, TypeError, ImportException)):
            csv_impexp.import_file("broken.csv")

        # The export should match the existing content.
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_missing_slug(self, csv_impexp: ImportExport) -> None:
        """
        Importing pages without slugs causes a validation error.

        (This uses missing-slug.csv.)
        """

        # One of the content page rows doesn't have a slug.
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-slug.csv")

        assert e.value.row_num == 3
        assert e.value.message == ["Missing slug value"]

    def test_missing_slug_on_index_page(self, csv_impexp: ImportExport) -> None:
        """
        Importing index pages without slugs causes a validation error.

        (This uses missing-slug-index.csv.)
        """

        # One of the index page rows doesn't have a slug.
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-slug-index.csv")

        assert e.value.row_num == 2
        assert e.value.message == ["Missing slug value"]

    def test_no_translation_key_default(self, csv_impexp: ImportExport) -> None:
        """
        Importing pages without translation keys in the default locale causes
        wagtail to generate new translation keys.

        (This uses no-translation-key-default.csv.)
        """
        csv_bytes = csv_impexp.import_file("no-translation-key-default.csv")

        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        # Check that the export has translation keys for all rows and clear
        # them to match the imported data
        for row in dst:
            assert len(row["translation_tag"]) == 36  # uuid with dashes
            row["translation_tag"] = ""

        assert dst == src

    def test_no_translation_key_nondefault(self, csv_impexp: ImportExport) -> None:
        """
        Importing pages without translation keys in the non-default locale
        causes a validation error.

        (This uses no-translation-key-cpi.csv and no-translation-key-cp.csv.)
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # A ContentPageIndex without a translation key fails
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("no-translation-key-cpi.csv")

        assert e.value.row_num == 4

        assert e.value.message == [
            "Validation error: translation_key - “” is not a valid UUID."
        ]

        # A ContentPage without a translation key fails
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("no-translation-key-cp.csv")

        assert e.value.row_num == 5

        assert e.value.message == [
            "Validation error: translation_key - “” is not a valid UUID."
        ]

    def test_invalid_locale_name(self, csv_impexp: ImportExport) -> None:
        """
        Importing pages with invalid locale names should raise an error that results
        in an error message that gets sent back to the user
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("invalid-locale-name.csv")

        assert e.value.row_num == 2
        assert e.value.message == ["Language not found: NotEnglish"]

    def test_multiple_locales_for_name(
        self, csv_impexp: ImportExport, settings: SettingsWrapper
    ) -> None:
        """
        Importing pages with locale names that represent multiple locales should raise
        an error that results in an error message that gets sent back to the user
        """
        settings.WAGTAIL_CONTENT_LANGUAGES = settings.LANGUAGES = [
            ("en1", "NotEnglish"),
            ("en2", "NotEnglish"),
        ]
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("invalid-locale-name.csv")

        assert e.value.row_num == 2
        assert e.value.message == [
            "Multiple codes for language: NotEnglish -> ['en1', 'en2']"
        ]

    def test_locale_HomePage_DNE(self, csv_impexp: ImportExport) -> None:
        """
        Importing files with non default locale HomePages that do not exist in the db should raise
        an error that results in an error message that gets sent back to the user
        """
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("content_without_locale_homepage.csv")
        assert e.value.row_num == 13
        assert e.value.message == [
            "You are trying to add a child page to a 'Portuguese' HomePage that does not exist. Please create the 'Portuguese' HomePage first"
        ]

    def test_missing_parent(self, csv_impexp: ImportExport) -> None:
        """
        If the import file specifies a parent title, but there are no pages with that
        title, then an error message should get sent back to the user.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-parent.csv")

        assert e.value.row_num == 2
        assert e.value.message == [
            "Cannot find parent page with title 'missing-parent' and locale "
            "'English'"
        ]

    def test_multiple_parents(self, csv_impexp: ImportExport) -> None:
        """
        Because we use the title to find a parent page, and it's possible to have
        multiple pages with the same title, it's possible to have the situation where
        we don't know which parent this import points to. In that case we should show
        the user an error message, with information that will allow them to fix it.
        """
        home_page = HomePage.objects.first()
        PageBuilder.build_cpi(home_page, "missing-parent1", "missing-parent")
        PageBuilder.build_cpi(home_page, "missing-parent2", "missing-parent")

        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-parent.csv", purge=False)
        assert e.value.row_num == 2
        assert e.value.message == [
            "Multiple pages with title 'missing-parent' and locale 'English' for "
            "parent page: ['missing-parent1', 'missing-parent2']"
        ]

    def test_message_for_missing_page(self, csv_impexp: ImportExport) -> None:
        """
        If we try to import a message for a page that isn't in the same import,
        an error message should get sent back to the user.

        FIXME:
         * We currently assume that messages belong to the content page immediately
           above them, but we don't check or enforce this.
         * We also get the locale from the content page immediately above the message.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("message-row-missing-page.csv")

        assert e.value.row_num == 4
        assert e.value.message == [
            "This is a message for page with slug 'not-cp-import-export' and locale "
            "'English', but no such page exists"
        ]

    def test_variation_for_missing_page(self, csv_impexp: ImportExport) -> None:
        """
        If we try to import a variation message for a page that isn't in the
        same import, an error message should get sent back to the user.

        FIXME:
         * We currently assume that messages belong to the content page immediately
           above them, but we don't check or enforce this.
         * We also get the locale from the content page immediately above the message.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("variation-row-missing-page.csv")

        assert e.value.row_num == 4
        assert e.value.message == [
            "This is a variation for the content page with slug 'not-cp-import-export' and locale "
            "'English', but no such page exists"
        ]

    def test_go_to_page_button_missing_page(self, csv_impexp: ImportExport) -> None:
        """
        Go to page buttons in the import file link to other pages using the slug. But
        if no page with that slug exists, then we should give the user an error message
        that tells them where and how to fix it.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-gotopage.csv")
        assert e.value.row_num == 2
        assert e.value.message == [
            "No pages found with slug 'missing' and locale 'English' for go_to_page "
            "button 'Missing' on page 'ma_import-export'"
        ]

    def test_go_to_form_button_missing_form(self, csv_impexp: ImportExport) -> None:
        """
        Go to form buttons in the import file link to other forms using the slug. But
        if no form with that slug exists, then we should give the user an error message
        that tells them where and how to fix it.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-gotoform.csv")
        assert e.value.row_num == 2
        assert e.value.message == [
            "No form found with slug 'missing' and locale 'English' for go_to_form "
            "button 'Missing' on page 'ma_import-export'"
        ]

    def test_go_to_form_list_missing_form(self, csv_impexp: ImportExport) -> None:
        """
        Go to form list items buttons in the import file link to other forms using the
        slug. But if no form with that slug exists, then we should give the user an
        error message that tells them where and how to fix it.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-gotoform-list.csv")
        assert e.value.row_num == 2
        assert e.value.message == [
            "No form found with slug 'missing' and locale 'English' for go_to_form "
            "list item 'Missing' on page 'ma_import-export'"
        ]

    def test_missing_related_pages(self, csv_impexp: ImportExport) -> None:
        """
        Related pages are listed as comma separated slugs in imported files. If there
        is a slug listed that we cannot find the page for, then we should show the
        user an error with information about the missing page.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("missing-related-page.csv")
        assert e.value.row_num == 2
        assert e.value.message == [
            "Cannot find related page with slug 'missing related' and locale "
            "'English'"
        ]

    def test_invalid_wa_template_category(self, csv_impexp: ImportExport) -> None:
        """
        Importing a WhatsApp template with an invalid category should raise an
        error that results in an error message that gets sent back to the user.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-whatsapp-template-category.csv")

        assert e.value.row_num == 3
        # FIXME: Find a better way to represent this.

        # print(e.value.message)
        assert e.value.message == [
            "Validation error: whatsapp_template_category - Select a valid choice. Marketing is not one of the available choices."
        ]

    def test_invalid_wa_template_vars(self, csv_impexp: ImportExport) -> None:
        """
        Importing a WhatsApp template with invalid variables should raise an
        error that results in an error message that gets sent back to the user.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-whatsapp-template-vars.csv")

        assert e.value.row_num == 3
        # FIXME: Find a better way to represent this.
        assert e.value.message == [
            "Validation error: example_values - The number of example values provided (1) does not match the number of variables used in the template (3)"
        ]

    def test_invalid_wa_template_vars_update(self, csv_impexp: ImportExport) -> None:
        """
        Updating a valid WhatsApp template with invalid variables should raise
        an error that results in an error message that gets sent back to the
        user. The update validation happens in a different code path from the
        initial import.
        """
        csv_impexp.import_file("good-whatsapp-template-vars.csv")

        # Update an existing page, which does the validation in
        # `page.save_revision()` rather than `parent.add_child()`.
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-whatsapp-template-vars.csv", purge=False)

        assert e.value.row_num == 3
        # FIXME: Find a better way to represent this.
        assert e.value.message == [
            "Validation error: example_values - The number of example values provided (1) does not match the number of variables used in the template (3)"
        ]

    def test_cpi_validation_failure(self, csv_impexp: ImportExport) -> None:
        """
        Importing a ContentPageIndex with an invalid translation key should
        raise an error that results in an error message that gets sent back to
        the user.
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-cpi-translation-key.csv")

        assert e.value.row_num == 2
        # FIXME: Find a better way to represent this.
        assert e.value.message == [
            "Validation error: translation_key - “BADUUID” is not a valid UUID."
        ]

    def test_cpi_validation_failure_update(self, csv_impexp: ImportExport) -> None:
        """
        Updating a valid ContentPageIndex with an invalid translation key
        should raise an error that results in an error message that gets sent
        back to the user. The update validation happens in a different code
        path from the initial import.
        """
        csv_impexp.import_file("good-cpi-translation-key.csv")

        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-cpi-translation-key.csv", purge=False)

        assert e.value.row_num == 2
        # FIXME: Find a better way to represent this.
        assert e.value.message == [
            "Validation error: translation_key - “BADUUID” is not a valid UUID."
        ]

    def test_ContentPageIndex_required_fields(self, csv_impexp: ImportExport) -> None:
        """
        Importing an CSV file with only the required fields for a ContentPageIndex shoud not break
        """

        csv_bytes = csv_impexp.import_file("contentpage_index_required_fields.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)

        # the importer adds extra fields, so we filter for the ones we want
        allowed_keys = ["message", "slug", "parent", "web_title", "locale"]
        dst = [{k: v for k, v in item.items() if k in allowed_keys} for item in dst]
        src = [{k: v for k, v in item.items() if k in allowed_keys} for item in src]

        [main_menu] = ContentPageIndex.objects.all()
        assert main_menu.slug == "main-menu"

        assert src == dst

    def test_ContentPage_required_fields(self, csv_impexp: ImportExport) -> None:
        """
        Importing an CSV file with only the required fields for a ContentPage shoud not break
        """

        csv_bytes = csv_impexp.import_file("contentpage_required_fields.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)

        # the importer adds extra fields, so we filter for the ones we want
        allowed_keys = ["message", "slug", "parent", "web_title", "locale"]
        dst = [{k: v for k, v in item.items() if k in allowed_keys} for item in dst]
        src = [{k: v for k, v in item.items() if k in allowed_keys} for item in src]

        [main_menu] = ContentPageIndex.objects.all()
        [first_time_user, health_info] = ContentPage.objects.all()

        assert main_menu.slug == "main_menu"
        assert first_time_user.slug == "first_time_user"
        assert health_info.slug == "health_info"

        assert src == dst

    def test_footer_maximum_characters(self, csv_impexp: ImportExport) -> None:
        """
        Importing an CSV file with footer and and footer characters exceeding maximum charactercount
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("whatsapp_footer_max_characters.csv")

        assert isinstance(e.value, ImportException)
        assert e.value.row_num == 4
        assert e.value.message == [
            "Validation error: footer - Ensure this value has at most 60 characters (it has 110)."
        ]

    def test_fields_containing_only_whitespace(self, csv_impexp: ImportExport) -> None:
        """
        A page_id or footer containing only whitespace is not an error.

        NOTE: This test passes if no exception is raised.
        """
        csv_impexp.import_file("whitespace-only-fields.csv")

    def test_list_items_maximum_num(self, csv_impexp: ImportExport) -> None:
        """
        Importing an CSV file with list_items and and list items characters exceeding maximum charactercount
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("whatsapp_list_items_max_num.csv")

        assert isinstance(e.value, ImportException)
        assert e.value.row_num == 4
        assert e.value.message == [
            "Validation error: list_items - The maximum number of items is 10"
        ]

    def test_list_items_maximum_characters(self, csv_impexp: ImportExport) -> None:
        """
        Importing an CSV file with list_items and and list items characters exceeding maximum character count
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("whatsapp_list_items_max_characters.csv")

        assert isinstance(e.value, ImportException)
        assert e.value.row_num == 4
        assert e.value.message == [
            "Validation error: list_items - Ensure this value has at most 24 characters (it has 31)."
        ]

    def test_max_char_variation(self, csv_impexp: ImportExport) -> None:
        """
        Importing a file with the variation message greater than 4096 characters should
        return a validation error to the user
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("max_char_variation.csv")

        assert e.value.row_num == 4
        assert e.value.message == [
            "Validation error: variation_messages - The minimum number of items is 1",
            "Validation error: variation_messages - Ensure this value has at most 4096 characters (it has 4097).",
        ]

    def test_invalid_JSON_button(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with an invalid JSON value for the button should return a detailed error message
        to the user
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("broken_button.csv")

        assert e.value.row_num == 3
        assert e.value.message == ["Bad JSON button, you have: Broken Button"]

    def test_invalid_JSON_button_xlsx(self, xlsx_impexp: ImportExport) -> None:
        """
        Importing a XLSX file with an invalid JSON value for the button should return a detailed error message
        to the user
        """
        with pytest.raises(ImportException) as e:
            xlsx_impexp.import_file("broken_button.xlsx", purge=True)
        assert e.value.row_num == 3
        assert e.value.message == ["Bad JSON button, you have: Broken button"]

    def test_max_char_button(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with a button chars greater than the limit should return a detailed error message
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("max-chars-button.csv")

        assert e.value.row_num == 15
        assert e.value.message == [
            "Validation error: buttons - Ensure this value has at most 20 characters (it has 23)."
        ]

    def test_max_varation_xlsx(self, xlsx_impexp: ImportExport) -> None:
        """
        Importing a XLSX file with the variation message greater than 4096 characters should
        return a validation error to the user.
        """
        with pytest.raises(ImportException) as e:
            xlsx_impexp.import_file("max_char_variation.xlsx", purge=True)
        assert e.value.row_num == 3
        assert e.value.message == [
            "Validation error: variation_messages - The minimum number of items is 1",
            "Validation error: variation_messages - Ensure this value has at most 4096 characters (it has 4319).",
        ]

    def test_max_WA_body_xlsx(self, xlsx_impexp: ImportExport) -> None:
        """
        Importing a XLSX file with the whatsapp body message greater than 4096 characters should
        return a validation error to the user.
        """
        with pytest.raises(ImportException) as e:
            xlsx_impexp.import_file("max_char_WA_body.xlsx", purge=True)
        assert e.value.row_num == 3
        assert e.value.message == [
            "Validation error: message - The minimum number of items is 1",
            "Validation error: message - Ensure this value has at most 4096 characters (it has 4319).",
        ]

    @pytest.mark.xfail(reason="Form creation during import needs to be fixed.")
    def test_multiple_field_errors(self, csv_impexp: ImportExport) -> None:
        """
        Importing a file with multiple errors for different fields should return errors for each of those fields,
        at the same time, for the same row
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("multiple-validation-errors.csv")

        assert e.value.row_num == 3
        assert e.value.message == [
            "Validation error: example_values - The number of example values provided (1) does not match the number of variables used in the template (2)",
            "Validation error: whatsapp_template_category - Select a valid choice. Marketing is not one of the available choices.",
        ]

    def test_import_ordered_content_sets_error(self, csv_impexp: ImportExport) -> None:
        """
        Importing a broken CSV for ordered content sets should throw an exception and give a detailed error message
        """
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content_broken.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.row_num == 2
        assert e.value.message == [
            "Row Test Set has 2 times, 2 units, 3 before_or_afters, 3 page_slugs and 3 contact_fields and they should all be equal."
        ]

    def test_import_ordered_content_sets_incorrect_time_values_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a broken CSV for ordered content sets should throw an exception and give a detailed error message
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content_incorrect_time_values.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.row_num == 2
        assert e.value.message == ["Validation error: time - Enter a whole number."]

    def test_import_ordered_content_sets_incorrect_unit_values_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a broken CSV for ordered content sets should throw an exception and give a detailed error message
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content_incorrect_unit_values.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.row_num == 2
        assert e.value.message == [
            "Validation error: unit - Select a valid choice. 1 is not one of the available choices."
        ]

    def test_import_ordered_content_sets_incorrect_before_or_after_values_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a broken CSV for ordered content sets should throw an exception and give a detailed error message
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes(
                "ordered_content_incorrect_before_or_after_values.csv"
            )
            csv_impexp.import_ordered_sets(content)

        assert e.value.row_num == 2
        assert e.value.message == [
            "Validation error: before_or_after - Select a valid choice. 1 is not one of the available choices."
        ]

    def test_import_ordered_content_sets_no_page_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets without pages should throw an error.
        """
        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.message == [
            "Content page not found for slug 'first_time_user' in locale 'English'"
        ]

    def test_import_ordered_content_sets_missing_slug(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets that has a page slug, but no
        Unit/Time/EDD/Before_or_After/Contact field, should result in a
        error message
        """
        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("bad_ordered_set.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.row_num == 2
        assert e.value.message == [
            "You are attempting to import an ordered content set with page details, but no page slug."
        ]

    def test_import_ordered_content_sets_no_locale_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets without a locale should throw an error.
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content_no_locale.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.message == ["No locale specified."]

    def test_import_ordered_content_sets_incorrect_locale_error(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets with an incorrect locale should throw an error.
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        with pytest.raises(ImportException) as e:
            content = csv_impexp.read_bytes("ordered_content_incorrect_locale.csv")
            csv_impexp.import_ordered_sets(content)

        assert e.value.message == ["Locale pt does not exist."]

    def test_import_ordered_content_sets_duplicate_name(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets with a duplicate name but unique slugs
        should not throw an error.
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        content = csv_impexp.read_bytes("ordered_content_same_name.csv")
        csv_impexp.import_ordered_sets(content)

        ordered_sets = OrderedContentSet.objects.all()
        assert len(ordered_sets) == 2

    def test_import_ordered_content_sets_duplicate_slug(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing CSV for ordered content sets with a duplicate and duplicate slugs
        should update the existing OrderedContentSet to the last one in the CSV.
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        content = csv_impexp.read_bytes("ordered_content_same_slug.csv")
        csv_impexp.import_ordered_sets(content)

        ordered_sets = OrderedContentSet.objects.all()
        assert len(ordered_sets) == 1

        ordered_set = ordered_sets[0]
        assert ordered_set.name == "Test Set"
        assert ordered_set.slug == "test_set"
        assert ordered_set.locale.language_code == "en"
        assert unwagtail(ordered_set.profile_fields) == [
            ("gender", "male"),
            ("relationship", "in_a_relationship"),
        ]

        pages = unwagtail(ordered_set.pages)
        assert len(pages) == 1

        page = pages[0][1]
        assert page["contentpage"].slug == "first_time_user"

    def test_import_simple_ordered_sets_csv(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with ordered content sets without any contentpage should not break
        """

        content = csv_impexp.read_bytes("simple_ordered_set.csv")
        csv_impexp.import_ordered_sets(content)

        en = Locale.objects.get(language_code="en")

        ordered_set = OrderedContentSet.objects.filter(slug="slug", locale=en).first()

        assert ordered_set.name == "Ordered"
        pages = unwagtail(ordered_set.pages)
        assert len(pages) == 0

    def test_import_ordered_sets_csv(self, csv_impexp: ImportExport) -> None:
        """
        Importing a CSV file with ordered content sets should not break
        """
        set_profile_field_options()
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        csv_impexp.import_file("contentpage_required_fields_multi_locale.csv")
        content = csv_impexp.read_bytes("ordered_content.csv")
        csv_impexp.import_ordered_sets(content)

        en = Locale.objects.get(language_code="en")

        ordered_set = OrderedContentSet.objects.filter(
            slug="test_set", locale=en
        ).first()

        assert ordered_set.name == "Test Set"
        pages = unwagtail(ordered_set.pages)
        assert len(pages) == 3

        page = pages[0][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "2"
        assert page["unit"] == "days"
        assert page["before_or_after"] == "before"
        assert page["contact_field"] == "edd"

        page = pages[1][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "3"
        assert page["unit"] == "months"
        assert page["before_or_after"] == "before"
        assert page["contact_field"] == "edd"

        page = pages[2][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "4"
        assert page["unit"] == "minutes"
        assert page["before_or_after"] == "after"
        assert page["contact_field"] == "edd"
        assert unwagtail(ordered_set.profile_fields) == [
            ("gender", "male"),
            ("relationship", "in_a_relationship"),
        ]

        ordered_set_pt = OrderedContentSet.objects.filter(
            slug="test_set", locale=pt
        ).first()

        pages = unwagtail(ordered_set_pt.pages)
        assert len(pages) == 1

        page = pages[0][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "2"
        assert page["unit"] == "days"
        assert page["before_or_after"] == "before"
        assert page["contact_field"] == "edd"

        assert unwagtail(ordered_set_pt.profile_fields) == [
            ("gender", "male"),
            ("relationship", "in_a_relationship"),
        ]

    def test_import_ordered_sets_duplicate_header_csv(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a CSV with duplicate headers should throw an error
        """
        set_profile_field_options()
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        csv_impexp.import_file("contentpage_required_fields_multi_locale.csv")
        content = csv_impexp.read_bytes("ordered_content_duplicate_header.csv")

        with pytest.raises(ImportException) as e:
            csv_impexp.import_ordered_sets(content)

        assert e.value.message == [
            "Invalid format. Please check that there are no duplicate headers."
        ]

    def test_import_ordered_sets_no_profile_fields_csv(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a CSV file with ordered content sets should not break
        """
        csv_impexp.import_file("contentpage_required_fields.csv")
        content = csv_impexp.read_bytes("ordered_content_no_profile_fields.csv")
        csv_impexp.import_ordered_sets(content)

        locale = Locale.objects.get(language_code="en")
        ordered_set = OrderedContentSet.objects.filter(
            name="Test Set", slug="test_set", locale=locale
        ).first()

        assert ordered_set.name == "Test Set"
        pages = unwagtail(ordered_set.pages)
        assert len(pages) == 1
        page = pages[0][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "2"
        assert page["unit"] == "days"
        assert page["before_or_after"] == "before"
        assert page["contact_field"] == "edd"
        assert unwagtail(ordered_set.profile_fields) == []

    def test_import_ordered_sets_xlsx(
        self, xlsx_impexp: ImportExport, csv_impexp: ImportExport
    ) -> None:
        """
        Importing a XLSX file with ordered content sets should not break
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")
        content = xlsx_impexp.read_bytes("ordered_content.xlsx")
        xlsx_impexp.import_ordered_sets(content)

        locale = Locale.objects.get(language_code="en")
        ordered_set = OrderedContentSet.objects.filter(
            name="Test Set", slug="test-set", locale=locale
        ).first()

        assert ordered_set.name == "Test Set"
        assert ordered_set.slug == "test-set"
        assert ordered_set.locale == locale
        pages = unwagtail(ordered_set.pages)
        assert len(pages) == 1
        page = pages[0][1]
        assert page["contentpage"].slug == "first_time_user"
        assert page["time"] == "2"
        assert page["unit"] == "days"
        assert page["before_or_after"] == "before"
        assert page["contact_field"] == "edd"
        assert unwagtail(ordered_set.profile_fields) == [
            ("gender", "male"),
            ("relationship", "in_a_relationship"),
        ]

    def test_changed_parentpage(self, csv_impexp: ImportExport) -> None:
        """
        Users should not be allowed to import a file where a parent of an existing contentpage. A descriptive error should be sent back.
        """
        home_page = HomePage.objects.first()

        _self_help = PageBuilder.build_cp(
            parent=home_page,
            slug="self-help",
            title="self help",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )

        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("changed_parent.csv", purge=False)
        assert e.value.row_num == 5
        assert e.value.message == [
            "Changing the parent from 'Home' to 'Main Menu' for the page with title 'self-help' during import is not allowed. Please use the UI"
        ]

    def test_import_pages_xlsx(self, xlsx_impexp: ImportExport) -> None:
        """
        Importing an XLSX file with content pages should not break
        """
        xlsx_impexp.import_file("content_pages.xlsx", purge=False)
        content_pages = ContentPage.objects.all()
        assert len(content_pages) > 0

    def test_import_pages_number_type(self, xlsx_impexp: ImportExport) -> None:
        """
        Importing an XLSX file where number fields have a number cell formatting
        shouldn't break
        """
        home_page = HomePage.objects.first()
        PageBuilder.build_cpi(home_page, "main-menu", "main menu first time user")
        xlsx_impexp.import_file("contentpage_number_type.xlsx", purge=False)
        content_pages = ContentPage.objects.all()
        assert len(content_pages) > 0

    def test_invalid_page(self, csv_impexp: ImportExport) -> None:
        """
        Import an invalid page that matches a valid page already in the db
        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="main-menu-first-time-user",
            title="main menu first time user",
            bodies=[
                WABody("HA menu", [WABlk("Welcome M")]),
            ],
        )

        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("broken_button.csv")

        assert e.value.row_num == 3
        assert e.value.message == ["Bad JSON button, you have: Broken Button"]

    def test_invalid_page_already_in_db(self, csv_impexp: ImportExport) -> None:
        """
        Import an invalid page that matches an invalid page already in the db
        """

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        wa_block = [WABlk("Vars {1} {2} {3}", example_values=["Example value 1"])]
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="main-menu-first-time-user",
            title="main menu first time user",
            bodies=[WABody("HA menu ", wa_block)],
        )

        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("bad-whatsapp-template-vars.csv")

        assert e.value.row_num == 3
        # FIXME: Find a better way to represent this.
        assert e.value.message == [
            "Validation error: example_values - The number of example values provided (1) does not match the number of variables used in the template (3)"
        ]

    def test_language_code_import(self, csv_impexp: ImportExport) -> None:
        """
        Import a page that doesn't have language code set
        """
        csv_impexp.import_file("language_code_import.csv")
        content = csv_impexp.export_content()

        new_export_content = csv_impexp.read_bytes(
            "language_code_import_output.csv", IMP_EXP_DATA_BASE
        )

        src, dst = csv_impexp.csvs2dicts(new_export_content, content)
        assert dst == src

    def test_button_max_length(self, csv_impexp: ImportExport) -> None:
        """
        Import a page that doesn't have language code set
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("button_max_length.csv")
        assert e.value.row_num == 3
        assert e.value.message == [
            "Validation error: message - A WhatsApp message with interactive items cannot be longer than 1024 characters, your message is 1075 characters long"
        ]

    def test_hidden_characters(self, csv_impexp: ImportExport) -> None:
        """
        Import a page that has hidden characters in the whatsapp body
        """
        csv_bytes = csv_impexp.import_file("test_special_chars.csv")
        assert "\u2028\u2028" in csv_bytes.decode("utf-8")
        assert "\u2028" not in ContentPage.objects.all().values()[0]["whatsapp_body"]

    def test_list_item_descriptive_error_message(
        self, csv_impexp: ImportExport
    ) -> None:
        """
        Import failure for list_items should return a descriptive error message
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("list_items_with_errors.csv")
        assert e.value.row_num == 3
        assert e.value.message == ["list item is missing key 'type'"]

    def test_list_item_type_error_message(self, csv_impexp: ImportExport) -> None:
        """
        Import invalid list item type should return a invalid type error message
        """
        with pytest.raises(ImportException) as e:
            csv_impexp.import_file("list_items_with_type_error.csv")
        assert e.value.row_num == 3
        assert e.value.message == ["list item with invalid type 'new_type'"]

    def test_media_link_warning(self, csv_impexp: ImportExport) -> None:
        """
        Import a page with media link it should return a warning
        no error should be raised
        """

        csv_impexp.import_file("contentpage_media_link_warning.csv")

        page = Page.objects.all()

        assert len(page) > 0

    def test_media_link_warning_response(self, csv_impexp: ImportExport) -> None:
        """
        Import a page with media link it should return a warning
        media_link will be excluded from uploaded data
        """

        resp = csv_impexp.import_file("contentpage_media_link_warning.csv")

        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(resp, content)

        assert "media_link" not in [item.keys() for item in dst]

    def test_import_success_no_warnings(self, csv_impexp: ImportExport) -> None:
        """
        Import a page that will not have warnings list
        """
        content = csv_impexp.read_bytes("contentpage_without_warning.csv")
        importer = csv_impexp.import_content(content)

        assert importer.import_warnings == []

    def test_import_success_with_warnings(self, csv_impexp: ImportExport) -> None:
        """
        Import a page that will return warnings if media_link is not empty
        """
        content = csv_impexp.read_bytes("contentpage_media_link_warning.csv")
        importer = csv_impexp.import_content(content)

        assert len(importer.import_warnings) == 2
        assert (
            importer.import_warnings[0].message
            == "Media import not supported, http://test.com/image.png not added to main-menu-first-time-user"
        )
        assert importer.import_warnings[0].row_num == 3


@pytest.mark.django_db
class TestExport:
    """
    Test that the export is valid.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_ordered_content_set_export(self, csv_impexp: ImportExport) -> None:
        """
        Ordered Content Sets should export all pages correctly
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        imported_content = csv_impexp.import_ordered_file(
            "ordered_content_multiple_contact_fields.csv"
        )

        exported_content = csv_impexp.export_ordered_content()

        src, dst = csv_impexp.csvs2dicts(imported_content, exported_content)
        assert src == dst

    def test_ordered_content_XLSX_export(
        self, xlsx_impexp: ImportExport, csv_impexp: ImportExport
    ) -> None:
        """
        Ordered Content Sets should export in XLSX format correctly
        """
        set_profile_field_options()
        csv_impexp.import_file("contentpage_required_fields.csv")

        imported_content = xlsx_impexp.import_ordered_file("ordered_content.xlsx")

        exported_content = xlsx_impexp.export_ordered_content()

        src, dst = csv_impexp.xlsxs2dicts(imported_content, exported_content)
        assert src == dst

    def test_export_wa_with_image(self, impexp: ImportExport) -> None:
        img_path = Path("home/tests/test_static") / "test.jpeg"
        img_wa = mk_img(img_path, "wa_image")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HA menu", [WABlk("Welcome WA", image=img_wa.id)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None

    def test_export_viber_with_image(self, impexp: ImportExport) -> None:
        img_path = Path("home/tests/test_static") / "test.jpeg"
        img_v = mk_img(img_path, "v_image")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                VBody("HA menu", [VBlk("Welcome V", image=img_v.id)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None

    def test_export_messenger_with_image(self, impexp: ImportExport) -> None:
        img_path = Path("home/tests/test_static") / "test.jpeg"
        img_m = mk_img(img_path, "m_image")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                MBody("HA menu", [MBlk("Welcome M", image=img_m.id)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None

    def test_export_wa_with_media(self, impexp: ImportExport) -> None:
        media_path = Path("home/tests/test_static") / "test.mp4"
        media_wa = mk_media(media_path, "wa_media")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HA menu", [WABlk("Welcome WA", media=media_wa.id)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None

    def test_export_wa_with_document(self, impexp: ImportExport) -> None:
        doc_path = Path("home/tests/test_static") / "test.txt"
        doc_wa = mk_doc(doc_path, "wa_document")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HA menu", [WABlk("Welcome WA", document=doc_wa.id)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None

    def test_export_wa_with_none_document(self, impexp: ImportExport) -> None:
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HA menu", [WABlk("Welcome WA", document=None)]),
            ],
        )
        content = impexp.export_content(locale="en")
        # Export should succeed
        assert content is not None


@pytest.fixture(params=["csv", "xlsx"])
def impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, request.param)


@pytest.fixture()
def tmp_media_path(tmp_path: Path, settings: SettingsWrapper) -> None:
    settings.MEDIA_ROOT = tmp_path


def mk_img(img_path: Path, title: str) -> Image:
    img = Image(title=title, file=ImageFile(img_path.open("rb"), name=img_path.name))
    img.save()
    return img


def mk_media(media_path: Path, title: str) -> File:
    media = Media(title=title, file=File(media_path.open("rb"), name=media_path.name))
    media.save()
    return media


def mk_doc(doc_path: Path, title: str) -> Document:
    doc = Document(title=title, file=File(doc_path.open("rb"), name=doc_path.name))
    doc.save()
    return doc


def add_go_to_page_button(whatsapp_block: Any, button: PageBtn) -> None:
    button_val = GoToPageButton().to_python(button.value_dict())
    whatsapp_block.value["buttons"].append(("go_to_page", button_val))


def add_go_to_form_button(whatsapp_block: Any, button: FormBtn) -> None:
    button_val = GoToFormButton().to_python(button.value_dict())
    whatsapp_block.value["buttons"].append(("go_to_form", button_val))


def add_go_to_form_list_item(whatsapp_block: Any, list_item: FormListItem) -> None:
    list_val = GoToFormListItem().to_python(list_item.value_dict())
    whatsapp_block.value["list_items"].append(("go_to_form", list_val))


@pytest.mark.usefixtures("tmp_media_path")
@pytest.mark.django_db
class TestExportImportRoundtrip:
    """
    Test that the db state after exporting and reimporting content is
    equivalent to what it was before.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_simple(self, impexp: ImportExport) -> None:
        """
        Exporting and then importing leaves the db in the same state it was
        before, except for page_ids, timestamps, and body item ids.

        FIXME:
         * Determine whether we need to maintain StreamField block ids. (I
           think we don't.)
         * Confirm that there's no meaningful difference between null and ""
           for the nullable fields that the importer sets to "", potentially
           make these fields non-nullable.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                SBody("HealthAlert menu", [SBlk("Welcome to HealthAlert S")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
            ],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody("health info", [WABlk("*Health information* WA")]),
                SBody("health info", [SBlk("*Health information* S")]),
                MBody("health info", [MBlk("*Health information* M")]),
            ],
        )
        _self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[
                WABody("self-help", [WABlk("*Self-help programs* WA")]),
                SBody("self-help", [SBlk("*Self-help programs* S")]),
                MBody("self-help", [MBlk("*Self-help programs* M")]),
                VBody("self-help", [VBlk("*Self-help programs* V")]),
            ],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_web_content(self, impexp: ImportExport) -> None:
        """
        ContentPages with web content are preserved across export/import.

        FIXME:
         * The exporter currently emits rendered web content instead of the source data.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[],
            web_body=["Paragraph 1.", "Paragraph 2."],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_multiple_messages(self, impexp: ImportExport) -> None:
        """
        ContentPages with multiple message blocks are preserved across
        export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                SBody("HealthAlert menu", [SBlk("Welcome to HealthAlert S")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
                VBody("HealthAlert menu", [VBlk("Welcome to HealthAlert V")]),
            ],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody("health info", [WABlk(f"wa{i}") for i in [1, 2, 3]]),
                SBody("health info", [SBlk(f"s{i}") for i in [1, 2, 3, 4]]),
                MBody("health info", [MBlk(f"m{i}") for i in [1, 2, 3, 4, 5]]),
                VBody("health info", [VBlk(f"v{i}") for i in [1, 2, 3, 4, 5, 6]]),
            ],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    @pytest.mark.xfail(reason="Image imports are currently broken.")
    def test_images(self, impexp: ImportExport) -> None:
        """
        ContentPages with images in multiple message types are preserved across
        export/import.

        """
        img_path = Path("home/tests/test_static") / "test.jpeg"
        img_wa = mk_img(img_path, "wa_image")
        img_m = mk_img(img_path, "m_image")
        img_v = mk_img(img_path, "v_image")

        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        _ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HA menu", [WABlk("Welcome WA", image=img_wa.id)]),
                MBody("HA menu", [MBlk("Welcome M", image=img_m.id)]),
                VBody("HA menu", [VBlk("Welcome V", image=img_v.id)]),
            ],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_variations(self, impexp: ImportExport) -> None:
        """
        ContentPages with variation messages (and buttons and next prompts) are
        preserved across export/import.

        """
        set_profile_field_options()

        home_page = HomePage.objects.first()
        imp_exp = PageBuilder.build_cpi(home_page, "import-export", "Import Export")

        cp_imp_exp_wablks = [
            WABlk(
                "Message 1",
                next_prompt="Next message",
                buttons=[NextBtn("Next message")],
                variation_messages=[
                    VarMsg("Var'n for Single", relationship="single"),
                    VarMsg("Var'n for Complicated", relationship="complicated"),
                ],
            ),
            WABlk(
                "Message 2",
                buttons=[PageBtn("Import Export", page=imp_exp)],
                variation_messages=[VarMsg("Var'n for Rather not say", gender="empty")],
            ),
            WABlk("Message 3 with no variation", next_prompt="Next message"),
        ]
        _cp_imp_exp = PageBuilder.build_cp(
            parent=imp_exp,
            slug="cp-import-export",
            title="CP-Import/export",
            bodies=[WABody("WA import export data", cp_imp_exp_wablks)],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_tags_and_related(self, impexp: ImportExport) -> None:
        """
        ContentPages with tags and related pages are preserved across
        export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
            ],
            tags=["tag1", "tag2"],
        )
        health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[MBody("health info", [MBlk("*Health information* M")])],
            tags=["tag2", "tag3"],
        )
        self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[WABody("self-help", [WABlk("*Self-help programs* WA")])],
            tags=["tag4"],
        )
        PageBuilder.link_related(health_info, [self_help])
        PageBuilder.link_related(self_help, [health_info, main_menu])

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_triggers_and_quick_replies(self, impexp: ImportExport) -> None:
        """
        ContentPages with triggers and quick replies are preserved across
        export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
            ],
            triggers=["trigger1", "trigger2"],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[MBody("health info", [MBlk("*Health information* M")])],
            triggers=["trigger2", "trigger3"],
            quick_replies=["button1", "button2"],
        )
        _self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[WABody("self-help", [WABlk("*Self-help programs* WA")])],
            quick_replies=["button3", "button2"],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_whatsapp_template(self, impexp: ImportExport) -> None:
        """
        ContentPages that are whatsapp templates are preserved across
        export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")])],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[WABody("health info", [WABlk("*Health information* WA")])],
            whatsapp_template_name="template-health-info",
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_translations(self, impexp: ImportExport) -> None:
        """
        ContentPages in multiple languages (with unique-per-locale slugs and
        titles) are preserved across export/import.
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # NOTE: For the results to match without a bunch of work reordering
        # pages and juggling ids, we need all CPIs to be created before all
        # CPs.

        # English CPIs
        home_en = HomePage.objects.get(locale__language_code="en")
        _app_rem = PageBuilder.build_cpi(
            parent=home_en, slug="appointment-reminders", title="Appointment reminders"
        )
        _sbm = PageBuilder.build_cpi(
            parent=home_en, slug="stage-based-messages", title="Stage-based messages"
        )
        _him = PageBuilder.build_cpi(
            parent=home_en, slug="health-info-messages", title="Health info messages"
        )
        _wtt = PageBuilder.build_cpi(
            parent=home_en,
            slug="whatsapp-template-testing",
            title="whatsapp template testing",
        )
        imp_exp = PageBuilder.build_cpi(
            parent=home_en, slug="import-export", title="Import Export"
        )

        # Portuguese CPIs
        home_pt = HomePage.objects.get(locale__language_code="pt")
        imp_exp_pt = PageBuilder.build_cpi(
            parent=home_pt,
            slug="import-export",
            title="Import Export (pt)",
            translated_from=imp_exp,
        )

        # English CPs
        non_templ_wablks = [
            WABlk("this is a non template message"),
            WABlk("this message has a doc"),
            WABlk("this message comes with audio"),
        ]
        non_tmpl = PageBuilder.build_cp(
            parent=imp_exp,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks)],
        )

        # Portuguese CPs
        non_templ_wablks_pt = [
            WABlk("this is a non template message (pt)"),
            WABlk("this message has a doc (pt)"),
            WABlk("this message comes with audio (pt)"),
        ]
        non_tmpl_pt = PageBuilder.build_cp(
            parent=imp_exp_pt,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks_pt)],
            translated_from=non_tmpl,
        )

        assert imp_exp.translation_key == imp_exp_pt.translation_key
        assert non_tmpl.translation_key == non_tmpl_pt.translation_key

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_translations_sep(self, impexp: ImportExport) -> None:
        """
        ContentPages in multiple languages (with globally-unique slugs and titles) are
        preserved across export/import with each language imported separately.
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # English pages
        home_en = HomePage.objects.get(locale__language_code="en")
        _app_rem = PageBuilder.build_cpi(
            parent=home_en, slug="appointment-reminders", title="Appointment reminders"
        )
        _sbm = PageBuilder.build_cpi(
            parent=home_en, slug="stage-based-messages", title="Stage-based messages"
        )
        _him = PageBuilder.build_cpi(
            parent=home_en, slug="health-info-messages", title="Health info messages"
        )
        _wtt = PageBuilder.build_cpi(
            parent=home_en,
            slug="whatsapp-template-testing",
            title="whatsapp template testing",
        )
        imp_exp = PageBuilder.build_cpi(
            parent=home_en, slug="import-export", title="Import Export"
        )
        non_templ_wablks = [
            WABlk("this is a non template message"),
            WABlk("this message has a doc"),
            WABlk("this message comes with audio"),
        ]
        non_tmpl = PageBuilder.build_cp(
            parent=imp_exp,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks)],
        )

        # Portuguese pages
        home_pt = HomePage.objects.get(locale__language_code="pt")
        imp_exp_pt = PageBuilder.build_cpi(
            parent=home_pt,
            slug="import-export-pt",
            title="Import Export (pt)",
            translated_from=imp_exp,
        )
        non_templ_wablks_pt = [
            WABlk("this is a non template message (pt)"),
            WABlk("this message has a doc (pt)"),
            WABlk("this message comes with audio (pt)"),
        ]
        non_tmpl_pt = PageBuilder.build_cp(
            parent=imp_exp_pt,
            slug="non-template-pt",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks_pt)],
            translated_from=non_tmpl,
        )

        assert imp_exp.translation_key == imp_exp_pt.translation_key
        assert non_tmpl.translation_key == non_tmpl_pt.translation_key

        orig = impexp.get_page_json()
        content_en = impexp.export_content(locale="en")
        content_pt = impexp.export_content(locale="pt")

        impexp.import_content(content_en, locale="en")
        impexp.import_content(content_pt, locale="pt", purge=False)
        imported = impexp.get_page_json()
        assert imported == orig

    def test_translations_split(self, impexp: ImportExport) -> None:
        """
        ContentPages in multiple languages (with unique-per-locale slugs and
        titles) are preserved across export/import with each language imported
        separately.
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # English pages
        home_en = HomePage.objects.get(locale__language_code="en")
        _app_rem = PageBuilder.build_cpi(
            parent=home_en, slug="appointment-reminders", title="Appointment reminders"
        )
        _sbm = PageBuilder.build_cpi(
            parent=home_en, slug="stage-based-messages", title="Stage-based messages"
        )
        _him = PageBuilder.build_cpi(
            parent=home_en, slug="health-info-messages", title="Health info messages"
        )
        _wtt = PageBuilder.build_cpi(
            parent=home_en,
            slug="whatsapp-template-testing",
            title="whatsapp template testing",
        )
        imp_exp = PageBuilder.build_cpi(
            parent=home_en, slug="import-export", title="Import Export"
        )
        non_templ_wablks = [
            WABlk("this is a non template message"),
            WABlk("this message has a doc"),
            WABlk("this message comes with audio"),
        ]
        non_tmpl = PageBuilder.build_cp(
            parent=imp_exp,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks)],
        )

        # Portuguese pages
        home_pt = HomePage.objects.get(locale__language_code="pt")
        imp_exp_pt = PageBuilder.build_cpi(
            parent=home_pt,
            slug="import-export",
            title="Import Export (pt)",
            translated_from=imp_exp,
        )
        non_templ_wablks_pt = [
            WABlk("this is a non template message (pt)"),
            WABlk("this message has a doc (pt)"),
            WABlk("this message comes with audio (pt)"),
        ]
        non_tmpl_pt = PageBuilder.build_cp(
            parent=imp_exp_pt,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks_pt)],
            translated_from=non_tmpl,
        )

        assert imp_exp.translation_key == imp_exp_pt.translation_key
        assert non_tmpl.translation_key == non_tmpl_pt.translation_key

        orig = impexp.get_page_json()
        content_en = impexp.export_content(locale="en")
        content_pt = impexp.export_content(locale="pt")

        impexp.import_content(content_en, locale="en")
        impexp.import_content(content_pt, locale="pt", purge=False)
        imported = impexp.get_page_json()
        assert imported == orig

    def test_translations_en(self, impexp: ImportExport) -> None:
        """
        ContentPages in multiple languages are that are imported with a locale
        specified have pages in that locale preserved and all other locales are
        removed.
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # English pages
        home_en = HomePage.objects.get(locale__language_code="en")
        _app_rem = PageBuilder.build_cpi(
            parent=home_en, slug="appointment-reminders", title="Appointment reminders"
        )
        _sbm = PageBuilder.build_cpi(
            parent=home_en, slug="stage-based-messages", title="Stage-based messages"
        )
        _him = PageBuilder.build_cpi(
            parent=home_en, slug="health-info-messages", title="Health info messages"
        )
        _wtt = PageBuilder.build_cpi(
            parent=home_en,
            slug="whatsapp-template-testing",
            title="whatsapp template testing",
        )
        imp_exp = PageBuilder.build_cpi(
            parent=home_en, slug="import-export", title="Import Export"
        )
        non_templ_wablks = [
            WABlk("this is a non template message"),
            WABlk("this message has a doc"),
            WABlk("this message comes with audio"),
        ]
        non_tmpl = PageBuilder.build_cp(
            parent=imp_exp,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks)],
        )

        # Portuguese pages
        home_pt = HomePage.objects.get(locale__language_code="pt")
        imp_exp_pt = PageBuilder.build_cpi(
            parent=home_pt,
            slug="import-export-pt",
            title="Import Export (pt)",
            translated_from=imp_exp,
        )
        non_templ_wablks_pt = [
            WABlk("this is a non template message (pt)"),
            WABlk("this message has a doc (pt)"),
            WABlk("this message comes with audio (pt)"),
        ]
        non_tmpl_pt = PageBuilder.build_cp(
            parent=imp_exp_pt,
            slug="non-template-pt",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks_pt)],
            translated_from=non_tmpl,
        )

        assert imp_exp.translation_key == imp_exp_pt.translation_key
        assert non_tmpl.translation_key == non_tmpl_pt.translation_key

        orig_en = impexp.get_page_json(locale="en")
        content = impexp.export_content()
        impexp.import_content(content, locale="en")
        imported = impexp.get_page_json()
        assert imported == orig_en

    def test_footer(self, impexp: ImportExport) -> None:
        """
        ContentPages with footer in whatsapp messages are preserved
        across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")])],
        )

        footer = "Test footer"
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody(
                    "health info",
                    [WABlk("*Health information* WA", footer=footer)],
                )
            ],
            whatsapp_template_name="template-health-info",
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_example_values(self, impexp: ImportExport) -> None:
        """
        ContentPages with example values in whatsapp messages are preserved
        across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")])],
        )

        example_values = ["Example value 1", "Example value 2"]
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody(
                    "health info",
                    [WABlk("*Health information* WA", example_values=example_values)],
                )
            ],
            whatsapp_template_name="template-health-info",
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_export_missing_related_page(self, impexp: ImportExport) -> None:
        """
        If a page has a related page that no longer exists, the missing related
        page is skipped during export.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )
        health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[WABody("health info", [WABlk("*Health information*")])],
        )
        self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[WABody("self-help", [WABlk("*Self-help programs*")])],
        )
        health_info = PageBuilder.link_related(health_info, [self_help])
        # This is what we expect to see after export/import, so we fetch the
        # JSON to compare with before adding the missing related page.
        orig_without_self_help = impexp.get_page_json()

        move_along = PageBuilder.build_cp(
            parent=ha_menu,
            slug="move-along",
            title="move-along",
            bodies=[WABody("move along", [WABlk("*Nothing to see here*")])],
        )
        PageBuilder.link_related(health_info, [move_along])
        move_along.delete()

        # Ideally, all related page links would be removed when the page they
        # link to is deleted. We don't currently do that, so for now we just
        # make sure that we skip such links during export.
        sh_page = Page.objects.get(pk=self_help.id)
        assert [rp.value for rp in health_info.related_pages] == [sh_page, None]

        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig_without_self_help

    def test_ussd_values(self, impexp: ImportExport) -> None:
        """
        ContentPages with USSD messages are preserved
        across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[UBody("HealthAlert menu", [UBlk("*Welcome to HealthAlert* USSD")])],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_sms_values(self, impexp: ImportExport) -> None:
        """
        ContentPages with SMS messages are preserved
        across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[SBody("HealthAlert menu", [SBlk("*Welcome to HealthAlert* SMS")])],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_list_items(self, impexp: ImportExport) -> None:
        """
        ContentPages with list items in whatsapp messages are preserved
        across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")])],
        )

        form = Assessment.objects.create(
            title="Test form", slug="test-form", locale=home_page.locale
        )
        list_items = [
            NextListItem("Item 1"),
            PageListItem("Item 2", page=ha_menu),
            FormListItem("Item 3", form=form),
        ]
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody(
                    "health info",
                    [WABlk("*Health information* WA", list_items=list_items)],
                )
            ],
            whatsapp_template_name="template-health-info",
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_export_import_page_with_go_to_button(self, impexp: ImportExport) -> None:
        """
        If pages linked to another page with go to button are not deleted, all buttons will be exported to a file.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        imp_exp = PageBuilder.build_cpi(home_page, "import-export", "Import Export")

        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )
        PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody(
                    "health info",
                    [
                        WABlk(
                            "*Health information*",
                            buttons=[NextBtn("Next message button")],
                        )
                    ],
                )
            ],
        )

        _self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[
                WABody(
                    "self-help",
                    [
                        WABlk(
                            "*Self-help programs*",
                            buttons=[PageBtn("Import Export", page=imp_exp)],
                        )
                    ],
                )
            ],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_export_missing_go_to_button(self, impexp: ImportExport) -> None:
        """
        If a page has a button go to page that no longer exists, the missing button
        is skipped during export.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )

        first_page = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody(
                    "health info",
                    [
                        WABlk(
                            "*Health information*",
                            buttons=[NextBtn("Next message button")],
                        )
                    ],
                )
            ],
        )

        orig_json = impexp.get_page_json()

        index = PageBuilder.build_cpi(home_page, "import-export", "Import Export")

        # Add another button to existing page (first_page)
        add_go_to_page_button(
            first_page.whatsapp_body[0], PageBtn("Go to Btn_2", page=index)
        )

        first_page.save()
        rev = first_page.save_revision()
        rev.publish()
        first_page.refresh_from_db()

        # Delete page linked to got to button
        index.delete()

        impexp.export_reimport()
        updated_json = impexp.get_page_json()

        assert orig_json == updated_json

    def test_buttons(self, impexp: ImportExport) -> None:
        """
        Content page buttons should import and export
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        form = Assessment.objects.create(
            title="Test form", slug="test-form", locale=home_page.locale
        )

        target_page = PageBuilder.build_cp(
            parent=main_menu,
            slug="target_page",
            title="Target page",
            bodies=[WABody("Target", [WABlk("Target page")])],
        )

        PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody(
                    "HealthAlert menu",
                    [
                        WABlk(
                            "*Welcome to HealthAlert*",
                            buttons=[
                                NextBtn("Go to next page"),
                                PageBtn("Go to page", page=target_page),
                                FormBtn("Start form", form=form),
                            ],
                        )
                    ],
                )
            ],
        )
        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_export_import_page_with_missing_form_button(
        self, impexp: ImportExport
    ) -> None:
        """
        If a go_to_form button links to a delete form, it should be excluded from exports
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        page = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )
        orig = impexp.get_page_json()

        # Add another button to the page, then delete the form it links to
        form = Assessment.objects.create(
            title="Test form", slug="test-form", locale=home_page.locale
        )
        add_go_to_form_button(page.whatsapp_body[0], FormBtn("Go to Btn_2", form=form))
        form.delete()

        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_export_import_page_with_missing_form_list(
        self, impexp: ImportExport
    ) -> None:
        """
        If a go_to_form list item links to a delete form, it should be excluded from exports
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")

        page = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert*")])],
        )
        orig = impexp.get_page_json()

        # Add another button to the page, then delete the form it links to
        form = Assessment.objects.create(
            title="Test form", slug="test-form", locale=home_page.locale
        )
        add_go_to_form_list_item(
            page.whatsapp_body[0], FormListItem("Go to Btn_2", form=form)
        )
        form.delete()

        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig
