import contextlib
import csv
from collections import defaultdict
from dataclasses import dataclass, field, fields
from datetime import datetime
from io import BytesIO, StringIO
from queue import Queue
from typing import Any

from django.core.exceptions import ObjectDoesNotExist, ValidationError  # type: ignore
from openpyxl import load_workbook
from taggit.models import Tag  # type: ignore
from treebeard.exceptions import NodeAlreadySaved  # type: ignore
from wagtail.blocks import (  # type: ignore
    StreamValue,
    StructValue,
)
from wagtail.coreutils import get_content_languages  # type: ignore
from wagtail.models import Locale, Page  # type: ignore
from wagtail.models.sites import Site  # type: ignore

from home.models import Assessment, ContentPage, HomePage  # type: ignore

PageId = tuple[str, Locale]


class ImportAssessmentException(Exception):
    """
    Base exception for all import related issues.
    """

    def __init__(self, message: str, row_num: int | None = None):
        self.row_num = row_num
        self.message = message
        super().__init__()


class ContentImporter:
    def __init__(
        self,
        file_content: bytes,
        file_type: str,
        progress_queue: Queue[int],
        purge: bool | str = True,
        locale: Locale | str | None = None,
    ):
        if isinstance(locale, str):
            locale = Locale.objects.get(language_code=locale)
        self.file_content = file_content
        self.file_type = file_type
        self.progress_queue = progress_queue
        self.purge = purge in ["True", "yes", True]
        self.locale = locale
        self.locale_map: dict[str, Locale] = {}
        self.shadow_pages: dict[PageId, ShadowAssessmentPage] = {}

    def locale_from_language_code(self, lang_code_entry: str) -> Locale:
        if lang_code_entry not in self.locale_map:
            codes = []
            lang_name = ""
            for lang_code, lang_dn in get_content_languages().items():
                if lang_code == lang_code_entry:
                    lang_name = lang_dn
                    codes.append(lang_code)
            if not codes:
                raise ImportAssessmentException(
                    f"Language code not found: {lang_code_entry}"
                )
            if len(codes) > 1:
                raise ImportAssessmentException(
                    f"Multiple codes for language: {lang_name} -> {codes}"
                )
            self.locale_map[lang_code_entry] = Locale.objects.get(
                language_code=codes[0]
            )
        return self.locale_map[lang_code_entry]

    def perform_import(self) -> None:
        rows = self.parse_file()
        self.set_progress("Loaded file", 5)

        if self.purge:
            self.delete_existing_content()
        self.set_progress("Deleted existing assessment", 10)

        self.process_rows(rows)
        self.save_pages_assessment()

    def process_rows(self, rows: list["ContentRow"]) -> None:
        for i, row in enumerate(rows, start=2):
            try:
                self.create_shadow_assessment_page_from_row(row, i)
            except ImportAssessmentException as e:
                e.row_num = i
                raise e

    def save_pages_assessment(self) -> None:
        for i, page in enumerate(reversed(self.shadow_pages.values())):
            parent = self.home_page(page.locale)
            page.save(parent)
            self.set_progress("Importing pages", 10 + 70 * i // len(self.shadow_pages))

    def parse_file(self) -> list["ContentRow"]:
        if self.file_type == "XLSX":
            return self.parse_excel()
        return self.parse_csv()

    def parse_excel(self) -> list["ContentRow"]:
        workbook = load_workbook(BytesIO(self.file_content), read_only=True)
        worksheet = workbook.worksheets[0]

        def clean_excel_cell(cell_value: str | float | datetime | None) -> str:
            return str(cell_value).replace("_x000D", "")

        first_row = next(worksheet.iter_rows(max_row=1, values_only=True))
        header = [clean_excel_cell(cell) if cell else None for cell in first_row]
        rows: list[ContentRow] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            r = {}
            for name, cell in zip(header, row):  # noqa: B905 (TODO: strict?)
                if name and cell:
                    r[name] = clean_excel_cell(cell)
            rows.append(ContentRow.from_flat(r))
        page_id_rows = group_rows_by_page_id(rows)
        return page_id_rows

    def parse_csv(self) -> list["ContentRow"]:
        reader = csv.DictReader(StringIO(self.file_content.decode()))
        rows = [ContentRow.from_flat(row) for row in reader]
        page_id_rows = group_rows_by_page_id(rows)
        return page_id_rows

    def set_progress(self, message: str, progress: int) -> None:
        self.progress_queue.put_nowait(progress)

    def delete_existing_content(self) -> None:
        Assessment.objects.all().delete()

    def home_page(self, locale: Locale) -> HomePage:
        try:
            return HomePage.objects.get(locale=locale)
        except ObjectDoesNotExist:
            raise ImportAssessmentException(
                f"You are trying to add a child page to a '{locale}' HomePage that does not exist. Please create the '{locale}' HomePage first"
            )

    def default_locale(self) -> Locale:
        site = Site.objects.get(is_default_site=True)
        return site.root_page.locale

    def create_shadow_assessment_page_from_row(
        self, row: "ContentRow", row_num: int
    ) -> None:
        locale = self.locale_from_language_code(row.locale)
        page = ShadowAssessmentPage(
            row_num=row_num,
            slug=row.slug,
            title=row.title,
            locale=locale,
            high_result_page=row.high_result_page,
            medium_result_page=row.medium_result_page,
            low_result_page=row.low_result_page,
            high_inflection=row.high_inflection,
            medium_inflection=row.medium_inflection,
            low_inflection=row.low_inflection,
            answers=row.answers,
            scores=row.scores,
            questions=row.question,
            generic_error=row.generic_error,
            error=row.error,
            tags=row.tags,
        )

        self.shadow_pages[(row.slug, locale)] = page


# Since wagtail page models are so difficult to create and modify programatically,
# we instead store all the pages as these shadow objects, which we can then at the end
# do a single write to the database from, and encapsulate all that complexity
@dataclass(slots=True)
class ShadowAssessmentPage:
    slug: str
    locale: Locale
    row_num: int
    title: str = ""
    tags: list[str] = field(default_factory=list)
    high_result_page: list[str] = field(default_factory=list)
    medium_result_page: list[str] = field(default_factory=list)
    low_result_page: list[str] = field(default_factory=list)
    high_inflection: str = ""
    medium_inflection: str = ""
    low_inflection: str = ""
    generic_error: str = ""
    questions: list[str] | list[dict[str, Any]] | str = field(default_factory=list)
    error: str = ""
    answers: list[str] = field(default_factory=list)
    scores: list[str] = field(default_factory=list)

    def save(self, parent: Page) -> None:
        try:
            page = Assessment.objects.get(slug=self.slug, locale=self.locale)
        except Assessment.DoesNotExist:
            page = Assessment(slug=self.slug, locale=self.locale)
        self.add_identifiers_to_page(page)
        self.add_tags_to_page(page)
        try:
            with contextlib.suppress(NodeAlreadySaved):
                parent.add_child(instance=page)
            page.save_revision().publish()
        except ValidationError as err:
            raise ImportAssessmentException(f"Validation error: {err}", self.row_num)

    def add_identifiers_to_page(self, page: Assessment) -> None:
        page.slug = self.slug
        page.title = self.title
        page.locale = self.locale
        page.high_inflection = self.high_inflection
        page.high_result_page_id = result_page_get_id_from_slug(self.high_result_page)
        page.medium_inflection = self.medium_inflection
        page.medium_result_page_id = result_page_get_id_from_slug(
            self.medium_result_page
        )
        page.low_inflection = self.low_inflection
        page.low_result_page_id = result_page_get_id_from_slug(self.low_result_page)
        page.generic_error = self.generic_error
        page.questions = create_question_streamfield(self.questions)

    def add_tags_to_page(self, page: Assessment) -> None:
        page.tags.clear()
        for tag_name in self.tags:
            tag, _ = Tag.objects.get_or_create(name=tag_name)
            page.tags.add(tag)


@dataclass(slots=True, frozen=True)
class ContentRow:
    slug: str
    title: str = ""
    page_id: int | None = None
    tags: list[str] = field(default_factory=list)
    locale: str = ""
    high_result_page: list[str] = field(default_factory=list)
    high_inflection: str = ""
    medium_result_page: list[str] = field(default_factory=list)
    medium_inflection: str = ""
    low_result_page: list[str] = field(default_factory=list)
    low_inflection: str = ""
    generic_error: str = ""
    question_count: str = ""
    question: list[str] | list[dict[str, Any]] | str = field(default_factory=list)
    error: str = ""
    answers: list[str] = field(default_factory=list)
    scores: list[str] = field(default_factory=list)

    @classmethod
    def from_flat(cls, row: dict[str, str]) -> "ContentRow":
        class_fields = {field.name for field in fields(cls)}
        row = {
            key.strip(): value.strip()
            for key, value in row.items()
            if value and key in class_fields
        }
        try:
            return cls(
                tags=deserialise_list(row.pop("tags", "")),
                question=[row.pop("question")] if row.get("question") else "",
                answers=deserialise_list(row.pop("answers", "")),
                scores=deserialise_list(row.pop("scores", "")),
                **row,  # type: ignore
            )
        except TypeError:
            raise ImportAssessmentException(
                "The import file is missing some required fields."
            )


def group_rows_by_page_id(rows: list[ContentRow]) -> list[ContentRow]:
    grouped_rows = defaultdict(list)

    for row in rows:
        grouped_rows[row.slug].append(row)
    output_rows = []
    for _slug, grouped_rows_list in grouped_rows.items():
        questions = []
        for grouped_row in grouped_rows_list:
            for _, question in enumerate(grouped_row.question):
                answers = grouped_row.answers if grouped_row.answers else []
                scores = grouped_row.scores if grouped_row.scores else []
                error = grouped_row.error if grouped_row.error else []
                question_entry = {
                    "question": question,
                    "answers": [
                        {"answer": ans.strip(), "score": sc.strip()}
                        for ans, sc in zip(answers, scores, strict=False)
                    ],  # Strip whitespace
                    "error": error,
                }
                questions.append(question_entry)

        output_rows.append(
            ContentRow(
                slug=grouped_rows_list[0].slug,
                title=grouped_rows_list[0].title,
                tags=grouped_rows_list[0].tags,
                locale=grouped_rows_list[0].locale,
                high_result_page=grouped_rows_list[0].high_result_page,
                high_inflection=grouped_rows_list[0].high_inflection,
                medium_result_page=grouped_rows_list[0].medium_result_page,
                medium_inflection=grouped_rows_list[0].medium_inflection,
                low_result_page=grouped_rows_list[0].low_result_page,
                low_inflection=grouped_rows_list[0].low_inflection,
                generic_error=grouped_rows_list[0].generic_error,
                question_count=str(len(questions)),
                question=questions,
                answers=[],
                scores=[],
                error="",
            )
        )

    return output_rows


def result_page_get_id_from_slug(slug: list[str]) -> int:
    try:
        page = ContentPage.objects.get(slug=slug)
    except ObjectDoesNotExist:
        raise ImportAssessmentException(
            f"You are trying to add an assessment, where one of the result pages "
            f"references the content page with slug {slug} which does not exist. "
            "Please create the content page first."
        )
    return page.id


def create_question_streamfield(questions: StreamValue) -> list[StructValue]:
    stream_data = []
    for question_data in questions:
        answers = [
            {"answer": ans_data["answer"], "score": float(ans_data["score"])}
            for ans_data in question_data["answers"]
        ]
        error = question_data["error"]
        stream_data.append(
            {
                "type": "question",
                "value": {
                    "question": question_data["question"],
                    "answers": answers,
                    "error": error,
                },
            }
        )
    return stream_data


def deserialise_list(value: str) -> list[str]:
    items = next(csv.reader([value]))
    return [item.strip() for item in items]
