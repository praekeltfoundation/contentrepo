import contextlib
import csv
import json
from collections import defaultdict
from dataclasses import dataclass, field, fields
from datetime import datetime
from io import BytesIO, StringIO
from queue import Queue
from typing import Any
from uuid import uuid4

from django.core.exceptions import ObjectDoesNotExist, ValidationError  # type: ignore
from django.forms import model_to_dict  # type: ignore
from openpyxl import load_workbook
from taggit.models import Tag  # type: ignore
from treebeard.exceptions import NodeAlreadySaved  # type: ignore
from wagtail.blocks import StreamValue, StructValue, StreamBlockValidationError  # type: ignore
from wagtail.blocks.list_block import ListValue  # type: ignore
from wagtail.coreutils import get_content_languages  # type: ignore
from wagtail.models import Locale, Page  # type: ignore
from wagtail.models.sites import Site  # type: ignore
from wagtail.rich_text import RichText  # type: ignore
from wagtail.test.utils.form_data import nested_form_data, streamfield  # type: ignore
from wagtail.admin.rich_text.converters.contentstate import ContentstateConverter  # type: ignore
from wagtail import blocks


from home.models import (
    ContentPage,
    ContentPageIndex,
    ContentQuickReply,
    ContentTrigger,
    HomePage,
    MessengerBlock,
    SMSBlock,
    USSDBlock,
    ViberBlock,
    WhatsappBlock,
)

PageId = tuple[str, Locale]


class ImportException(Exception):
    """
    Base exception for all import related issues.
    """

    def __init__(
        self,
        message: str,
        row_num: int | None = None,
        slug: str | None = None,
        locale: Locale | None = None,
    ):
        self.row_num = row_num
        self.message = message
        self.slug = slug
        self.locale = locale
        super().__init__()


class ContentImporter:
    def __init__(
        self,
        file_content: bytes,
        file_type: str,
        progress_queue: Queue[int],
        purge: bool | str = True,
        locale: Locale | str | None = None,
    ):
        if isinstance(locale, str):
            locale = Locale.objects.get(language_code=locale)
        self.file_content = file_content
        self.file_type = file_type
        self.progress_queue = progress_queue
        self.purge = purge in ["True", "yes", True]
        self.locale = locale
        self.locale_map: dict[str, Locale] = {}
        self.shadow_pages: dict[PageId, ShadowContentPage] = {}
        self.go_to_page_buttons: dict[PageId, dict[int, list[dict[str, Any]]]] = (
            defaultdict(lambda: defaultdict(list))
        )

    def locale_from_display_name(self, langname: str) -> Locale:
        if langname not in self.locale_map:
            codes = []
            for lang_code, lang_dn in get_content_languages().items():
                if lang_dn == langname:
                    codes.append(lang_code)
            if not codes:
                raise ImportException(f"Language not found: {langname}")
            if len(codes) > 1:
                raise ImportException(
                    f"Multiple codes for language: {langname} -> {codes}"
                )
            self.locale_map[langname] = Locale.objects.get(language_code=codes[0])
        return self.locale_map[langname]

    def perform_import(self) -> None:
        rows = self.parse_file()
        self.set_progress("Loaded file", 5)

        if self.purge:
            self.delete_existing_content()
        self.set_progress("Deleted existing content", 10)

        self.process_rows(rows)
        self.save_pages()
        self.link_related_pages()
        self.add_go_to_page_buttons()

    def process_rows(self, rows: list["ContentRow"]) -> None:
        # Non-page rows don't have a locale, so we need to remember the last
        # row that does have a locale.
        prev_locale: Locale | None = None
        for i, row in enumerate(rows, start=2):
            try:
                if row.is_page_index:
                    if self.locale and row.locale != self.locale.get_display_name():
                        # This page index isn't for the locale we're importing, so skip it.
                        continue
                    self.create_content_page_index_from_row(row)
                    prev_locale = self.locale_from_display_name(row.locale)
                elif row.is_content_page:
                    self.create_shadow_content_page_from_row(row, i)
                    prev_locale = self.locale_from_display_name(row.locale)
                elif row.is_variation_message:
                    self.add_variation_to_shadow_content_page_from_row(row, prev_locale)
                else:
                    self.add_message_to_shadow_content_page_from_row(row, prev_locale)
            except ImportException as e:
                e.row_num = i
                e.slug = row.slug
                e.locale = row.locale
                raise e

    def save_pages(self) -> None:
        for i, page in enumerate(self.shadow_pages.values()):
            if self.locale and page.locale != self.locale:
                # This page isn't for the locale we're importing, so skip it.
                continue
            if page.parent:
                # TODO: We should need to use something unique for `parent`
                try:
                    parent = Page.objects.get(title=page.parent, locale=page.locale)
                except Page.DoesNotExist:
                    raise ImportException(
                        f"Cannot find parent page with title '{page.parent}' and "
                        f"locale '{page.locale}'",
                        page.row_num,
                    )
                except Page.MultipleObjectsReturned:
                    parents = Page.objects.filter(
                        title=page.parent, locale=page.locale
                    ).values_list("slug", flat=True)
                    raise ImportException(
                        f"Multiple pages with title '{page.parent}' and locale "
                        f"'{page.locale}' for parent page: {list(parents)}",
                        page.row_num,
                    )

                try:
                    child = Page.objects.get(slug=page.slug, locale=page.locale)
                except Page.DoesNotExist:
                    # Nothing to check if the child doesn't exist yet.
                    pass
                else:
                    if child.get_parent().title != page.parent:
                        raise ImportException(
                            f"Changing the parent from '{child.get_parent()}' to '{page.parent}' "
                            f"for the page with title '{page.title}' during import is not allowed. Please use the UI",
                            page.row_num,
                        )

            else:
                parent = self.home_page(page.locale)
            page.save(parent)
            self.set_progress("Importing pages", 10 + 70 * i // len(self.shadow_pages))

    def link_related_pages(self) -> None:
        for i, page in enumerate(self.shadow_pages.values()):
            if page.related_pages:
                page.link_related_pages()
            self.set_progress(
                "Linking related pages", 80 + 10 * i // len(self.shadow_pages)
            )

    def add_go_to_page_buttons(self) -> None:
        for (slug, locale), messages in self.go_to_page_buttons.items():
            page = ContentPage.objects.get(slug=slug, locale=locale)
            for message_index, buttons in messages.items():
                for button in buttons:
                    title = button["title"]
                    try:
                        related_page = Page.objects.get(
                            slug=button["slug"], locale=locale
                        )
                    except Page.DoesNotExist:
                        row = self.shadow_pages[(slug, locale)]
                        raise ImportException(
                            f"No pages found with slug '{button['slug']}' and locale "
                            f"'{locale}' for go_to_page button '{button['title']}' on "
                            f"page '{slug}'",
                            row.row_num,
                        )
                    page.whatsapp_body[message_index].value["buttons"].append(
                        ("go_to_page", {"page": related_page, "title": title})
                    )
            page.save_revision().publish()

    def parse_file(self) -> list["ContentRow"]:
        if self.file_type == "XLSX":
            return self.parse_excel()
        return self.parse_csv()

    def parse_excel(self) -> list["ContentRow"]:
        workbook = load_workbook(BytesIO(self.file_content), read_only=True)
        worksheet = workbook.worksheets[0]

        def clean_excel_cell(cell_value: str | float | datetime | None) -> str:
            return str(cell_value).replace("_x000D", "")

        first_row = next(worksheet.iter_rows(max_row=1, values_only=True))
        header = [clean_excel_cell(cell) if cell else None for cell in first_row]
        rows: list[ContentRow] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            r = {}
            for name, cell in zip(header, row):  # noqa: B905 (TODO: strict?)
                if name and cell:
                    r[name] = clean_excel_cell(cell)
            rows.append(ContentRow.from_flat(r))
        return rows

    def parse_csv(self) -> list["ContentRow"]:
        reader = csv.DictReader(StringIO(self.file_content.decode()))
        return [ContentRow.from_flat(row) for row in reader]

    def set_progress(self, message: str, progress: int) -> None:
        self.progress_queue.put_nowait(progress)

    def delete_existing_content(self) -> None:
        ContentPage.objects.all().delete()
        ContentPageIndex.objects.all().delete()

    def home_page(self, locale: Locale) -> HomePage:
        try:
            return HomePage.objects.get(locale=locale)
        except ObjectDoesNotExist:
            raise ImportException(
                f"You are trying to add a child page to a '{locale}' HomePage that does not exist. Please create the '{locale}' HomePage first"
            )

    def default_locale(self) -> Locale:
        site = Site.objects.get(is_default_site=True)
        return site.root_page.locale

    def create_content_page_index_from_row(self, row: "ContentRow") -> None:
        locale = self.locale_from_display_name(row.locale)
        try:
            index = ContentPageIndex.objects.get(slug=row.slug, locale=locale)
        except ContentPageIndex.DoesNotExist:
            index = ContentPageIndex(slug=row.slug, locale=locale)
        index.title = row.web_title
        # Translation keys are required for pages with a non-default locale,
        # but optional for the default locale.
        if row.translation_tag or locale != self.default_locale():
            index.translation_key = row.translation_tag
        locale = self.locale_from_display_name(row.locale)
        try:
            with contextlib.suppress(NodeAlreadySaved):
                self.home_page(locale).add_child(instance=index)
            index.save_revision().publish()
        except ValidationError as err:
            # FIXME: Find a better way to represent this.
            raise ImportException(f"Validation error: {err}")

    def create_shadow_content_page_from_row(
        self, row: "ContentRow", row_num: int
    ) -> None:
        locale = self.locale_from_display_name(row.locale)
        page = ShadowContentPage(
            row_num=row_num,
            slug=row.slug,
            title=row.web_title,
            locale=locale,
            subtitle=row.web_subtitle,
            body=row.web_body,
            enable_web=bool(row.web_body),
            tags=row.tags,
            quick_replies=row.quick_replies,
            triggers=row.triggers,
            parent=row.parent,
            related_pages=row.related_pages,
        )

        self.shadow_pages[(row.slug, locale)] = page

        self.add_message_to_shadow_content_page_from_row(row, locale)

        if row.is_whatsapp_message:
            page.whatsapp_title = row.whatsapp_title
            if row.is_whatsapp_template_message:
                page.is_whatsapp_template = True
                page.whatsapp_template_name = row.whatsapp_template_name
                page.whatsapp_template_category = row.whatsapp_template_category
            # TODO: Media
            # Currently media is exported as a URL, which just has an ID. This doesn't
            # actually help us much, as IDs can differ between instances, so we need
            # a better way of exporting and importing media here

        if row.is_sms_message:
            page.sms_title = row.sms_title

        if row.is_ussd_message:
            page.ussd_title = row.ussd_title

        if row.is_messenger_message:
            page.messenger_title = row.messenger_title

        if row.is_viber_message:
            page.viber_title = row.viber_title

        # Translation keys are required for pages with a non-default locale,
        # but optional for the default locale.
        if row.translation_tag or locale != self.default_locale():
            page.translation_key = row.translation_tag

    def add_variation_to_shadow_content_page_from_row(
        self, row: "ContentRow", locale: Locale
    ) -> None:
        try:
            page = self.shadow_pages[(row.slug, locale)]
        except KeyError:
            raise ImportException(
                f"This is a variation for the content page with slug '{row.slug}' and locale '{locale}', but no such page exists"
            )
        whatsapp_block = page.whatsapp_body[-1]
        whatsapp_block.variation_messages.append(
            ShadowVariationBlock(
                message=row.variation_body, variation_restrictions=row.variation_title
            )
        )

    def add_message_to_shadow_content_page_from_row(
        self, row: "ContentRow", locale: Locale
    ) -> None:
        try:
            page = self.shadow_pages[(row.slug, locale)]
        except KeyError:
            raise ImportException(
                f"This is a message for page with slug '{row.slug}' and locale '{locale}', but no such page exists"
            )
        if row.is_whatsapp_message:
            page.enable_whatsapp = True
            buttons = []
            for button in row.buttons:
                if button["type"] == "next_message":
                    buttons.append(
                        {
                            "id": str(uuid4()),
                            "type": button["type"],
                            "value": {"title": button["title"]},
                        }
                    )
                elif button["type"] == "go_to_page":
                    page_gtps = self.go_to_page_buttons[(row.slug, locale)]
                    page_gtps[len(page.whatsapp_body)].append(button)
            page.whatsapp_body.append(
                ShadowWhatsappBlock(
                    message=row.whatsapp_body,
                    next_prompt=row.next_prompt,
                    example_values=row.example_values,
                    buttons=buttons,
                    footer=row.footer,
                    list_items=row.list_items,
                )
            )
        if row.is_sms_message:
            page.enable_sms = True
            page.sms_body.append(ShadowSMSBlock(message=row.sms_body))

        if row.is_ussd_message:
            page.enable_ussd = True
            page.ussd_body.append(ShadowUSSDBlock(message=row.ussd_body))

        if row.is_messenger_message:
            page.enable_messenger = True
            page.messenger_body.append(ShadowMessengerBlock(message=row.messenger_body))

        if row.is_viber_message:
            page.enable_viber = True
            page.viber_body.append(ShadowViberBlock(message=row.viber_body))


def wagtail_to_formdata(val: Any) -> Any:
    """
    Convert a model dict field that may be a nested streamfield (or associated
    type) into something we can turn into form data.
    """
    match val:
        case StreamValue():  # type: ignore[misc] # No type info
            return streamfield([(b.block_type, wagtail_to_formdata(b.value)) for b in val])
        case StructValue():  # type: ignore[misc] # No type info
            return {k: wagtail_to_formdata(v) for k, v in val.items()}
        case ListValue():  # type: ignore[misc] # No type info
            # Wagtail doesn't have an equivalent of streamfield() for
            # listvalue, so we have to do it by hand.
            list_val: dict[str, Any] = {
                str(i): {"deleted": "", "order": str(i), "value": wagtail_to_formdata(v)}
                for i, v in enumerate(val)
            }
            list_val["count"] = str(len(val))
            return list_val
        case RichText():  # type: ignore[misc] # No type info
            # FIXME: The only RichTextBlock() we currently have is in the web
            #        body and we don't appear to do any validation on it.
            #        There's probably a better way to convert and/or ignore these.
            return ContentstateConverter([]).from_database_format(val.source)
        case _:
            return val


# Since wagtail page models are so difficult to create and modify programatically,
# we instead store all the pages as these shadow objects, which we can then at the end
# do a single write to the database from, and encapsulate all that complexity
@dataclass(slots=True)
class ShadowContentPage:
    slug: str
    parent: str
    locale: Locale
    row_num: int
    enable_web: bool = False
    title: str = ""
    subtitle: str = ""
    body: str = ""
    enable_whatsapp: bool = False
    is_whatsapp_template: bool = False
    whatsapp_title: str = ""
    whatsapp_body: list["ShadowWhatsappBlock"] = field(default_factory=list)
    whatsapp_template_name: str = ""
    whatsapp_template_category: str = "UTILITY"
    enable_sms: bool = False
    sms_title: str = ""
    sms_body: list["ShadowSMSBlock"] = field(default_factory=list)
    enable_ussd: bool = False
    ussd_title: str = ""
    ussd_body: list["ShadowUSSDBlock"] = field(default_factory=list)
    enable_messenger: bool = False
    messenger_title: str = ""
    messenger_body: list["ShadowMessengerBlock"] = field(default_factory=list)
    enable_viber: bool = False
    viber_title: str = ""
    viber_body: list["ShadowViberBlock"] = field(default_factory=list)
    translation_key: str | None = None
    tags: list[str] = field(default_factory=list)
    quick_replies: list[str] = field(default_factory=list)
    triggers: list[str] = field(default_factory=list)
    related_pages: list[str] = field(default_factory=list)

    def validate_page_using_form(self, page: Page) -> None:
        edit_handler = page.edit_handler.bind_to_model(ContentPage)
        form_class = edit_handler.get_form_class()

        form_data = nested_form_data(
            {k: wagtail_to_formdata(v) for k, v in model_to_dict(page).items()}
        )

        print("FORM_DATA:", form_data)
        form = form_class(form_data)
        if not form.is_valid():
            errs = form.errors.as_data()
            if "slug" in errs:
                errs["slug"] = [err for err in errs["slug"] if err.code != "slug-in-use"]
                if not errs["slug"]:
                    del errs["slug"]
            # TODO: better error stuff
            if errs:
                error_messsage = self.errors_to_strings(errs)

                raise ImportException(f"Validation error: {error_messsage}", self.row_num)
            
    def errors_to_strings(self, errs):
        errors = errs[next(iter(errs))][0]

        if isinstance(errors, dict):
            return {key: self.errors_to_strings(value) for key, value in errs.items()}
        elif isinstance(errors, list):
            return [self.errors_to_strings(value) for value in errs]
        elif isinstance(errors, StreamBlockValidationError):
            json_data_errors = errors.as_json_data()
            error_messages = []
            for value in json_data_errors['blockErrors'][0]['blockErrors'].values():
                if isinstance(value, dict) and 'messages' in value:
                    error_messages.extend(value['messages'])
            return error_messages[0]          
        elif isinstance(errors, ValidationError):
            return errors.message

    def save(self, parent: Page) -> None:
        try:
            page = ContentPage.objects.get(slug=self.slug, locale=self.locale)
        except ContentPage.DoesNotExist:
            page = ContentPage(slug=self.slug, locale=self.locale)

        self.add_web_to_page(page)
        self.add_whatsapp_to_page(page)
        self.add_sms_to_page(page)
        self.add_ussd_to_page(page)
        self.add_messenger_to_page(page)
        self.add_viber_to_page(page)
        self.add_tags_to_page(page)
        self.add_quick_replies_to_page(page)
        self.add_triggers_to_page(page)
        self.validate_page_using_form(page)

        try:
            with contextlib.suppress(NodeAlreadySaved):
                parent.add_child(instance=page)
            page.save_revision().publish()
        except ValidationError as err:
            # FIXME: Find a better way to represent this.
            raise ImportException(f"Validation error: {err}", self.row_num)

    def add_web_to_page(self, page: ContentPage) -> None:
        page.enable_web = self.enable_web
        page.title = self.title
        page.subtitle = self.subtitle
        page.body = self.formatted_body
        if self.translation_key is not None:
            page.translation_key = self.translation_key

    def add_whatsapp_to_page(self, page: ContentPage) -> None:
        page.enable_whatsapp = self.enable_whatsapp
        page.is_whatsapp_template = self.is_whatsapp_template
        page.whatsapp_title = self.whatsapp_title
        page.whatsapp_template_name = self.whatsapp_template_name
        page.whatsapp_template_category = self.whatsapp_template_category
        page.whatsapp_body.clear()
        for message in self.formatted_whatsapp_body:
            page.whatsapp_body.append(("Whatsapp_Message", message))

    def add_sms_to_page(self, page: ContentPage) -> None:
        page.enable_sms = self.enable_sms
        page.sms_title = self.sms_title
        page.sms_body.clear()
        for message in self.formatted_sms_body:
            page.sms_body.append(("SMS_Message", message))

    def add_ussd_to_page(self, page: ContentPage) -> None:
        page.enable_ussd = self.enable_ussd
        page.ussd_title = self.ussd_title
        page.ussd_body.clear()
        for message in self.formatted_ussd_body:
            page.ussd_body.append(("USSD_Message", message))

    def add_messenger_to_page(self, page: ContentPage) -> None:
        page.enable_messenger = self.enable_messenger
        page.messenger_title = self.messenger_title
        page.messenger_body.clear()
        for message in self.formatted_messenger_body:
            page.messenger_body.append(("messenger_block", message))

    def add_viber_to_page(self, page: ContentPage) -> None:
        page.enable_viber = self.enable_viber
        page.viber_title = self.viber_title
        page.viber_body.clear()
        for message in self.formatted_viber_body:
            page.viber_body.append(("viber_message", message))

    def add_tags_to_page(self, page: ContentPage) -> None:
        page.tags.clear()
        for tag_name in self.tags:
            tag, _ = Tag.objects.get_or_create(name=tag_name)
            page.tags.add(tag)

    def add_quick_replies_to_page(self, page: ContentPage) -> None:
        for quick_reply_name in self.quick_replies:
            quick_reply, _ = ContentQuickReply.objects.get_or_create(
                name=quick_reply_name
            )
            page.quick_replies.add(quick_reply)

    def add_triggers_to_page(self, page: ContentPage) -> None:
        for trigger_name in self.triggers:
            trigger, _ = ContentTrigger.objects.get_or_create(name=trigger_name)
            page.triggers.add(trigger)

    def link_related_pages(self) -> None:
        page = ContentPage.objects.get(slug=self.slug, locale=self.locale)
        related_pages = []
        for related_page_slug in self.related_pages:
            try:
                related_page = Page.objects.get(
                    slug=related_page_slug, locale=self.locale
                )
            except Page.DoesNotExist:
                raise ImportException(
                    f"Cannot find related page with slug '{related_page_slug}' and "
                    f"locale '{self.locale}'",
                    self.row_num,
                )
            related_pages.append(("related_page", related_page))
        page.related_pages = related_pages
        page.save_revision().publish()

    @property
    def formatted_body(self) -> list[tuple[str, RichText]]:
        if not self.body:
            return []
        formatted = []
        for line in self.body.splitlines():
            if line:
                formatted.append(("paragraph", RichText(line)))
        return formatted

    @property
    def formatted_whatsapp_body(self) -> list[StructValue]:
        return [WhatsappBlock().to_python(m.wagtail_format) for m in self.whatsapp_body]

    @property
    def formatted_sms_body(self) -> list[StructValue]:
        return [SMSBlock().to_python(m.wagtail_format) for m in self.sms_body]

    @property
    def formatted_ussd_body(self) -> list[StructValue]:
        return [USSDBlock().to_python(m.wagtail_format) for m in self.ussd_body]

    @property
    def formatted_messenger_body(self) -> list[StructValue]:
        return [
            MessengerBlock().to_python(m.wagtail_format) for m in self.messenger_body
        ]

    @property
    def formatted_viber_body(self) -> list[StructValue]:
        return [ViberBlock().to_python(m.wagtail_format) for m in self.viber_body]


@dataclass(slots=True)
class ShadowWhatsappBlock:
    message: str = ""
    next_prompt: str = ""
    buttons: list[dict[str, Any]] = field(default_factory=list)
    example_values: list[str] = field(default_factory=list)
    variation_messages: list["ShadowVariationBlock"] = field(default_factory=list)
    list_items: list[str] = field(default_factory=list)
    footer: str = ""

    @property
    def wagtail_format(
        self,
    ) -> dict[str, str | list[dict[str, str | list[dict[str, str]]]] | list[str]]:
        return {
            "message": self.message,
            "next_prompt": self.next_prompt,
            "example_values": self.example_values,
            "buttons": self.buttons,
            "variation_messages": [m.wagtail_format for m in self.variation_messages],
            "list_items": self.list_items,
            "footer": self.footer,
        }


@dataclass(slots=True)
class ShadowVariationBlock:
    message: str = ""
    variation_restrictions: dict[str, str] = field(default_factory=dict)

    @property
    def wagtail_format(self) -> dict[str, str | list[dict[str, str]]]:
        return {
            "message": self.message,
            "variation_restrictions": [
                {"type": t, "value": v} for t, v in self.variation_restrictions.items()
            ],
        }


@dataclass(slots=True)
class ShadowSMSBlock:
    message: str = ""

    @property
    def wagtail_format(self) -> dict[str, str]:
        return {"message": self.message}


@dataclass(slots=True)
class ShadowUSSDBlock:
    message: str = ""

    @property
    def wagtail_format(self) -> dict[str, str]:
        return {"message": self.message}


@dataclass(slots=True)
class ShadowMessengerBlock:
    message: str = ""

    @property
    def wagtail_format(self) -> dict[str, str]:
        return {"message": self.message}


@dataclass(slots=True)
class ShadowViberBlock:
    message: str = ""

    @property
    def wagtail_format(self) -> dict[str, str]:
        return {"message": self.message}


@dataclass(slots=True, frozen=True)
class ContentRow:
    slug: str
    page_id: int | None = None
    parent: str = ""
    web_title: str = ""
    web_subtitle: str = ""
    web_body: str = ""
    whatsapp_title: str = ""
    whatsapp_body: str = ""
    whatsapp_template_name: str = ""
    whatsapp_template_category: str = ""
    example_values: list[str] = field(default_factory=list)
    variation_title: dict[str, str] = field(default_factory=dict)
    variation_body: str = ""
    list_items: list[str] = field(default_factory=list)
    sms_title: str = ""
    sms_body: str = ""
    ussd_title: str = ""
    ussd_body: str = ""
    messenger_title: str = ""
    messenger_body: str = ""
    viber_title: str = ""
    viber_body: str = ""
    translation_tag: str = ""
    tags: list[str] = field(default_factory=list)
    quick_replies: list[str] = field(default_factory=list)
    triggers: list[str] = field(default_factory=list)
    locale: str = ""
    next_prompt: str = ""
    buttons: list[dict[str, Any]] = field(default_factory=list)
    image_link: str = ""
    doc_link: str = ""
    media_link: str = ""
    related_pages: list[str] = field(default_factory=list)
    footer: str = ""

    @classmethod
    def from_flat(cls, row: dict[str, str]) -> "ContentRow":
        class_fields = {field.name for field in fields(cls)}
        row = {
            key.strip(): value.strip()
            for key, value in row.items()
            if value and key in class_fields
        }

        return cls(
            page_id=int(row.pop("page_id")) if row.get("page_id") else None,
            variation_title=deserialise_dict(row.pop("variation_title", "")),
            tags=deserialise_list(row.pop("tags", "")),
            quick_replies=deserialise_list(row.pop("quick_replies", "")),
            triggers=deserialise_list(row.pop("triggers", "")),
            related_pages=deserialise_list(row.pop("related_pages", "")),
            example_values=deserialise_list(row.pop("example_values", "")),
            buttons=json.loads(row.pop("buttons", "")) if row.get("buttons") else [],
            list_items=deserialise_list(row.pop("list_items", "")),
            footer=row.pop("footer") if row.get("footer") else "",
            **row,
        )

    @property
    def is_page_index(self) -> bool:
        return bool(self.web_title) and not any(
            [
                self.parent,
                self.web_body,
                self.whatsapp_body,
                self.sms_body,
                self.ussd_body,
                self.messenger_body,
                self.viber_body,
            ]
        )

    @property
    def is_content_page(self) -> bool:
        return bool(self.web_title)

    @property
    def is_whatsapp_message(self) -> bool:
        return bool(self.whatsapp_body)

    @property
    def is_whatsapp_template_message(self) -> bool:
        return bool(self.whatsapp_template_name)

    @property
    def is_sms_message(self) -> bool:
        return bool(self.sms_body)

    @property
    def is_ussd_message(self) -> bool:
        return bool(self.ussd_body)

    @property
    def is_messenger_message(self) -> bool:
        return bool(self.messenger_body)

    @property
    def is_viber_message(self) -> bool:
        return bool(self.viber_body)

    @property
    def is_variation_message(self) -> bool:
        return bool(self.variation_body)


def deserialise_dict(value: str) -> dict[str, str]:
    if not value:
        return {}
    result = {}
    for item in value.strip().split(","):
        key, value = item.split(":")
        result[key.strip()] = value.strip()
    return result


def deserialise_list(value: str) -> list[str]:
    if not value:
        return []

    items = list(csv.reader([value]))[0]
    return [item.strip() for item in items]
