import copy
import csv
from collections.abc import Iterable
from dataclasses import asdict, astuple, dataclass, fields
from math import ceil

from django.http import HttpResponse  # type: ignore
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from wagtail.query import QuerySet  # type: ignore  # No typing available


@dataclass
class ExportRow:
    """
    All the data for a single row of an ordered content set
    """

    name: str
    profile_field: str
    page_slugs: str
    time: int | None
    unit: str | None
    before_or_after: str | None
    contact_field: str | None
    slug: str | None
    locale: str | None

    @classmethod
    def headings(cls) -> list[str]:
        """
        The field names, used to write the header rows in exports
        """
        return [f.name for f in fields(cls)]

    def to_dict(self) -> dict[str, str]:
        return asdict(self)

    def to_tuple(self) -> tuple[str]:
        return astuple(self)


class OrderedSetExporter:
    rows: list[ExportRow]

    def __init__(self, queryset: QuerySet):
        self.queryset = queryset
        self.rows = []

    def perform_export(self) -> Iterable[ExportRow]:
        for item in self.queryset:
            yield ExportRow(
                name=item.name,
                profile_field=item.profile_field,
                page_slugs=item.page_slugs,
                time=item.time,
                unit=item.unit,
                before_or_after=str(item.before_or_after),
                contact_field=item.contact_field,
                slug=item.slug,
                locale=item.locale.language_code,
            )


class OrderedSetsExportWriter:
    rows: Iterable[ExportRow]

    def __init__(self, rows: Iterable[ExportRow]):
        self.rows = rows

    def write_xlsx(self, response: HttpResponse) -> None:
        workbook = Workbook()
        worksheet: Worksheet = workbook.active  # type: ignore  # This is not a write or read only workbook
        worksheet.append(ExportRow.headings())
        for row in self.rows:
            worksheet.append(row.to_tuple())
        _set_xlsx_styles(workbook, worksheet)
        workbook.save(response)  # type: ignore

    def write_csv(self, response: HttpResponse) -> None:
        writer = csv.DictWriter(f=response, fieldnames=ExportRow.headings())
        writer.writeheader()

        for row in self.rows:
            writer.writerow(row.to_dict())


def _set_xlsx_styles(wb: Workbook, sheet: Worksheet) -> None:
    """
    Sets the style for the workbook adding any formatting that will make the sheet more
    aesthetically pleasing
    """
    # Adjustment is because the size in openxlsx and google sheets are not equivalent
    adjustment = 7

    # Set columns based on best size
    column_widths_in_pts = {
        "name": 130,
        "profile_field": 110,
        "page_slugs": 100,
        "time": 100,
        "unit": 110,
        "before_or_after": 120,
        "contact_field": 100,
        "slug": 100,
        "locale": 100,
    }
    for column in sheet.iter_cols(max_row=1):
        [cell] = column
        width = column_widths_in_pts[str(cell.value)]
        sheet.column_dimensions[get_column_letter(cell.col_idx)].width = ceil(
            width / adjustment
        )

    # Named Styles
    header_style = NamedStyle(name="header_style")

    # Set attributes to styles
    header_style.font = Font(bold=True, size=10)

    # Add named styles to wb
    wb.add_named_style(header_style)

    # Set header style for row 1
    for row in sheet["1:2"]:
        for cell in row:
            cell.style = header_style

    # set font on all cells initially to 10pt and row height
    general_font = Font(size=10)
    for index, row in enumerate(sheet.iter_rows()):
        if index > 2:
            sheet.row_dimensions[index].height = 60  # type: ignore # Bad annotation.
        for cell in row:
            cell.font = general_font
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment  # type: ignore # Broken typeshed update, maybe?
