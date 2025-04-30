import copy
import csv
import io
import json
from collections.abc import Iterable
from dataclasses import asdict, astuple, dataclass, fields
from math import ceil

from django.http import HttpResponse  # type: ignore
from openpyxl.styles import Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from wagtail import blocks  # type: ignore
from wagtail.query import PageQuerySet  # type: ignore


@dataclass
class ExportRow:
    name: str = ""
    category: str = ""
    buttons: str = ""
    locale: str = ""
    image: str = ""
    message: str = ""
    example_values: str = ""
    submission_name: str = ""
    submission_status: str = ""
    submission_result: str = ""

    @classmethod
    def headings(cls) -> list[str]:
        return [f.name for f in fields(cls)]

    def to_dict(self) -> dict[str, str | int]:
        return asdict(self)

    def to_tuple(self) -> tuple[str | int, ...]:
        return astuple(self)


class WhatsAppTemplateExporter:
    def __init__(self, queryset: PageQuerySet):
        self.rows = []  #  type:  list[dict[str | int, str | int]]
        self.queryset = queryset

    def perform_export(self) -> Iterable[ExportRow]:
        for item in self.queryset:
            image_link = ""
            if item.image:
                image_link = item.image.file.url
            yield ExportRow(
                name=item.name,
                category=item.category,
                buttons=self.serialise_buttons(item.buttons),
                locale=str(item.locale.language_code),
                image=str(image_link),
                message=str(item.message),
                example_values=serialize_list(
                    [v["value"] for v in item.example_values.raw_data]
                ),
                submission_name=str(item.submission_name),
                submission_status=str(item.submission_status),
                submission_result=(item.submission_result),
            )

    @staticmethod
    def serialise_buttons(buttons: blocks.StreamValue.StreamChild) -> str:
        button_dicts = []

        for button in buttons:
            button_dict = {
                "type": button.block_type,
                "title": button.value["title"],
                "slug": "",
            }
            if button.block_type == "go_to_page":
                # Exclude buttons that has deleted pages that they are linked to it
                if button.value.get("page") is None:
                    continue
                button_dict["slug"] = button.value["page"].slug
            if button.block_type == "go_to_form":
                # Exclude buttons that has deleted forms that they are linked to it
                if button.value.get("form") is None:
                    continue
                button_dict["slug"] = button.value["form"].slug

            button_dicts.append(button_dict)
        return json.dumps(button_dicts)


def serialize_list(items: Iterable[str]) -> str:
    """
    Uses CSV formatting to seralize a list of strings, handling escaping
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(items)
    return output.getvalue().rstrip("\r\n")


"""
Definining ExportWriter class for Whatsapp Templates
"""


class WhatsAppTemplateExportWriter:
    rows: Iterable[ExportRow]

    def __init__(self, rows: Iterable[ExportRow]):
        self.rows = rows

    def write_xlsx(self, response: HttpResponse) -> None:
        workbook = Workbook()
        worksheet: Worksheet = workbook.active  # type: ignore  # This is not a write or read only workbook
        worksheet.append(ExportRow.headings())
        for row in self.rows:
            worksheet.append(row.to_tuple())
        _set_xlsx_styles(workbook, worksheet)
        workbook.save(response)

    def write_csv(self, response: HttpResponse) -> None:
        writer = csv.DictWriter(f=response, fieldnames=ExportRow.headings())
        writer.writeheader()

        for row in self.rows:
            writer.writerow(row.to_dict())


def _set_xlsx_styles(wb: Workbook, sheet: Worksheet) -> None:
    """Sets the style for the workbook adding any formatting that will make the sheet more aesthetically pleasing"""
    # Adjustment is because the size in openxlsx and google sheets are not equivalent
    adjustment = 7
    # Padding
    sheet.insert_cols(1)

    # Set columns based on best size

    column_widths_in_pts = {
        "name": 110,
        "category": 110,
        "buttons": 110,
        "locale": 118,
        "image": 110,
        "message": 110,
        "example_values": 110,
        "submission_name": 110,
        "submission_status": 118,
        "submission_result": 110,
    }

    for index, column_width in enumerate(column_widths_in_pts.values(), 2):
        sheet.column_dimensions[get_column_letter(index)].width = ceil(
            column_width / adjustment
        )

    # Freeze heading row and side panel, 1 added because it freezes before the column
    panel_column = get_column_letter(5)
    sheet.freeze_panes = sheet[f"{panel_column}2"]

    # Colours
    blue = Color(rgb="0099CCFF")

    # Boarders
    left_border = Border(left=Side(border_style="thin", color="FF000000"))

    # Fills
    blue_fill = PatternFill(patternType="solid", fgColor=blue)

    # Named Styles
    header_style = NamedStyle(name="header_style")
    menu_style = NamedStyle(name="menu_style")

    # Set attributes to styles
    header_style.font = Font(bold=True, size=10)
    menu_style.fill = blue_fill
    menu_style.font = Font(bold=True, size=10)

    # Add named styles to wb
    wb.add_named_style(header_style)
    wb.add_named_style(menu_style)

    # column widths

    # Set header style for row 1 and 2
    for row in sheet["1:2"]:
        for cell in row:
            cell.style = header_style

    # Set dividing border for side panel
    for cell in sheet[f"{panel_column}:{panel_column}"]:
        cell.border = left_border

    # set font on all cells initially to 10pt and row height
    general_font = Font(size=10)
    for index, row in enumerate(sheet.iter_rows()):
        if index > 2:
            sheet.row_dimensions[index].height = 60  # type: ignore # Bad annotation.
        for cell in row:
            cell.font = general_font
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment  # type: ignore # Broken typeshed update, maybe?
