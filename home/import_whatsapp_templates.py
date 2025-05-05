import csv
from collections import defaultdict
from dataclasses import dataclass, field, fields
from datetime import datetime
from io import BytesIO, StringIO
from queue import Queue
from typing import Any
from uuid import uuid4

from django.core.exceptions import ValidationError  # type: ignore
from openpyxl import load_workbook
from wagtail.coreutils import get_content_languages  # type: ignore
from wagtail.models import Locale  # type: ignore
from wagtail.models.sites import Site  # type: ignore

from home.import_helpers import ImportException, JSON_loader
from home.models import Assessment, WhatsAppTemplate  # type: ignore

PageId = tuple[str, Locale]


class WhatsAppTemplateImporter:
    def __init__(
        self,
        file_content: bytes,
        file_type: str,
        progress_queue: Queue[int],
        purge: bool | str = True,
        locale: Locale | str | None = None,
    ):
        if isinstance(locale, str):
            locale = Locale.objects.get(language_code=locale)
        self.file_content = file_content
        self.file_type = file_type
        self.progress_queue = progress_queue
        self.purge = purge in ["True", "yes", True]
        self.locale = locale
        self.locale_map: dict[str, Locale] = {}
        self.go_to_page_buttons: dict[PageId, list[dict[str, Any]]] = defaultdict(list)

    def locale_from_language_code(self, lang_code_entry: str) -> Locale:
        if lang_code_entry not in self.locale_map:
            codes = []
            lang_name = ""
            for lang_code, lang_dn in get_content_languages().items():
                if lang_code == lang_code_entry:
                    lang_name = lang_dn
                    codes.append(lang_code)
            if not codes:
                raise ImportException(f"Language code not found: {lang_code_entry}")
            if len(codes) > 1:
                raise ImportException(
                    f"Multiple codes for language: {lang_name} -> {codes}"
                )
            self.locale_map[lang_code_entry] = Locale.objects.get(
                language_code=codes[0]
            )
        return self.locale_map[lang_code_entry]

    def perform_import(self) -> None:
        rows = self.parse_file()
        self.set_progress("Loaded file", 5)

        if self.purge:
            self.delete_existing_content()
        self.set_progress("Deleted existing WhatsApp Template", 10)

        self.process_rows(rows)

    def process_rows(self, rows: list["ContentRow"]) -> None:
        for i, row in reversed(list(enumerate(rows, start=2))):
            try:
                self.create_whatsapp_template_from_row(row)
            except ImportException as e:
                e.row_num = i
                raise e
            except ValidationError as errors:
                err = []
                for error in errors:
                    field_name = error[0]
                    for msg in error[1]:
                        err.append(f"{field_name} - {msg}")
                raise ImportException(
                    [f"Validation error: {msg}" for msg in err], i, locale=row.locale
                )

    def create_whatsapp_template_from_row(self, row: "ContentRow") -> None:
        locale = self.locale_from_language_code(row.locale)

        if row.category not in WhatsAppTemplate.Category.values:
            raise ImportException(
                f"Validation error: whatsapp_template_category - Select a valid choice. {row.category} is not one of the available choices."
            )

        template = self._update_or_create_whatsapp_template(row, locale)
        template.full_clean()
        template.save()

        buttons = self._create_interactive_items(
            row.buttons, template.name, locale, "button"
        )
        template.buttons = buttons

        template.full_clean()
        template.save()
        return

    def _update_or_create_whatsapp_template(
        self, row: "ContentRow", locale: Locale
    ) -> WhatsAppTemplate:
        try:
            template = WhatsAppTemplate.objects.get(name=row.name, locale=locale)
            template.category = row.category
            template.message = row.message
            template.example_values = [
                ("example_values", v) for v in row.example_values
            ]
            template.submission_status = (
                row.submission_status
                if row.submission_status
                else WhatsAppTemplate.SubmissionStatus.NOT_SUBMITTED_YET
            )
            template.submission_result = row.submission_result
            template.submission_name = row.submission_name
            return template
        except WhatsAppTemplate.DoesNotExist:
            return WhatsAppTemplate(
                name=row.name,
                category=row.category,
                locale=locale,
                message=row.message,
                example_values=[("example_values", v) for v in row.example_values],
                submission_status=(
                    row.submission_status
                    if row.submission_status
                    else WhatsAppTemplate.SubmissionStatus.NOT_SUBMITTED_YET
                ),
                submission_result=row.submission_result,
                submission_name=row.submission_name,
            )

    def parse_file(self) -> list["ContentRow"]:
        if self.file_type == "XLSX":
            return self.parse_excel()
        return self.parse_csv()

    def parse_excel(self) -> list["ContentRow"]:
        workbook = load_workbook(BytesIO(self.file_content), read_only=True)
        worksheet = workbook.worksheets[0]

        def clean_excel_cell(cell_value: str | float | datetime | None) -> str:
            return str(cell_value).replace("_x000D", "")

        first_row = next(worksheet.iter_rows(max_row=1, values_only=True))
        header = [clean_excel_cell(cell) if cell else None for cell in first_row]
        rows: list[ContentRow] = []
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {}
            for name, cell in zip(header, row):  # noqa: B905 (TODO: strict?)
                if name and cell:
                    r[name] = clean_excel_cell(cell)
            rows.append(ContentRow.from_flat(r, row_num))
        # page_id_rows = group_rows_by_page_id(rows)
        return rows

    def parse_csv(self) -> list["ContentRow"]:
        reader = list(csv.DictReader(StringIO(self.file_content.decode())))
        rows = [ContentRow.from_flat(row, i) for i, row in enumerate(reader, start=2)]
        # page_id_rows = group_rows_by_page_id(rows)
        return rows

    def set_progress(self, message: str, progress: int) -> None:
        self.progress_queue.put_nowait(progress)

    def delete_existing_content(self) -> None:
        WhatsAppTemplate.objects.all().delete()

    def default_locale(self) -> Locale:
        site = Site.objects.get(is_default_site=True)
        return site.root_page.locale

    def _create_interactive_items(
        self,
        row_field: list[dict[str, Any]],
        template_name: str,
        locale: Locale,
        item_type: str,
    ) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for index, item in enumerate(row_field):
            try:
                if item["type"] == "next_message":
                    items.append(
                        {
                            "id": str(uuid4()),
                            "type": item["type"],
                            "value": {"title": item["title"]},
                        }
                    )
                elif item["type"] == "go_to_page":
                    item["index"] = index
                    if item_type == "button":
                        go_to_page = self.go_to_page_buttons
                    else:
                        raise ImportException(
                            f"{item_type} with invalid type '{item['type']}'"
                        )
                    page_gtp = go_to_page[(template_name, locale)]
                    page_gtp.append(item)
                elif item["type"] == "go_to_form":
                    form = self._get_form(
                        item["slug"],
                        locale,
                        item["title"],
                        item_type,
                    )
                    items.append(
                        {
                            "id": str(uuid4()),
                            "type": item["type"],
                            "value": {
                                "title": item["title"],
                                "form": form.id,
                            },
                        }
                    )
                elif not item["type"]:
                    pass
                else:
                    raise ImportException(
                        f"{item_type} with invalid type '{item['type']}'"
                    )
            except KeyError as e:
                raise ImportException(f"{item_type} is missing key {e}")
        return items

    def _get_form(
        self, slug: str, locale: Locale, title: str, item_type: str
    ) -> Assessment:
        try:
            return Assessment.objects.get(slug=slug, locale=locale)
        except Assessment.DoesNotExist:
            raise ImportException(
                f"No form found with slug '{slug}' and locale '{locale}' for go_to_form {item_type} '{title}'"
            )


@dataclass(slots=True, frozen=True)
class ContentRow:
    name: str = ""
    category: str = ""
    locale: str = ""
    buttons: list[dict[str, Any]] = field(default_factory=list)
    image_link: str = ""
    message: str = ""
    example_values: list[str] = field(default_factory=list)
    submission_name: str = ""
    submission_status: str = ""
    submission_result: str = ""

    @classmethod
    def from_flat(cls, row: dict[str, str], row_num: int) -> "ContentRow":
        class_fields = {field.name for field in fields(cls)}
        row = {
            key.strip(): value.strip()
            for key, value in row.items()
            if value and key in class_fields
        }
        return cls(
            name=str(row.pop("name", "")),
            category=str(row.pop("category", "")),
            buttons=(
                JSON_loader(row_num, row.pop("buttons", ""))
                if row.get("buttons")
                else []
            ),
            example_values=deserialise_list(row.pop("example_values", "")),
            submission_name=str(row.pop("submission_name", "")),
            submission_status=str(row.pop("submission_status", "")),
            submission_result=str(row.pop("submission_result", "")),
            **row,
        )


def deserialise_list(value: str) -> list[str]:
    if not value:
        return []

    items = list(csv.reader([value]))[0]
    return [item.strip() for item in items]
