import copy
import csv
from dataclasses import asdict, astuple, dataclass, fields
from itertools import zip_longest
from json import dumps
from math import ceil

from django.http import HttpResponse  # type: ignore
from openpyxl.styles import Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from wagtail import blocks  # type: ignore
from wagtail.models import Locale, Page  # type: ignore
from wagtail.query import PageQuerySet  # type: ignore

from .models import (
    ContentPage,
    ContentPageIndex,
    HomePage,
    MessengerBlock,
    SMSBlock,
    USSDBlock,
    VariationBlock,
    ViberBlock,
    WhatsappBlock,
    WhatsAppTemplate,
)
from .xlsx_helpers import get_active_sheet

HP_CTYPE = HomePage._meta.verbose_name
CP_CTYPE = ContentPage._meta.verbose_name
CPI_CTYPE = ContentPageIndex._meta.verbose_name


MsgBlocks = tuple[
    WhatsappBlock | None,
    SMSBlock | None,
    USSDBlock | None,
    MessengerBlock | None,
    ViberBlock | None,
]


@dataclass
class ExportRow:
    structure: str = ""
    message: int = 0
    page_id: int = 0
    slug: str = ""
    parent: str = ""
    web_title: str = ""
    web_subtitle: str = ""
    web_body: str = ""
    whatsapp_title: str = ""
    whatsapp_body: str = ""
    whatsapp_template_name: str = ""
    variation_title: str = ""
    variation_body: str = ""
    list_title: str = ""
    list_items: str = ""
    sms_title: str = ""
    sms_body: str = ""
    ussd_title: str = ""
    ussd_body: str = ""
    messenger_title: str = ""
    messenger_body: str = ""
    viber_title: str = ""
    viber_body: str = ""
    translation_tag: str = ""
    tags: str = ""
    quick_replies: str = ""
    triggers: str = ""
    next_prompt: str = ""
    buttons: str = ""
    image_link: str = ""
    doc_link: str = ""
    media_link: str = ""
    related_pages: str = ""
    footer: str = ""
    language_code: str = ""

    @classmethod
    def headings(cls) -> list[str]:
        return [f.name for f in fields(cls)]

    def to_dict(self) -> dict[str, str | int]:
        return asdict(self)

    def to_tuple(self) -> tuple[str | int, ...]:
        return astuple(self)

    def new_message_row(self) -> "ExportRow":
        return ExportRow(message=self.message + 1, page_id=self.page_id, slug=self.slug)

    def new_variation_row(self, variation: VariationBlock) -> "ExportRow":
        return ExportRow(
            message=self.message,
            page_id=self.page_id,
            slug=self.slug,
            variation_title=", ".join(
                f"{pf['type']}: {pf['value']}"
                for pf in variation.get("variation_restrictions").raw_data
            ),
            variation_body=variation.get("message"),
        )

    def add_message_fields(self, msg_blocks: MsgBlocks) -> None:
        whatsapp, sms, ussd, messenger, viber = msg_blocks
        # We do these in reverse order to pick the same image as the old
        # exporter if there's more than one.
        if viber:
            self.viber_body = viber.value["message"].strip()
            if "image" in viber.value and viber.value["image"] is not None:
                self.image_link = viber.value["image"].file.url
        if messenger:
            self.messenger_body = messenger.value["message"].strip()
            if "image" in messenger.value and messenger.value["image"] is not None:
                self.image_link = messenger.value["image"].file.url
        if sms:
            self.sms_body = sms.value["message"].strip()
        if ussd:
            self.ussd_body = ussd.value["message"].strip()
        if whatsapp:
            if isinstance(whatsapp.value, WhatsAppTemplate):
                self.whatsapp_template_name = whatsapp.value.name
            else:
                self.whatsapp_body = whatsapp.value["message"].strip()
                if "image" in whatsapp.value and whatsapp.value["image"] is not None:
                    self.image_link = whatsapp.value["image"].file.url
                if (
                    "document" in whatsapp.value
                    and whatsapp.value["document"] is not None
                ):
                    self.doc_link = whatsapp.value["document"].file.url
                if "media" in whatsapp.value and whatsapp.value["media"] is not None:
                    self.media_link = whatsapp.value["media"].file.url
                if "next_prompt" in whatsapp.value:
                    self.next_prompt = whatsapp.value["next_prompt"]
                if "buttons" in whatsapp.value:
                    self.buttons = self.serialise_buttons(whatsapp.value["buttons"])
                if "footer" in whatsapp.value:
                    self.footer = whatsapp.value["footer"]
                if "list_title" in whatsapp.value:
                    self.list_title = whatsapp.value["list_title"]
                if "list_items" in whatsapp.value:
                    self.list_items = self.serialise_buttons(
                        whatsapp.value["list_items"]
                    )

    @staticmethod
    def serialise_buttons(buttons: blocks.StreamValue.StreamChild) -> str:
        button_dicts = []

        for button in buttons:
            button_dict = {"type": button.block_type, "title": button.value["title"]}
            if button.block_type == "go_to_page":
                # Exclude buttons that has deleted pages that they are linked to it
                if button.value.get("page") is None:
                    continue
                button_dict["slug"] = button.value["page"].slug
            if button.block_type == "go_to_form":
                # Exclude buttons that has deleted forms that they are linked to it
                if button.value.get("form") is None:
                    continue
                button_dict["slug"] = button.value["form"].slug

            button_dicts.append(button_dict)
        return dumps(button_dicts)


class ContentExporter:
    rows: list[ExportRow]

    def __init__(self, queryset: PageQuerySet):
        self.rows = []
        self.queryset = queryset

    def perform_export(self) -> list[ExportRow]:
        for locale in Locale.objects.all():
            home = HomePage.objects.get(locale_id=locale.id)
            self._export_locale(home)
        return self.rows

    def _export_locale(self, home: HomePage) -> None:
        main_menu_pages = home.get_children()
        for index, page in enumerate(main_menu_pages, 1):
            structure_string = f"Menu {index}"
            self._export_page(page, structure_string)

    def _export_page(self, page: Page, structure: str) -> None:
        if page.content_type.name == CPI_CTYPE:
            self._export_cpi(ContentPageIndex.objects.get(id=page.id), structure)
        elif page.content_type.name == CP_CTYPE:
            content_page = self.queryset.filter(id=page.id).first()
            if content_page:
                self._export_content_page(content_page, structure)
        else:
            raise ValueError(f"Unexpected page type: {page.content_type.name}")
        # Now handle any child pages.
        if page.get_children_count() > 0:
            for index, child in enumerate(page.get_children(), 1):
                child_structure = f"{structure.replace('Menu', 'Sub')}.{index}"
                self._export_page(child, child_structure)

    def _export_content_page(self, page: ContentPage, structure: str) -> None:
        """
        Export a ContentPage.

        FIXME:
         * We should use the parent slug (which is expected to be unique per
           locale (probably?)) instead of the parent title.
        """
        row = ExportRow(
            structure=structure,
            message=1,
            page_id=page.id,
            slug=page.slug,
            parent=self._parent_title(page),
            web_title=page.title,
            web_subtitle=page.subtitle,
            web_body=str(page.body),
            whatsapp_title=page.whatsapp_title,
            whatsapp_template_name=page.whatsapp_template_name,
            sms_title=page.sms_title,
            ussd_title=page.ussd_title,
            messenger_title=page.messenger_title,
            viber_title=page.viber_title,
            translation_tag=str(page.translation_key),
            tags=self._comma_sep_qs(page.tags.all()),
            quick_replies=self._comma_sep_qs(page.quick_replies.all()),
            triggers=self._comma_sep_qs(page.triggers.all()),
            related_pages=self._comma_sep_qs(self._related_pages(page)),
            language_code=page.locale.language_code,
        )
        self.rows.append(row)
        message_bodies = list(
            zip_longest(
                page.whatsapp_body,
                page.sms_body,
                page.ussd_body,
                page.messenger_body,
                page.viber_body,
            )
        )
        for msg_blocks in message_bodies:
            self._export_row_message(row, msg_blocks)
            row = row.new_message_row()

    def _export_row_message(self, row: ExportRow, msg_blocks: MsgBlocks) -> None:
        row.add_message_fields(msg_blocks)
        if self.rows[-1] is not row:
            self.rows.append(row)
        # Only WhatsappBlock has variations at present.
        if msg_blocks[0] is None or isinstance(msg_blocks[0].value, WhatsAppTemplate):
            return
        for variation in msg_blocks[0].value["variation_messages"]:
            self.rows.append(row.new_variation_row(variation))

    def _export_cpi(self, page: ContentPageIndex, structure: str) -> None:
        """
        Export a ContentPageIndex.

        FIXME:
         * We should use the parent slug (which is expected to be unique per
           locale (probably?)) instead of the parent title.
        """
        row = ExportRow(
            structure=structure,
            page_id=page.id,
            slug=page.slug,
            parent=self._parent_title(page),
            web_title=page.title,
            translation_tag=str(page.translation_key),
            language_code=page.locale.language_code,
        )
        self.rows.append(row)

    @staticmethod
    def _parent_title(page: Page) -> str:
        parent = page.get_parent()
        # If the parent is a HomePage, we treat the page as parentless.
        if parent.content_type.name == HP_CTYPE:
            return ""
        return parent.title

    @staticmethod
    def _related_pages(page: ContentPage) -> list[str]:
        # Ideally, all related page links would be removed when the page they
        # link to is deleted. We don't currently do that, so for now we just
        # make sure that we skip such links during export.
        return [rp.value.slug for rp in page.related_pages if rp.value is not None]

    @staticmethod
    def _comma_sep_qs(unformatted_query: PageQuerySet) -> str:
        return ", ".join(str(x) for x in unformatted_query if str(x) != "")


@dataclass
class ExportWriter:
    rows: list[ExportRow]

    def write_csv(self, response: HttpResponse) -> None:
        writer = csv.DictWriter(response, ExportRow.headings())
        writer.writeheader()
        for row in self.rows:
            writer.writerow(row.to_dict())

    def write_xlsx(self, response: HttpResponse) -> None:
        workbook = Workbook()
        worksheet = get_active_sheet(workbook)

        worksheet.append(ExportRow.headings())
        for row in self.rows:
            worksheet.append(row.to_tuple())
        _set_xlsx_styles(workbook, worksheet)
        workbook.save(response)  # type: ignore


def _set_xlsx_styles(wb: Workbook, sheet: Worksheet) -> None:
    """Sets the style for the workbook adding any formatting that will make the sheet more aesthetically pleasing"""
    # Adjustment is because the size in openxlsx and google sheets are not equivalent
    adjustment = 7
    # Padding
    sheet.insert_cols(1)

    # Set columns based on best size

    column_widths_in_pts = {
        "structure": 110,
        "message": 70,
        "page_id": 110,
        "slug": 110,
        "parent": 110,
        "web_title": 110,
        "web_subtitle": 110,
        "web_body": 370,
        "whatsapp_title": 118,
        "whatsapp_body": 370,
        "whatsapp_template_name": 118,
        "variation_title": 118,
        "variation_body": 370,
        "sms_title": 118,
        "sms_body": 370,
        "ussd_title": 118,
        "ussd_body": 370,
        "messenger_title": 118,
        "messenger_body": 370,
        "viber_title": 118,
        "viber_body": 370,
        "translation_tag": 118,
        "tags": 118,
        "quick_replies": 118,
        "triggers": 118,
        "next_prompt": 118,
        "buttons": 118,
        "image_link": 118,
        "doc_link": 118,
        "media_link": 118,
        "related": 118,
        "footer": 118,
        "language_code": 118,
    }

    for index, column_width in enumerate(column_widths_in_pts.values(), 2):
        sheet.column_dimensions[get_column_letter(index)].width = ceil(
            column_width / adjustment
        )

    # Freeze heading row and side panel, 1 added because it freezes before the column
    panel_column = get_column_letter(5)
    sheet.freeze_panes = sheet[f"{panel_column}2"]

    # Colours
    blue = Color(rgb="0099CCFF")

    # Boarders
    left_border = Border(left=Side(border_style="thin", color="FF000000"))

    # Fills
    blue_fill = PatternFill(patternType="solid", fgColor=blue)

    # Named Styles
    header_style = NamedStyle(name="header_style")
    menu_style = NamedStyle(name="menu_style")

    # Set attributes to styles
    header_style.font = Font(bold=True, size=10)
    menu_style.fill = blue_fill
    menu_style.font = Font(bold=True, size=10)

    # Add named styles to wb
    wb.add_named_style(header_style)
    wb.add_named_style(menu_style)

    # column widths

    # Set menu style for any "Menu" row
    for row in sheet.iter_rows():
        if isinstance(row[1].value, str) and "Menu" in row[1].value:
            for cell in row:
                cell.style = menu_style

    # Set header style for row 1 and 2
    for row in sheet["1:2"]:
        for cell in row:
            cell.style = header_style

    # Set dividing border for side panel
    for cell in sheet[f"{panel_column}:{panel_column}"]:
        cell.border = left_border

    # set font on all cells initially to 10pt and row height
    general_font = Font(size=10)
    for index, row in enumerate(sheet.iter_rows()):
        if index > 2:
            sheet.row_dimensions[index].height = 60  # type: ignore # Bad annotation.
        for cell in row:
            cell.font = general_font
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment  # type: ignore # Broken typeshed update, maybe?
