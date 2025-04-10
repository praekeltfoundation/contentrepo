# The error messages are processed and parsed into a list of messages we return to the user
import csv
import json
from collections.abc import Iterator
from datetime import datetime
from io import BytesIO, StringIO
from json.decoder import JSONDecodeError
from typing import Any

from django.core.exceptions import ValidationError  # type: ignore
from django.db.models import Model  # type: ignore
from django.forms import model_to_dict  # type: ignore
from openpyxl import load_workbook
from wagtail.admin.rich_text.converters.contentstate import (  # type: ignore
    ContentstateConverter,  # type: ignore
)
from wagtail.blocks import (  # type: ignore
    StreamBlockValidationError,
    StreamValue,
    StructValue,  # type: ignore
)
from wagtail.blocks.list_block import ListValue  # type: ignore
from wagtail.models import Locale  # type: ignore
from wagtail.rich_text import RichText  # type: ignore
from wagtail.test.utils.form_data import nested_form_data, streamfield  # type: ignore

from .xlsx_helpers import get_active_sheet


class ImportException(Exception):
    """
    Base exception for all import related issues.
    """

    def __init__(
        self,
        message: str | list[str],
        row_num: int | None = None,
        slug: str | None = None,
        locale: Locale | None = None,
    ):
        self.row_num = row_num
        self.message = [message] if isinstance(message, str) else message
        self.slug = slug
        self.locale = locale
        super().__init__()


class ImportAssessmentException(Exception):
    """
    Base exception for all import related issues.
    """

    def __init__(self, message: str, row_num: int | None = None):
        self.row_num = row_num
        self.message = message
        super().__init__()


class ImportWarning:
    """
    Base warnings for all import related issues.
    """

    def __init__(
        self,
        message: str | list[str],
        row_num: int | None = None,
    ):
        self.row_num = row_num
        self.message = message


def wagtail_to_formdata(val: Any) -> Any:
    """
    Convert a model dict field that may be a nested streamfield (or associated
    type) into something we can turn into form data.
    """
    match val:
        case StreamValue():  # type: ignore[misc] # No type info
            return streamfield(
                [(b.block_type, wagtail_to_formdata(b.value)) for b in val]
            )
        case StructValue():  # type: ignore[misc] # No type info
            return {k: wagtail_to_formdata(v) for k, v in val.items()}
        case ListValue():  # type: ignore[misc] # No type info
            # Wagtail doesn't have an equivalent of streamfield() for
            # listvalue, so we have to do it by hand.
            list_val: dict[str, Any] = {
                str(i): {
                    "deleted": "",
                    "order": str(i),
                    "value": wagtail_to_formdata(v),
                }
                for i, v in enumerate(val)
            }
            list_val["count"] = str(len(val))
            return list_val
        case RichText():  # type: ignore[misc] # No type info
            # FIXME: The only RichTextBlock() we currently have is in the web
            #        body and we don't appear to do any validation on it.
            #        There's probably a better way to convert and/or ignore these.
            return ContentstateConverter([]).from_database_format(val.source)
        case _:
            return val


def validate_using_form(edit_handler: Any, model: Model, row_num: int) -> None:
    form_class = edit_handler.get_form_class()

    form_data = nested_form_data(
        {k: wagtail_to_formdata(v) for k, v in model_to_dict(model).items()}
    )

    form = form_class(form_data)
    if not form.is_valid():
        errs = form.errors.as_data()
        if "slug" in errs:
            errs["slug"] = [err for err in errs["slug"] if err.code != "slug-in-use"]
            if not errs["slug"]:
                del errs["slug"]
        # TODO: better error stuff
        if errs:
            errors = []
            error_message = errors_to_list(errs)
            for err in error_message:
                errors.append(f"Validation error: {err}")
            raise ImportException(errors, row_num)


def errors_to_list(errs: dict[str, list[str]]) -> str | list[str]:

    def _extract_errors(
        data: dict[str | int, Any] | list[str]
    ) -> Iterator[tuple[list[str], str]]:
        if isinstance(data, dict):
            items = list(data.items())
        elif isinstance(data, list):
            items = list(enumerate(data))

        for key, value in items:
            key = str(key)
            if isinstance(value, dict | list):
                for child_keys, child_value in _extract_errors(value):
                    yield [key] + child_keys, child_value
            else:
                yield [key], value

    def extract_errors(data: dict[str | int, Any] | list[str]) -> dict[str, str]:
        return {"-".join(key): value for key, value in _extract_errors(data)}

    errors = errs[next(iter(errs))][0]

    if isinstance(errors, dict):
        error_message = {key: errors_to_list(value) for key, value in errs.items()}
    elif isinstance(errors, list):
        error_message = [errors_to_list(value) for value in errs]

    elif isinstance(errors, StreamBlockValidationError):
        json_data_errors = errors.as_json_data()
        error_messages = []
        error_message = ["An unknown error occurred"]
        if isinstance(json_data_errors["blockErrors"], dict):
            error_level = list(json_data_errors["blockErrors"].keys())[0]

            field_name = "Unknown Field"
            if "blockErrors" in json_data_errors["blockErrors"][error_level]:
                field_name = list(
                    json_data_errors["blockErrors"][error_level]["blockErrors"].keys()
                )[0]

                for val in json_data_errors["blockErrors"][error_level][
                    "blockErrors"
                ].values():
                    messages = list(extract_errors(val).values())
            else:
                messages = list(
                    json_data_errors["blockErrors"][error_level]["messages"]
                )

            error_messages.extend(messages)

            error_message = [f"{field_name} - {msg}" for msg in error_messages]

    elif isinstance(errors, ValidationError):
        field_name = list(errs.keys())[0]
        error_messages = errors.messages[0]
        error_message = [f"{field_name} - {error_messages}"]
    else:
        pass

    return error_message


def check_empty_rows(rows: list[dict[str, Any]], row_num: int) -> None:
    """
    Checks if the list of rows is empty and raises an exception if true.
    """
    if not rows:
        raise ImportException(
            "The import file is empty or contains no valid rows.", row_num=row_num
        )


# TODO:
# Move to shared code once we're able to work on the contentpage import.
# Contentsets uses pascal case headers
def convert_headers_to_snake_case(headers: list[str], row_num: int) -> dict[str, str]:
    """
    Converts a list of headers to snake_case and returns a mapping.
    """
    return {header: to_snake_case(header, row_num) for header in headers}


def to_snake_case(s: str, row_num: int) -> str:
    """
    Converts a given string to snake_case if it contains spaces or hyphens.
    """
    return s.replace(" ", "_").replace("-", "_").strip("_")


def fix_rows(rows: Iterator[dict[str | Any, Any]]) -> Iterator[dict[str, str | None]]:
    """
    Fix keys for all rows by lowercasing keys and removing whitespace from keys and values
    """
    try:
        first_row = next(rows)
    except StopIteration:
        return iter([])

    if len(first_row) != len(fix_row(first_row)):
        raise ImportException(
            "Invalid format. Please check that there are no duplicate headers."
        )
    yield fix_row(first_row)

    for row in rows:
        yield fix_row(row)


def fix_row(row: dict[str, str | None]) -> dict[str, str | None]:
    """
    Fix a single row by lowercasing the key and removing whitespace from the key and value
    """
    try:
        return {_normalise_key(k): _normalise_value(v) for k, v in row.items()}
    except AttributeError:
        raise ImportException(
            "Invalid format. Please check that all row values have headers."
        )


def _normalise_key(key: str) -> str:
    return key.lower().strip()


def _normalise_value(value: str | None) -> str | None:
    if value is None:
        return None
    return value.strip()


def parse_file(
    file_content: bytes, file_type: str
) -> Iterator[tuple[int, dict[str, Any]]]:
    read_rows = read_xlsx if file_type == "XLSX" else read_csv
    rows = list(fix_rows(read_rows(file_content)))

    check_empty_rows(rows, row_num=1)

    return enumerate(rows, start=2)


def read_csv(file_content: bytes) -> Iterator[dict[str, Any]]:
    return csv.DictReader(StringIO(file_content.decode()))


def remove_trailing_nones(row: list[Any]) -> list[Any]:
    while row and row[-1] is None:
        row = row[:-1]
    return row


def clean_excel_cell(cell_value: str | float | datetime | None) -> str:
    return "" if cell_value is None else str(cell_value).replace("_x000D", "").strip()


def read_xlsx(file_content: bytes) -> Iterator[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    worksheet = get_active_sheet(workbook)

    first_row = next(worksheet.iter_rows(max_row=1, values_only=True))
    header = [clean_excel_cell(cell) if cell else None for cell in first_row]
    header = remove_trailing_nones(header)

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row = remove_trailing_nones(row)  # type: ignore # Mypy cannot guarantee row is a list; rows may be tuples so we bypass.
        r = {}
        if len(row) > len(header):
            raise ImportException(
                "Invalid format. Please check that all row values have headers."
            )
        for name, cell in zip(header, row, strict=False):
            if name:
                r[name] = clean_excel_cell(cell)
        if r:
            yield r


def JSON_loader(row_num: int, value: str) -> list[dict[str, Any]]:
    if not value:
        return []

    try:
        button = json.loads(value)
    except JSONDecodeError:
        raise ImportException(f"Bad JSON button, you have: {value}", row_num)

    return button
