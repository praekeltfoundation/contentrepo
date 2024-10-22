import csv
import io
from io import BytesIO
from logging import getLogger

from django.db import transaction
from django.http import HttpResponse
from openpyxl import load_workbook
from wagtail.query import PageQuerySet

from home.import_helpers import ImportException
from home.models import ContentPage, OrderedContentSet

logger = getLogger(__name__)


@transaction.atomic
def import_content(file, filetype, progress_queue, purge=True, locale=None) -> None:
    from .import_content_pages import ContentImporter

    importer = ContentImporter(file.read(), filetype, progress_queue, purge, locale)
    importer.perform_import()


def export_xlsx_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_content_pages import ContentExporter, ExportWriter

    exporter = ContentExporter(queryset)
    export_rows = exporter.perform_export()
    ExportWriter(export_rows).write_xlsx(response)


def export_csv_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_content_pages import ContentExporter, ExportWriter

    exporter = ContentExporter(queryset)
    export_rows = exporter.perform_export()
    ExportWriter(export_rows).write_csv(response)


"""Ordered Content Sets Imports/Export"""


def export_xlsx_ordered_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_ordered_sets import OrderedSetExporter, OrderedSetsExportWriter

    exporter = OrderedSetExporter(queryset)
    export_rows = exporter.perform_export()
    OrderedSetsExportWriter(export_rows).write_xlsx(response)


def export_csv_ordered_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_ordered_sets import OrderedSetExporter, OrderedSetsExportWriter

    exporter = OrderedSetExporter(queryset)
    export_rows = exporter.perform_export()
    OrderedSetsExportWriter(export_rows).write_csv(response)


def import_ordered_sets(file, filetype, progress_queue, purge=False):
    def create_ordered_set_from_row(index, row):
        set_name = row["Name"]

        ordered_set = OrderedContentSet.objects.filter(name=set_name).first()
        if not ordered_set:
            ordered_set = OrderedContentSet(name=set_name)

        ordered_set.profile_fields = []
        for field in [f.strip() for f in (row["Profile Fields"] or "").split(",")]:
            if not field or field == "-":
                continue
            [field_name, field_value] = field.split(":")
            ordered_set.profile_fields.append((field_name, field_value))

        ordered_set.pages = []

        times = [p.strip() for p in row["Time"].split(",")]
        units = [p.strip() for p in row["Unit"].split(",")]
        before_or_afters = [p.strip() for p in row["Before Or After"].split(",")]
        page_slugs = [p.strip() for p in row["Page Slugs"].split(",")]
        contact_fields = row["Contact Field"].split(",")
        contact_fields = (
            [p.strip() for p in contact_fields]
            if len(contact_fields) > 1
            else [contact_fields[0]] * len(times)
        )
        if (
            len(times) != 0
            and len(times) != len(units)
            or len(times) != len(before_or_afters)
            or len(times) != len(page_slugs)
            or len(times) != len(contact_fields)
        ):
            raise ImportException(
                f"Row {row['Name']} has {len(times)} times, {len(units)} units, {len(before_or_afters)} before_or_afters, {len(page_slugs)} page_slugs and {len(contact_fields)} contact_fields and they should all be equal.",
                index,
            )
        for idx, page_slug in enumerate(page_slugs):
            if not page_slug or page_slug == "-":
                continue
            page = ContentPage.objects.filter(slug=page_slug).first()
            time = times[idx]
            unit = units[idx]
            before_or_after = before_or_afters[idx]
            contact_field = contact_fields[idx]
            if page:
                ordered_set.pages.append(
                    (
                        "pages",
                        {
                            "contentpage": page,
                            "time": time or "",
                            "unit": unit or "",
                            "before_or_after": before_or_after or "",
                            "contact_field": contact_field or "",
                        },
                    )
                )
            else:
                logger.warning(f"Content page not found for slug '{page_slug}'")

        ordered_set.save()
        return ordered_set

    file = file.read()
    lines = []
    if filetype == "XLSX":
        wb = load_workbook(filename=BytesIO(file))
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        for row in ws.iter_rows(values_only=True):
            row_dict = {
                "Name": row[0],
                "Profile Fields": row[1],
                "Page Slugs": row[2],
                "Time": row[3],
                "Unit": row[4],
                "Before Or After": row[5],
                "Contact Field": row[6],
            }
            lines.append(row_dict)
    else:
        if isinstance(file, bytes):
            try:
                file = file.decode("utf-8")
            except UnicodeDecodeError:
                file = file.decode("latin-1")

        reader = csv.DictReader(io.StringIO(file))
        for dictionary in reader:
            lines.append(dictionary)

    # 10% progress for loading file
    progress_queue.put_nowait(10)

    for index, row in enumerate(lines):
        os = create_ordered_set_from_row(index, row)
        if not os:
            print(f"Ordered Content Set not created for row {index + 1}")
        # 10-100% for loading ordered content sets
        progress_queue.put_nowait(10 + index * 90 / len(lines))
