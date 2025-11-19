import copy
import csv
import io
from collections.abc import Iterable, Iterator
from dataclasses import asdict, astuple, dataclass, fields
from math import ceil

from django.http import HttpResponse  # type: ignore  # No typing available
from openpyxl.styles import Border, Font, NamedStyle, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from wagtail.query import PageQuerySet  # type: ignore  # No typing available


@dataclass
class ExportRow:
    """
    All the data for a single row of an assessment export
    """

    title: str
    question_type: str
    tags: str
    slug: str
    version: str
    language_code: str
    high_result_page: str | None
    high_inflection: str | None
    medium_result_page: str | None
    medium_inflection: str | None
    low_result_page: str | None
    skip_threshold: str | None
    skip_high_result_page: str | None
    generic_error: str
    question: str
    explainer: str
    error: str
    min: int
    max: int
    answers: str
    scores: str
    answer_semantic_ids: str
    question_semantic_id: str
    answer_responses: str

    @classmethod
    def headings(cls) -> list[str]:
        """
        The field names, used to write the header rows in exports
        """
        return [f.name for f in fields(cls)]

    def to_dict(self) -> dict[str, str]:
        return asdict(self)

    def to_tuple(self) -> tuple[str]:
        return astuple(self)


class AssessmentExporter:
    def __init__(self, queryset: PageQuerySet):
        self.queryset = queryset

    def perform_export(self) -> Iterable[ExportRow]:
        """
        Converts the queryset into an iterable of ExportRows, ready to be written
        """
        for item in self.queryset:
            for question in item.questions:
                answers = [a["answer"] for a in question.value.get("answers", [])]
                scores = [a["score"] for a in question.value.get("answers", [])]
                answer_semantic_ids = [
                    a["semantic_id"] for a in question.value.get("answers", [])
                ]
                answer_responses = [
                    a["response"] for a in question.value.get("answers") or []
                ]
                yield ExportRow(
                    title=item.title,
                    tags=serialize_list(
                        filter_non_empty(t.name for t in item.tags.all())
                    ),
                    question_type=question.block_type,
                    slug=item.slug,
                    version=item.version,
                    language_code=item.locale.language_code,
                    high_result_page=getattr(item.high_result_page, "slug", None),
                    high_inflection=getattr(item, "high_inflection", None),
                    medium_result_page=getattr(item.medium_result_page, "slug", None),
                    medium_inflection=getattr(item, "medium_inflection", None),
                    low_result_page=getattr(item.low_result_page, "slug", None),
                    skip_threshold=getattr(item, "skip_threshold", "0.0"),
                    skip_high_result_page=getattr(
                        item.skip_high_result_page, "slug", None
                    ),
                    generic_error=item.generic_error,
                    question=question.value["question"],
                    explainer=question.value["explainer"],
                    error=question.value.get("error"),
                    min=question.value.get("min"),
                    max=question.value.get("max"),
                    answers=serialize_list(answers),
                    scores=serialize_list(scores),
                    answer_semantic_ids=serialize_list(answer_semantic_ids),
                    question_semantic_id=question.value.get("semantic_id"),
                    answer_responses=serialize_list(answer_responses),
                )


def filter_non_empty(items: Iterable[str]) -> Iterator[str]:
    """
    Ensures only truthy values are present in the iterable
    """
    for item in items:
        if item:
            yield item


def serialize_list(items: Iterable[str]) -> str:
    """
    Uses CSV formatting to seralize a list of strings, handling escaping
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(items)
    return output.getvalue().rstrip("\r\n")


class AssessmentExportWriter:
    """
    Responsible for taking an in-memory representation of export rows, and writing it
    out in our required export format
    """

    def __init__(self, rows: Iterable[ExportRow]):
        self.rows = rows

    def write_xlsx(self, response: HttpResponse) -> None:
        workbook = Workbook()
        worksheet: Worksheet = workbook.active  # type: ignore  # This is not a write or read only workbook
        worksheet.append(ExportRow.headings())
        for row in self.rows:
            worksheet.append(row.to_tuple())
        _set_xlsx_styles(workbook, worksheet)
        workbook.save(response)  # type: ignore

    def write_csv(self, response: HttpResponse) -> None:
        writer = csv.DictWriter(f=response, fieldnames=ExportRow.headings())
        writer.writeheader()

        for row in self.rows:
            writer.writerow(row.to_dict())


def _set_xlsx_styles(wb: Workbook, sheet: Worksheet) -> None:
    """
    Sets the style for the workbook adding any formatting that will make the sheet more
    aesthetically pleasing
    """
    # Adjustment is because the size in openxlsx and google sheets are not equivalent
    adjustment = 7

    # Set columns based on best size
    column_widths_in_pts = {
        "title": 110,
        "question_type": 110,
        "tags": 110,
        "slug": 110,
        "version": 90,
        "language_code": 50,
        "high_result_page": 110,
        "high_inflection": 110,
        "medium_result_page": 120,
        "medium_inflection": 110,
        "low_result_page": 110,
        "skip_threshold": 110,
        "skip_high_result_page": 110,
        "generic_error": 370,
        "question": 370,
        "explainer": 370,
        "error": 370,
        "min": 110,
        "max": 110,
        "answers": 370,
        "scores": 110,
        "answer_semantic_ids": 110,
        "question_semantic_id": 110,
        "answer_responses": 110,
    }
    for column in sheet.iter_cols(max_row=1):
        [cell] = column
        width = column_widths_in_pts[str(cell.value)]
        sheet.column_dimensions[get_column_letter(cell.col_idx)].width = ceil(
            width / adjustment
        )

    # Freeze heading row and side panel, 1 added because it freezes before the column
    panel_column = get_column_letter(5)
    sheet.freeze_panes = sheet[f"{panel_column}2"]

    # Borders
    left_border = Border(left=Side(border_style="thin", color="FF000000"))

    # Named Styles
    header_style = NamedStyle(name="header_style")

    # Set attributes to styles
    header_style.font = Font(bold=True, size=10)

    # Add named styles to wb
    wb.add_named_style(header_style)

    # Set header style for row 1
    for row in sheet["1:2"]:
        for cell in row:
            cell.style = header_style

    # Set dividing border for side panel
    for cell in sheet[f"{panel_column}:{panel_column}"]:
        cell.border = left_border

    # set font on all cells initially to 10pt and row height
    general_font = Font(size=10)
    for index, row in enumerate(sheet.iter_rows()):
        if index > 2:
            sheet.row_dimensions[index].height = 60  # type: ignore # Bad annotation.
        for cell in row:
            cell.font = general_font
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment  # type: ignore # Broken typeshed update, maybe?
