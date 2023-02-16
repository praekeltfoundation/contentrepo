import copy
import csv
import io
from dataclasses import dataclass
from io import BytesIO
from math import ceil
from typing import List, Tuple, Union

from django.http import HttpResponse
from django.db.models.query import QuerySet
from openpyxl import load_workbook
from openpyxl.styles import Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from taggit.models import Tag
from wagtail import blocks
from wagtail.documents.blocks import DocumentChooserBlock
from wagtail.documents.models import Document
from wagtail.images.blocks import ImageChooserBlock
from wagtail.images.models import Image
from wagtail.models import Locale
from wagtail.models.sites import Site
from wagtail.query import PageQuerySet
from wagtail.rich_text import RichText
from wagtailmedia.models import Media

from home.models import (  # isort:skip
    ContentPage,
    ContentPageIndex,
    ContentQuickReply,
    ContentTrigger,
    HomePage,
    MediaBlock,
    Page,
    VariationBlock,
    OrderedContentSet,
)


EXPORT_FIELDNAMES = [
    "page_id",
    "slug",
    "parent",
    "web_title",
    "web_subtitle",
    "web_body",
    "whatsapp_title",
    "whatsapp_body",
    "variation_title",
    "variation_body",
    "messenger_title",
    "messenger_body",
    "viber_title",
    "viber_body",
    "translation_tag",
    "tags",
    "quick_replies",
    "triggers",
    "locale",
    "next_prompt",
    "image_link",
    "doc_link",
    "media_link",
    "related_pages",
]


def import_content(file, filetype, purge=True, locale="en"):
    def set_variation_blocks(body_values, message_number):
        variation_blocks = []
        for variation_message in variation_messages:
            if variation_message["message_number"] == str(message_number):
                profile_field, profile_field_value = variation_message[
                    "variation_title"
                ].split(": ")
                site = Site.objects.get(is_default_site=True)
                for profile_block in site.sitesettings.profile_field_options:
                    if profile_block.block_type == profile_field:
                        variation_blocks.append(
                            {
                                "variation_restrictions": [
                                    {
                                        "type": profile_field,
                                        "value": profile_field_value,
                                    }
                                ],
                                "message": variation_message["variation_body"],
                            }
                        )
                        break
            body_values["variation_messages"] = variation_blocks
        return body_values

    def get_rich_text_body(text=None):
        if text:
            body = []
            lines = text.splitlines()
            for line in lines:
                if len(line) != 0:
                    body = body + [("paragraph", RichText(line))]
            return body

    def get_body(messages, type_of_message, variation_messages=[]):
        struct_blocks = []
        for message_number, raw in enumerate(messages, 1):
            im = None
            doc = None
            media = None
            if not raw:
                return struct_blocks
            split_body = raw.split("/n")
            for line in split_body:
                if "image_link:" in line:
                    link = raw[raw.index(":") + 1 :]
                    im = Image.objects.get(url=link).id
                if "doc_link:" in line:
                    link = raw[raw.index(":") + 1 :]
                    doc = Document.objects.get(url=link).id
                if "media_link:" in line:
                    link = raw[raw.index(":") + 1 :]
                    doc = Media.objects.get(url=link).id
            message_body = raw
            if message_body:
                body_blocks = [
                    ("message", blocks.TextBlock()),
                ]
                if im:
                    body_blocks.append(("image", ImageChooserBlock()))
                if doc:
                    body_blocks.append(
                        ("document", DocumentChooserBlock()),
                    )
                if media:
                    body_blocks.append(
                        ("media", MediaBlock()),
                    )
                if type_of_message == "Whatsapp_Message" and variation_messages:
                    body_blocks.append(
                        ("variation_messages", blocks.ListBlock(VariationBlock()))
                    )
                block = blocks.StructBlock(body_blocks)
                body_values = {
                    "message": message_body,
                }
                if im:
                    body_values["image"] = im
                if doc:
                    body_values["document"] = doc
                if media:
                    body_values["media"] = media

                if type_of_message == "Whatsapp_Message" and variation_messages:
                    body_values = set_variation_blocks(
                        body_values=body_values,
                        message_number=message_number,
                    )

                block_value = block.to_python(body_values)
                struct_blocks.append((type_of_message, block_value))
        return struct_blocks

    def create_tags(row, page):
        if row["tags"]:
            tags = row["tags"].split(",")
            for tag in tags:
                created_tag, _ = Tag.objects.get_or_create(name=tag.strip())
                page.tags.add(created_tag)
        else:
            # clear tags if content sheet has no tags
            page.tags.clear()

    def add_quick_replies(row, page):
        if row["quick_replies"]:
            replies = row["quick_replies"].split(",")
            for reply in replies:
                created_tag, _ = ContentQuickReply.objects.get_or_create(
                    name=reply.strip()
                )
                page.quick_replies.add(created_tag)

    def add_triggers(row, page):
        if row["triggers"]:
            triggers = row["triggers"].split(",")
            for trigger in triggers:
                created_tag, _ = ContentTrigger.objects.get_or_create(
                    name=trigger.strip()
                )
                page.triggers.add(created_tag)

    def add_parent(row, page, home_page):
        if row["parent"]:
            parent = Page.objects.filter(title=row["parent"], locale=page.locale).last()
            parent.add_child(instance=page)
        else:
            home_page.add_child(instance=page)

    def add_web(row, page=None):
        if "web_title" not in row.keys() or not row["web_title"]:
            return

        if page:  # update
            page.title = row["web_title"]
            page.subtitle = row["web_subtitle"]
            page.body = get_rich_text_body(row["web_body"])
            page.enable_web = True
            page.locale = home_page.locale
            return page
        return ContentPage(
            title=row["web_title"],
            slug=row["slug"],
            subtitle=row["web_subtitle"],
            body=get_rich_text_body(row["web_body"]),
            enable_web=True,
            locale=home_page.locale,
        )

    def add_whatsapp(row, whatsapp_messages, page=None, variation_messages=[]):
        if "whatsapp_title" not in row.keys() or not row["whatsapp_title"]:
            return page

        if not page:
            return ContentPage(
                title=row["whatsapp_title"],
                slug=row["slug"],
                enable_whatsapp=True,
                whatsapp_title=row["whatsapp_title"],
                whatsapp_body=get_body(
                    whatsapp_messages, "Whatsapp_Message", variation_messages
                ),
                locale=home_page.locale,
            )
        else:
            page.enable_whatsapp = True
            page.whatsapp_title = row["whatsapp_title"]
            page.whatsapp_body = get_body(
                whatsapp_messages, "Whatsapp_Message", variation_messages
            )
            return page

    def add_messenger(row, messenger_messages, page=None):
        if "messenger_title" not in row.keys() or not row["messenger_title"]:
            return page

        if not page:
            return ContentPage(
                title=row["messenger_title"],
                slug=row["slug"],
                enable_messenger=True,
                messenger_title=row["messenger_title"],
                messenger_body=get_body(messenger_messages, "messenger_block"),
                locale=home_page.locale,
            )
        else:
            page.enable_messenger = True
            page.messenger_title = row["messenger_title"]
            page.messenger_body = get_body(messenger_messages, "messenger_block")
            return page

    def add_viber(row, viber_messages, page=None):
        if "viber_title" not in row.keys() or not row["viber_title"]:
            return page

        if not page:
            return ContentPage(
                title=row["viber_title"],
                slug=row["slug"],
                enable_viberr=True,
                viber_title=row["viber_title"],
                viber_body=get_body(viber_messages, "viber_message"),
                locale=home_page.locale,
            )
        else:
            page.enable_viber = True
            page.viber_title = row["viber_title"]
            page.viber_body = get_body(viber_messages, "viber_message")
            return page

    def clean_row(row):
        for field in (
            "web_title",
            "web_subtitle",
            "web_body",
            "whatsapp_title",
            "whatsapp_body",
            "variation_title",
            "variation_body",
            "next_prompt",
            "viber_title",
            "viber_body",
            "messenger_title",
            "messenger_body",
            "parent",
        ):
            if field in row and row[field]:
                row[field] = str(row[field]).strip()
        return row

    def get_side_panel_width(ws, column_heading):
        column_lookup = {}
        iterator = 1
        for column in ws.iter_cols(1, ws.max_column):
            column_lookup[column[1].value] = iterator
            iterator += 1
        if column_heading in column_lookup:
            return column_lookup[column_heading]
        raise KeyError(f"{column_heading} not found in column headings {column_lookup}")

    file = file.read()
    lines = []
    if filetype == "XLSX":
        wb = load_workbook(filename=BytesIO(file))
        ws = wb.worksheets[0]
        side_pannel = get_side_panel_width(ws, "message")
        if side_pannel:
            ws.delete_cols(1, side_pannel)
        ws.delete_rows(1, 2)
        for row in ws.iter_rows():
            row_dict = {}
            for index, x in enumerate(row):
                row_dict[EXPORT_FIELDNAMES[index]] = (
                    x.value.replace("_x000D_", "") if x.value else None
                )
            lines.append(row_dict)
    else:
        if isinstance(file, bytes):
            file = file.decode("utf-8")
        reader = csv.DictReader(io.StringIO(file))
        for dictionary in reader:
            lines.append(dictionary)

    if purge in ["True", "yes", True]:
        ContentPage.objects.all().delete()
        ContentPageIndex.objects.all().delete()

    if isinstance(locale, str):
        locale = Locale.objects.get(language_code=locale)

    home_page = HomePage.objects.get(locale__pk=locale.pk)

    for index, row in enumerate(lines):
        slug = row["slug"]

        if row["web_title"] in ["", None]:
            continue
        if (
            row["parent"] in ["", None]
            and row["web_body"] in ["", None]
            and row["whatsapp_body"] in ["", None]
            and row["messenger_body"] in ["", None]
        ):
            cpi = ContentPageIndex.objects.filter(slug=slug).first()
            if cpi:
                cpi.title = row["web_title"]
                cpi.save_revision().publish()
            else:
                cpi = ContentPageIndex(title=row["web_title"], slug=slug)
                home_page.add_child(instance=cpi)
                cpi.save_revision().publish()
            continue
        row = clean_row(row)
        variation_messages = []
        whatsapp_messages = [row["whatsapp_body"]]
        messenger_messages = [row["messenger_body"]]
        viber_messages = [row["viber_body"]]
        for next_row in lines[index + 1 :]:
            if next_row["web_title"] not in ["", None]:
                break
            if next_row["whatsapp_body"] not in ["", None]:
                whatsapp_messages.append(next_row["whatsapp_body"])
            if next_row["variation_body"] not in ["", None]:
                variation_messages.append(
                    {
                        "variation_body": next_row["variation_body"],
                        "variation_title": next_row["variation_title"],
                        "message_number": next_row["message"],
                    }
                )

            if next_row["messenger_body"] not in ["", None]:
                messenger_messages.append(next_row["messenger_body"])
            if next_row["viber_body"] not in ["", None]:
                viber_messages.append(next_row["viber_body"])

        exiting_contentpage = ContentPage.objects.filter(slug=slug).first()

        contentpage = add_web(row=row, page=exiting_contentpage)
        contentpage = add_whatsapp(
            row=row,
            whatsapp_messages=whatsapp_messages,
            page=contentpage,
            variation_messages=variation_messages,
        )
        contentpage = add_messenger(
            row=row, messenger_messages=messenger_messages, page=contentpage
        )
        contentpage = add_viber(
            row=row, viber_messages=viber_messages, page=contentpage
        )
        if contentpage:
            create_tags(row, contentpage)
            add_quick_replies(row, contentpage)
            add_triggers(row, contentpage)
            if not exiting_contentpage:
                add_parent(row, contentpage, home_page)
            contentpage.save_revision().publish()
            try:
                pages = contentpage.tags.first().home_contentpagetag_items.all()
                for page in pages:
                    if page.content_object.pk != contentpage.pk:
                        contentpage.translation_key = (
                            page.content_object.translation_key
                        )
                        contentpage.save_revision().publish()
            except Exception as e:
                print(e)
        else:
            print(f"Content page not created for {row}")


def style_sheet(wb: Workbook, sheet: Worksheet) -> Tuple[Workbook, Worksheet]:
    """Sets the style for the workbook adding any formatting that will make the sheet more aesthetically pleasing"""
    # Adjustment is because the size in openxlsx and google sheets are not equivalent
    adjustment = 7
    # Padding
    sheet.insert_cols(1)

    # Set columns based on best size

    column_widths_in_pts = {
        "structure": 110,
        "message": 70,
        "page_id": 110,
        "slug": 110,
        "parent": 110,
        "web_title": 110,
        "web_subtitle": 110,
        "web_body": 370,
        "whatsapp_title": 118,
        "whatsapp_body": 370,
        "variation_title": 118,
        "variation_body": 370,
        "messenger_title": 118,
        "messenger_body": 370,
        "viber_title": 118,
        "viber_body": 370,
        "translation_tag": 118,
        "tags": 118,
        "quick_replies": 118,
        "triggers": 118,
        "locale": 118,
        "next_prompt": 118,
        "image_link": 118,
        "doc_link": 118,
        "media_link": 118,
        "related": 118,
    }

    for index, column_width in enumerate(column_widths_in_pts.values(), 2):
        sheet.column_dimensions[get_column_letter(index)].width = ceil(
            column_width / adjustment
        )

    # Freeze heading row and side panel, 1 added because it freezes before the column
    panel_column = get_column_letter(5)
    sheet.freeze_panes = sheet[f"{panel_column}2"]

    # Colours
    blue = Color(rgb="0099CCFF")

    # Boarders
    left_border = Border(left=Side(border_style="thin", color="FF000000"))

    # Fills
    blue_fill = PatternFill(patternType="solid", fgColor=blue)

    # Named Styles
    header_style = NamedStyle(name="header_style")
    menu_style = NamedStyle(name="menu_style")

    # Set attributes to styles
    header_style.font = Font(bold=True, size=10)
    menu_style.fill = blue_fill
    menu_style.font = Font(bold=True, size=10)

    # Add named styles to wb
    wb.add_named_style(header_style)
    wb.add_named_style(menu_style)

    # column widths

    # Set menu style for any "Menu" row
    for row in sheet.iter_rows():
        if isinstance(row[1].value, str) and "Menu" in row[1].value:
            for cell in row:
                cell.style = menu_style

    # Set header style for row 1 and 2
    for row in sheet["1:2"]:
        for cell in row:
            cell.style = header_style

    # Set dividing border for side panel
    for cell in sheet[f"{panel_column}:{panel_column}"]:
        cell.border = left_border

    # set font on all cells initially to 10pt and row height
    general_font = Font(size=10)
    for index, row in enumerate(sheet.iter_rows()):
        if index > 2:
            sheet.row_dimensions[index].height = 60
        for cell in row:
            cell.font = general_font
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment
    return wb, sheet


def get_rows(
    page: Union[ContentPage, ContentPageIndex],
    structure_string: str,
) -> List[list]:
    """Sets up row for each page including the side panel.
    Each page is returned as a list of rows, this accounts for pages with multiple messages
    """
    rows = []
    base_row = ContentSheetRow.from_page(page=page, structure_string=structure_string)
    if base_row:
        rows.append(base_row.get_row())
        if base_row.variation_messages_rows:
            for variation_row in base_row.variation_messages_rows:
                rows.append(variation_row.get_row())
        if base_row.tuple_of_extra_rows:
            for extra_row in base_row.tuple_of_extra_rows:
                rows.append(extra_row.get_row())
                if extra_row.variation_messages_rows:
                    for variation_row in extra_row.variation_messages_rows:
                        rows.append(variation_row.get_row())
    return rows


def add_children(
    temp_sheet: List[list],
    children: PageQuerySet,
    queryset: PageQuerySet,
    parent_structure: str,
) -> List[list]:
    """Recursive function that traverses the children of a page with a depth first search algorithm"""
    for index, child in enumerate(children, 1):
        content_page = get_page(child, queryset)
        structure_string = f"{parent_structure.replace('Menu','Sub')}.{index}"
        row = get_rows(content_page, structure_string)
        if row:
            temp_sheet.extend(row)
        if content_page and content_page.has_children:
            add_children(
                temp_sheet,
                content_page.get_children(),
                queryset,
                structure_string,
            )
    return temp_sheet


def get_page(page: Union[ContentPage, ContentPageIndex], queryset: PageQuerySet = None):
    if page.content_type.name == "content page index":
        return ContentPageIndex.objects.filter(id=page.id).first()
    elif page.content_type.name == "content page" and queryset:
        return queryset.filter(id=page.id).first()
    else:
        return ContentPage.objects.filter(id=page.id).first()


def get_content_depth(queryset: PageQuerySet) -> list[int]:
    if queryset:
        distance = max([x.depth for x in queryset]) - 1
        headings = [x for x in range(1, distance)]
        return headings
    else:
        return [1, 2, 3, 4, 5]


def get_content_sheet(queryset: PageQuerySet) -> List[list]:
    content_sheet = []
    headings = ["structure", "message"] + EXPORT_FIELDNAMES
    content_sheet.append(headings)
    for locale in Locale.objects.all():
        home = HomePage.objects.filter(locale_id=locale.id).first()
        if home:
            main_menu_pages = home.get_children()
            for index, page in enumerate(main_menu_pages, 1):
                structure_string = f"Menu {index}"
                content_page = get_page(page, queryset)
                if content_page:
                    row = get_rows(content_page, structure_string)
                    content_sheet.extend(row)
                    if content_page.has_children:
                        content_sheet = add_children(
                            content_sheet,
                            content_page.get_children(),
                            queryset,
                            structure_string,
                        )
    return content_sheet


def export_xlsx_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    """Export contentpages within the queryset to an xlsx"""
    workbook = Workbook()
    worksheet = workbook.active

    content_sheet = get_content_sheet(queryset)
    for row in content_sheet:
        worksheet.append(row)
    workbook, worksheet = style_sheet(workbook, worksheet)
    workbook.save(response)


def export_csv_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    """Export contentpages within the queryset to a csv"""
    content_sheet = get_content_sheet(queryset)
    writer = csv.writer(response)
    for row in content_sheet:
        writer.writerow(row)


def get_serialized_ordered_set(ordered_set: OrderedContentSet,) -> List[list]:
    set_as_list = [ordered_set.name]

    profile_fields_list = []
    for field in ordered_set.profile_fields.raw_data:
        profile_fields_list.append(f"{field['type']}:{field['value']}")
    set_as_list.append(", ".join(profile_fields_list))

    pages_list = []
    for page in ordered_set.pages:
        pages_list.append(page.value.slug)
    set_as_list.append(", ".join(pages_list))

    return set_as_list


def get_ordered_set_sheet(queryset: QuerySet) -> List[list]:
    ordered_set_sheet = []
    headings = ["name", "profile fields", "page slugs"]
    ordered_content_sheet.append(headings)
    for ordered_set in queryset:
        ordered_content_sheet.append(get_serialized_ordered_set(ordered_set))
    return ordered_content_sheet


def export_csv_ordered_sets(queryset: QuerySet, response: HttpResponse) -> None:
    """Export ordered content sets within the queryset to a csv"""
    ordered_set_sheet = get_ordered_set_sheet(queryset)
    writer = csv.writer(response)
    for row in ordered_set_sheet:
        writer.writerow(row)


def export_xlsx_ordered_sets(queryset: QuerySet, response: HttpResponse) -> None:
    """Export ordered content sets within the queryset to an xlsx"""
    workbook = Workbook()
    worksheet = workbook.active

    ordered_content_sheet = get_ordered_set_sheet(queryset)
    for row in content_sheet:
        worksheet.append(row)
    workbook.save(response)


def import_ordered_sets(file, filetype):
    def create_ordered_set_from_row(row):
        set_name = row["name"]
        set_profile_fields = []
        for field in [f.strip() for f in row["profile fields"].split(",")]:
            [field_name, field_value] = field.split(":")
            set_profile_fields.append({
                "type": field_name,
                "value": field_value
            })
        set_pages = []
        for page_slug in [p.strip() for p in row["page slugs"].split(",")]:
            page = ContentPage.objects.filter(slug=page_slug).first()
            if page:
                set_pages.append(
                    {"type": "pages", "value": page.id}
                )
            else:
                print(f"Content page not found for slug '{page_slug}'")

        existing_ordered_set = OrderedContentSet.objects.filter(name=set_name).first()
        if existing_ordered_set:
            existing_ordered_set.profile_fields = set_profile_fields
            existing_ordered_set.pages = set_pages
            existing_ordered_set.save()
            ordered_set = existing_ordered_set
        else:
            ordered_set = OrderedContentSet(
                name=set_name,
                profile_fields=set_profile_fields,
                pages=set_pages,
            )
            ordered_set.save()
        return ordered_set or None

    file = file.read()
    lines = []
    if filetype == "XLSX":
        wb = load_workbook(filename=BytesIO(file))
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        for row in ws.iter_rows(values_only=True):
            row_dict = {
                "name": row[0],
                "profile fields": row[1],
                "page slugs": row[2]}
            lines.append(row_dict)
    else:
        if isinstance(file, bytes):
            file = file.decode("utf-8")
        reader = csv.DictReader(io.StringIO(file))
        for dictionary in reader:
            lines.append(dictionary)

    for index, row in enumerate(lines):
        os = create_ordered_set_from_row(row)
        if not os:
            print(f"Ordered Content Set not created for row {index + 1}")


@dataclass
class Message:
    body: str = ""
    image_name: str = None
    media_name: str = None
    document_name: str = None
    next_prompt: str = None

    @classmethod
    def from_platform_body_element(
        cls, platform_body_element: blocks.StreamValue.StreamChild
    ):
        message = cls()
        message.body = (
            platform_body_element.value["message"].strip()
            if "message" in platform_body_element.value
            else None
        )
        message.image_name = (
            platform_body_element.value["image"]
            if "image" in platform_body_element.value
            else None
        )
        message.document_name = (
            platform_body_element.value["document"]
            if "document" in platform_body_element.value
            else None
        )
        message.media_name = (
            platform_body_element.value["media"]
            if "media" in platform_body_element.value
            else None
        )
        message.next_prompt = (
            platform_body_element.value["next_prompt"]
            if "next_prompt" in platform_body_element.value
            else None
        )
        return message


@dataclass
class VariationMessage:
    body: str = ""
    title: str = ""

    @classmethod
    def from_variation(cls, variation: blocks.StreamValue.StreamChild):
        message = cls()
        message.body = variation.get("message")
        profile_fields = []
        for profile_field in variation.get("variation_restrictions").raw_data:
            profile_fields.append(f"{profile_field['type']}: {profile_field['value']}")
        message.title = ", ".join(profile_fields)
        return message


@dataclass
class VariationMessageList:
    variations: List[VariationMessage]

    @classmethod
    def from_platform_body_element(cls, whatsapp_msg: blocks.StreamValue.StreamChild):
        variations = []
        for variation in whatsapp_msg.value["variation_messages"]:
            variations.append(VariationMessage.from_variation(variation=variation))
        return cls(variations)


@dataclass
class MessageContainer:
    whatsapp: Tuple[Message]
    messenger: Tuple[Message]
    viber: Tuple[Message]
    variation_messages: Tuple[VariationMessage]

    @classmethod
    def from_platform_body(cls, whatsapp_body, messenger_body, viber_body):
        whatsapp = []
        whatsapp_variation_messages = []
        messenger = []
        viber = []
        for whatsapp_msg in whatsapp_body:
            whatsapp.append(Message.from_platform_body_element(whatsapp_msg))
            whatsapp_variation_messages.append(
                VariationMessageList.from_platform_body_element(whatsapp_msg)
            )

        for messenger_msg in messenger_body:
            messenger.append(Message.from_platform_body_element(messenger_msg))

        for viber_msg in viber_body:
            viber.append(Message.from_platform_body_element(viber_msg))
        return cls(whatsapp, messenger, viber, whatsapp_variation_messages)

    def find_first_attachment(self, index: int, attachment_type: str) -> str:
        for message_list in [self.whatsapp, self.messenger, self.viber]:
            try:
                value = getattr(message_list[index], attachment_type)
                if value:
                    return value
            except IndexError:
                continue
        return ""

    @staticmethod
    def message_from_platform_body_element(
        platform_body_element: blocks.StreamValue.StreamChild,
    ):
        message = {}
        message["message"] = (
            platform_body_element.value["message"].strip()
            if "message" in platform_body_element.value
            else None
        )
        message["image_name"] = (
            platform_body_element.value["image_name"].strip()
            if "image_name" in platform_body_element.value
            else None
        )
        message["document_name"] = (
            platform_body_element.value["document_name"].strip()
            if "document_name" in platform_body_element.value
            else None
        )
        message["media_name"] = (
            platform_body_element.value["media_name"].strip()
            if "media_name" in platform_body_element.value
            else None
        )
        message["next_prompt"] = (
            platform_body_element.value["next_prompt"].strip()
            if "next_prompt" in platform_body_element.value
            else None
        )

        return message


@dataclass
class ContentSheetRow:
    structure: str = ""
    side_panel_message_number: int = 0
    parent: str = ""
    slug: str = ""
    web_title: str = ""
    web_subtitle: str = ""
    web_body: str = ""
    whatsapp_title: str = ""
    whatsapp_body: str = ""
    messenger_title: str = ""
    messenger_body: str = ""
    viber_title: str = ""
    viber_body: str = ""
    translation_tag: str = ""
    tags: str = ""
    quick_replies: str = ""
    triggers: str = ""
    locale: str = ""
    next_prompt: str = ""
    image_link: str = ""
    doc_link: str = ""
    media_link: str = ""
    related_pages: str = ""
    variation_body: str = ""
    variation_title: str = ""
    tuple_of_extra_rows: tuple = ()
    variation_messages_rows: tuple = ()
    page_id: int = 0

    @classmethod
    def from_page(
        cls,
        page: ContentPage,
        structure_string: str,
        side_panel_message_number: int = 0,
    ):
        if isinstance(page, ContentPage):
            return cls.from_content_page(
                page, structure_string, side_panel_message_number
            )
        if isinstance(page, ContentPageIndex):
            return cls.from_index_page(page, structure_string)

    @classmethod
    def from_content_page(
        cls,
        page: ContentPage,
        structure_string: str,
        side_panel_message_number: int = 0,
    ):
        content_sheet_row = cls()
        message_container = MessageContainer.from_platform_body(
            page.whatsapp_body,
            page.messenger_body,
            page.viber_body,
        )
        content_sheet_row._set_messages(
            page, side_panel_message_number, message_container
        )

        content_sheet_row.structure = structure_string
        content_sheet_row.page_id = page.id
        content_sheet_row.slug = page.slug

        if side_panel_message_number == 0:
            content_sheet_row.parent = content_sheet_row._get_parent_page(page)
            content_sheet_row.web_title = page.title
            content_sheet_row.web_subtitle = page.subtitle
            content_sheet_row.web_body = str(page.body)
            content_sheet_row.whatsapp_title = page.whatsapp_title
            content_sheet_row.messenger_title = page.messenger_title
            content_sheet_row.viber_title = page.viber_title
            content_sheet_row.translation_tag = str(page.translation_key)
            content_sheet_row.tags = content_sheet_row._format_list_from_query_set(
                page.tags.all()
            )
            content_sheet_row.quick_replies = (
                content_sheet_row._format_list_from_query_set(page.quick_replies.all())
            )
            content_sheet_row.triggers = content_sheet_row._format_list_from_query_set(
                page.triggers.all()
            )
            content_sheet_row.locale = str(page.locale)
            content_sheet_row.related_pages = (
                content_sheet_row._format_list_from_query_set(
                    content_sheet_row._get_related_pages(page)
                )
            )
        return content_sheet_row

    @classmethod
    def from_index_page(
        cls,
        page: ContentPageIndex,
        structure_string: str,
    ):
        content_sheet_row = cls()
        content_sheet_row.page_id = page.id
        content_sheet_row.slug = page.slug
        content_sheet_row.structure = structure_string
        content_sheet_row.parent = content_sheet_row._get_parent_page(page)
        content_sheet_row.web_title = page.title
        content_sheet_row.translation_tag = str(page.translation_key)
        content_sheet_row.locale = str(page.locale)
        return content_sheet_row

    @classmethod
    def from_variation_message(
        cls,
        page,
        variation_title,
        side_panel_message_number,
        variation_body,
    ):
        content_sheet_row = cls()
        content_sheet_row.page_id = page.id
        content_sheet_row.slug = page.slug
        content_sheet_row.side_panel_message_number = side_panel_message_number + 1
        content_sheet_row.variation_title = variation_title
        content_sheet_row.variation_body = variation_body
        return content_sheet_row

    def get_row(self):
        """Returns object as an exportable row"""
        return [
            self.structure,
            self.side_panel_message_number,
            self.page_id,
            self.slug,
            self.parent,
            self.web_title,
            self.web_subtitle,
            self.web_body,
            self.whatsapp_title,
            self.whatsapp_body,
            self.variation_title,
            self.variation_body,
            self.messenger_title,
            self.messenger_body,
            self.viber_title,
            self.viber_body,
            self.translation_tag,
            self.tags,
            self.quick_replies,
            self.triggers,
            self.locale,
            self.next_prompt,
            self.image_link,
            self.doc_link,
            self.media_link,
            self.related_pages,
        ]

    @staticmethod
    def _format_list_from_query_set(unformatted_query: PageQuerySet) -> str:
        list_delimiter = ", "
        return list_delimiter.join(str(x) for x in unformatted_query if str(x) != "")

    @staticmethod
    def _get_parent_page(page: ContentPage) -> str:
        """Get parent page title as string"""
        if not HomePage.objects.filter(id=page.get_parent().id).exists():
            return page.get_parent().title

    def _set_messages(
        self,
        page: ContentPage,
        side_panel_message_number: int,
        message_container: MessageContainer,
    ):
        """Sets all message level content, including any message level content such as documents, images, next prompts and media.
        Takes the possibility of different numbers of messages on different platforms.
        Adds multiple messages as blank rows with only the extra messages and the message number to the row to match the template
        """
        self.side_panel_message_number = side_panel_message_number + 1
        most_messages = max(
            [
                len(message_container.whatsapp),
                len(message_container.messenger),
                len(message_container.viber),
                len(page.body),
            ]
        )
        variation_messages = []

        if len(message_container.whatsapp) == 1 or (
            side_panel_message_number == 0 and len(message_container.whatsapp) > 1
        ):
            self.whatsapp_body = message_container.whatsapp[0].body
            if message_container.variation_messages:
                variation_messages.extend(
                    self._make_variation_rows(page, message_container, 0)
                )

        if len(message_container.messenger) == 1 or (
            side_panel_message_number == 0 and len(message_container.messenger) > 1
        ):
            self.messenger_body = message_container.messenger[0].body

        if len(message_container.viber) == 1 or (
            side_panel_message_number == 0 and len(message_container.viber) > 1
        ):
            self.viber_body = message_container.viber[0].body

        self.next_prompt = self._get_next_prompt(message_container)
        self.doc_link = self._get_doc_link(message_container)
        self.image_link = self._get_image_link(message_container)
        self.media_link = self._get_media_link(message_container)

        temp_rows = []
        if side_panel_message_number == 0 and most_messages > 1:
            # Set tuple_of_extra_rows rows to account for multiple messages
            for message_index in range(1, most_messages):
                new_content_sheet_row = ContentSheetRow.from_page(
                    page=page,
                    structure_string="",
                    side_panel_message_number=message_index,
                )
                if message_index < len(message_container.whatsapp):
                    new_content_sheet_row.whatsapp_body = message_container.whatsapp[
                        message_index
                    ].body
                    if message_container.variation_messages:
                        new_content_sheet_row.variation_messages_rows = (
                            self._make_variation_rows(
                                page, message_container, message_index
                            )
                        )

                if message_index < len(message_container.messenger):
                    new_content_sheet_row.messenger_body = message_container.messenger[
                        message_index
                    ].body
                if message_index < len(message_container.viber):
                    new_content_sheet_row.viber_body = message_container.viber[
                        message_index
                    ].body
                new_content_sheet_row.next_prompt = self._get_next_prompt(
                    message_container, message_index
                )
                new_content_sheet_row.doc_link = self._get_doc_link(
                    message_container, message_index
                )
                new_content_sheet_row.image_link = self._get_image_link(
                    message_container, message_index
                )
                new_content_sheet_row.media_link = self._get_media_link(
                    message_container, message_index
                )
                temp_rows.append(new_content_sheet_row)
            self.tuple_of_extra_rows = tuple(temp_rows)
            self.variation_messages_rows = tuple(variation_messages)

    def _make_variation_rows(self, page, message_container, message_index):
        temp_rows = []
        for variation in message_container.variation_messages[message_index].variations:
            temp_rows.append(
                ContentSheetRow.from_variation_message(
                    page=page,
                    variation_title=variation.title,
                    side_panel_message_number=message_index,
                    variation_body=variation.body,
                )
            )
        return temp_rows

    @staticmethod
    def _get_related_pages(page: ContentPage) -> List[str]:
        """Returns a list of strings of the related page slugs"""
        related_pages = []
        for related_page in page.related_pages:
            related_pages.append(related_page.value.slug)
        return related_pages

    def _get_image_link(
        self, message_container: MessageContainer, index: int = 0
    ) -> str:
        """Iterate over a dict of whatsapp, messenger and viber messages to find a valid image for that message level,
        if an image is found in any of the platforms, the url will be saved to the sheet.
        This will take the first one found, can be extended to a list of urls
        """
        image_name = message_container.find_first_attachment(index, "image_name")
        image = Image.objects.filter(title=image_name).first()
        if image:
            return image.usage_url
        return ""

    def _get_doc_link(self, message_container: MessageContainer, index: int = 0) -> str:
        """Iterate over a dict of all whatsapp, messenger and viber messages to find a valid document,
        if a document is found in any of the platforms, the url will be saved to the sheet
        This will take the first one found, can be extended to a list of urls
        """
        document_name = message_container.find_first_attachment(index, "document_name")
        document = Document.objects.filter(title=document_name).first()
        if document:
            return document.usage_url
        return ""

    def _get_media_link(
        self, message_container: MessageContainer, index: int = 0
    ) -> str:
        """Iterate over a dict of all whatsapp, messenger and viber messages to find a valid media,
        if a media is found in any of the platforms, the url will be saved to the sheet
        This will take the first one found, can be extended to a list of urls
        """
        media_name = message_container.find_first_attachment(index, "media_name")
        media = Media.objects.filter(title=media_name).first()
        if media:
            return media.usage_url
        return ""

    def _get_next_prompt(
        self, message_container: MessageContainer, index: int = 0
    ) -> str:
        """Iterate over a dict of all whatsapp, messenger and viber messages to find next prompts,
        if a next prompt is found in any of the platforms, the url will be saved to the sheet
        This will take the one found, can be extended to a list of urls
        """
        next_prompt = message_container.find_first_attachment(index, "next_prompt")
        if next_prompt:
            return next_prompt
        return ""
