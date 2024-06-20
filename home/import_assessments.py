import contextlib
import csv
from dataclasses import dataclass, field, fields
from datetime import datetime
from io import BytesIO, StringIO
from queue import Queue
from typing import Any

from django.core.exceptions import ObjectDoesNotExist, ValidationError  # type: ignore
from openpyxl import load_workbook
from taggit.models import Tag  # type: ignore
from treebeard.exceptions import NodeAlreadySaved  # type: ignore
from wagtail.coreutils import get_content_languages  # type: ignore
from wagtail.models import Locale, Page  # type: ignore

from home.models import Assessment, ContentPage, HomePage  # type: ignore

AssessmentId = tuple[str, Locale]


class ImportAssessmentException(Exception):
    """
    Base exception for all import related issues.
    """

    def __init__(self, message: str, row_num: int | None = None):
        self.row_num = row_num
        self.message = message
        super().__init__()


class AssessmentImporter:
    def __init__(
        self,
        file_content: bytes,
        file_type: str,
        progress_queue: Queue[int],
        purge: bool | str = True,
        locale: Locale | str | None = None,
    ):
        if isinstance(locale, str):
            locale = Locale.objects.get(language_code=locale)
        self.file_content = file_content
        self.file_type = file_type
        self.progress_queue = progress_queue
        self.purge = purge in ["True", "yes", True]
        self.locale = locale
        self.locale_map: dict[str, Locale] = {}
        self.shadow_assessments: dict[AssessmentId, ShadowAssessment] = {}

    def locale_from_language_code(self, lang_code_entry: str) -> Locale:
        if lang_code_entry not in self.locale_map:
            codes = []
            lang_name = ""
            for lang_code, lang_dn in get_content_languages().items():
                if lang_code == lang_code_entry or lang_dn == lang_code_entry:
                    lang_name = lang_dn
                    codes.append(lang_code)
            if not codes:
                raise ImportAssessmentException(
                    f"Language code not found: {lang_code_entry}"
                )
            if len(codes) > 1:
                raise ImportAssessmentException(
                    f"Multiple codes for language: {lang_name} -> {codes}"
                )
            self.locale_map[lang_code_entry] = Locale.objects.get(
                language_code=codes[0]
            )
        return self.locale_map[lang_code_entry]

    def perform_import(self) -> None:
        rows = self.parse_file()
        self.set_progress("Loaded file", 5)

        if self.purge:
            self.delete_existing_assessments()
        self.set_progress("Deleted existing assessment", 10)

        self.process_rows(rows)
        self.save_assessments()

    def process_rows(self, rows: list["AssessmentRow"]) -> None:
        for i, row in enumerate(rows, start=2):
            try:
                self.create_shadow_assessment_from_row(row, i)
            except ImportAssessmentException as e:
                e.row_num = i
                raise e

    def save_assessments(self) -> None:
        for i, assessment in enumerate(reversed(self.shadow_assessments.values())):
            parent = self.home_page(assessment.locale)
            assessment.save(parent)
            self.set_progress(
                "Importing assessments", 10 + 70 * i // len(self.shadow_assessments)
            )

    def parse_file(self) -> list["AssessmentRow"]:
        if self.file_type == "XLSX":
            return self.parse_excel()
        return self.parse_csv()

    def parse_excel(self) -> list["AssessmentRow"]:
        workbook = load_workbook(
            BytesIO(self.file_content), read_only=True, data_only=True
        )
        worksheet = workbook.worksheets[0]

        def clean_excel_cell(cell_value: str | float | datetime | None) -> str:
            return str(cell_value).replace("_x000D", "")

        first_row = next(worksheet.iter_rows(max_row=1, values_only=True))
        header = [clean_excel_cell(cell) if cell else None for cell in first_row]
        rows: list[AssessmentRow] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            r = {}
            for name, cell in zip(header, row):  # noqa: B905 (TODO: strict?)
                if name and cell:
                    r[name] = clean_excel_cell(cell)
            if r:
                rows.append(AssessmentRow.from_flat(r))
        return rows

    def parse_csv(self) -> list["AssessmentRow"]:
        reader = csv.DictReader(StringIO(self.file_content.decode()))
        rows = [AssessmentRow.from_flat(row) for row in reader]
        return rows

    def set_progress(self, message: str, progress: int) -> None:
        self.progress_queue.put_nowait(progress)

    def delete_existing_assessments(self) -> None:
        Assessment.objects.all().delete()

    def home_page(self, locale: Locale) -> HomePage:
        try:
            return HomePage.objects.get(locale=locale)
        except ObjectDoesNotExist:
            raise ImportAssessmentException(
                f"The HomePage for the locale {locale} does not exist. Please add "
                f"it before importing an assessment for {locale}"
            )

    def create_shadow_assessment_from_row(
        self, row: "AssessmentRow", row_num: int
    ) -> None:
        locale = self.locale_from_language_code(row.locale)
        if not (assessment := self.shadow_assessments.get((row.slug, locale))):
            assessment = self.shadow_assessments[(row.slug, locale)] = ShadowAssessment(
                row_num=row_num,
                title=row.title,
                slug=row.slug,
                version=row.version,
                locale=locale,
                high_result_page=row.high_result_page,
                medium_result_page=row.medium_result_page,
                low_result_page=row.low_result_page,
                high_inflection=float(row.high_inflection),
                medium_inflection=float(row.medium_inflection),
                generic_error=row.generic_error,
                tags=row.tags,
            )

        answers = [
            ShadowAnswerBlock(answer=answer, score=score, semantic_id=semantic_id)
            for (answer, score, semantic_id) in zip(
                row.answers, row.scores, row.semantic_ids, strict=False
            )
        ]
        question = ShadowQuestionBlock(
            question=row.question,
            error=row.error,
            min=row.min,
            max=row.max,
            answers=answers,
            type=row.question_type,
            explainer=row.explainer,
        )
        assessment.questions.append(question)


# Since wagtail page models are so difficult to create and modify programatically,
# we instead store all the pages as these shadow objects, which we can then at the end
# do a single write to the database from, and encapsulate all that complexity
@dataclass(slots=True)
class ShadowAnswerBlock:
    answer: str
    score: float
    semantic_id: str


@dataclass(slots=True)
class ShadowQuestionBlock:
    question: str
    answers: list[ShadowAnswerBlock]
    explainer: str = ""
    error: str = ""
    min: str = ""
    max: str = ""
    type: str = ""


@dataclass(slots=True)
class ShadowAssessment:
    row_num: int
    title: str
    slug: str
    version: str
    locale: Locale
    high_result_page: str
    medium_result_page: str
    low_result_page: str
    high_inflection: float
    medium_inflection: float
    generic_error: str
    questions: list[ShadowQuestionBlock] = field(default_factory=list)
    tags: list[str] = field(default_factory=list)

    def save(self, parent: Page) -> None:
        try:
            assessment = Assessment.objects.get(slug=self.slug, locale=self.locale)
        except Assessment.DoesNotExist:
            assessment = Assessment(slug=self.slug, locale=self.locale)
        self.add_field_values_to_assessment(assessment)
        self.add_tags_to_assessment(assessment)
        try:
            with contextlib.suppress(NodeAlreadySaved):
                parent.add_child(instance=assessment)
            assessment.save_revision().publish()
        except ValidationError as err:
            raise ImportAssessmentException(f"Validation error: {err}", self.row_num)

    def add_field_values_to_assessment(self, assessment: Assessment) -> None:
        assessment.slug = self.slug
        assessment.title = self.title
        assessment.version = self.version
        assessment.locale = self.locale
        assessment.high_inflection = self.high_inflection
        assessment.high_result_page_id = get_content_page_id_from_slug(
            self.high_result_page
        )
        assessment.medium_inflection = self.medium_inflection
        assessment.medium_result_page_id = get_content_page_id_from_slug(
            self.medium_result_page
        )
        assessment.low_result_page_id = get_content_page_id_from_slug(
            self.low_result_page
        )
        assessment.generic_error = self.generic_error
        assessment.questions = self.questions_as_streamfield

    def add_tags_to_assessment(self, assessment: Assessment) -> None:
        assessment.tags.clear()
        for tag_name in self.tags:
            tag, _ = Tag.objects.get_or_create(name=tag_name)
            assessment.tags.add(tag)

    @property
    def questions_as_streamfield(self) -> list[dict[str, Any]]:
        """
        Formats the questions in the wagtail streamfield formatting
        """
        stream_data = []
        for question in self.questions:
            answers = [
                {
                    "answer": answer.answer,
                    "score": answer.score,
                    "semantic_id": answer.semantic_id,
                }
                for answer in question.answers
            ]
            stream_data.append(
                {
                    "type": question.type,
                    "value": {
                        "question": question.question,
                        "answers": answers,
                        "explainer": question.explainer,
                        "error": question.error,
                        "min": question.min,
                        "max": question.max,
                    },
                }
            )
        return stream_data


@dataclass(slots=True, frozen=True)
class AssessmentRow:
    """
    Represents a single, deserialised row from an import file
    """

    slug: str
    title: str = ""
    version: str = ""
    tags: list[str] = field(default_factory=list)
    question_type: str = ""
    locale: str = ""
    high_result_page: str = ""
    high_inflection: str = ""
    medium_result_page: str = ""
    medium_inflection: str = ""
    low_result_page: str = ""
    generic_error: str = ""
    question: str = ""
    explainer: str = ""
    error: str = ""
    min: str = ""
    max: str = ""
    answers: list[str] = field(default_factory=list)
    scores: list[float] = field(default_factory=list)
    semantic_ids: list[str] = field(default_factory=list)

    @classmethod
    def fields(cls) -> list[str]:
        return [field.name for field in fields(cls)]

    @classmethod
    def from_flat(cls, row: dict[str, str]) -> "AssessmentRow":
        """
        Creates an AssessmentRow instance from a row in the import, deserialising the
        fields.
        """
        row = {
            key.strip(): value.strip()
            for key, value in row.items()
            if value and key.strip() in cls.fields()
        }
        try:
            return cls(
                tags=deserialise_list(row.pop("tags", "")),
                answers=deserialise_list(row.pop("answers", "")),
                scores=[float(i) for i in deserialise_list(row.pop("scores", ""))],
                semantic_ids=deserialise_list(row.pop("semantic_ids", "")),
                **row,
            )
        except TypeError:
            raise ImportAssessmentException(
                "The import file is missing some required fields."
            )


def get_content_page_id_from_slug(slug: str) -> int:
    try:
        page = ContentPage.objects.get(slug=slug)
    except ObjectDoesNotExist:
        raise ImportAssessmentException(
            f"You are trying to add an assessment, where one of the result pages "
            f"references the content page with slug {slug} which does not exist. "
            "Please create the content page first."
        )
    return page.id


def deserialise_list(value: str) -> list[str]:
    """
    Takes a comma separated value serialised by the CSV library, and returns it as a
    deserialised list
    """
    items = next(csv.reader([value]))
    return [item.strip() for item in items]
